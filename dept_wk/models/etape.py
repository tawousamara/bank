from odoo import models, fields, api, _
import base64
import datetime
from io import BytesIO
import numpy as np
import matplotlib
import openpyxl
matplotlib.use('Agg')
from matplotlib import pyplot as plt
from odoo.exceptions import ValidationError, UserError
import magic
import xlrd
import xlsxwriter
#import pandas as pd


List_items = ['هل العميل شخص مقرب سياسيا؟',
              'هل أحد الشركاء/المساهمين/مسير مقرب سياسيا؟',
              'هل العميل أو أحد الشركاء/المساهمين/مسير مقرب من البنك؟',
              'هل للعميل شركات زميلة / مجموعة؟',
              'المتعامل / أحد الشركاء مدرج ضمن القوائم السوداء',
              'المتعامل / أحد الشركاء مدرج ضمن قائمة الزبائن المتعثرين بمركزية المخاطر لبنك الجزائر']

List_risque = [
    'المباشرة قصيرة الأجل',
    'المباشرة متوسطة الأجل',
    'الغير المباشرة',
    'الاجمالي'
]
list_mouvement = [
    'السنة',
    'الإيداعات (1)',
    'الإيرادات (2)',
    '(1)/(2)',
    'الربحية',
]
List_position = [
    'الوضعية الجبائية',
    'الوضعية الشبه جبائية',
    'الوضعية الشبه جبائية لغير الاجراء'
]
list_fisc = [
    'السنة',
    'حقوق الملكية',
    'مجموع الديون',
    'نسبة المديونية leverage',
     'نسبة الالتزامات تجاه البنوك /Gearing',
    'رقم الاعمال',
    'EBIDTA',
    'صافي الربح',
    'راس المال العامل',
    'احتياجات راس المال العامل'
]

List_Bilan = [
    'حقوق الملكية',
    'رأس المال',
    'نتائج متراكمة',
    'مجموع المطلوبات',
    'التزامات بنكية قصيرة الأجل',
    'التزامات بنكية متوسطة الأجل',
    'تسهيلات الموردين',
    'مستحقات ضرائب',
    'مطلوبات أخرى متداولة',
    'Leverage',
    'مجموع الميزانية',
    'رقم الأعمال',
    'EBITDA',
    'صافي الأرباح',
    'صافي الأرباح/المبيعات',
    'قدرة التمويل الذاتي CAF',
    'صافي رأس المال العامل',
    'احتياجات رأس المال العامل',
    'نسبة التداول (السيولة)',
    'نسبة السيولة السريعة',
    'حقوق عند الزبائن',
    'المخزون',
    'متوسط دوران المخزون (يوم)',
    'متوسط فترة التحصيل (يوم)',
    'متوسط مدة تسهيلات الموردين (يوم)'
]
list_recap = [
    'فترة التحصيل بالأيام',
    'فترة دوران المخزون',
    'مدة تسهيلات الموردين',
    'فترة دوران رأس المال العامل',
    'المبلغ المستحق لتسهيلات قصيرة الأمد',
]
list_var = [
    'المبيعات',
    'كلفة المبيعات',
    'الذمم المدينة',
    'المخزون',
    'الذمم الدائنة',
]
list_garantie = [
    'وجود التأمين على العقارات والضمانات و صلاحيتها',
    'التعهد بتحويل الإيجارات في الحساب / توطين الصفقات في الحساب',
    'تقديم الحسابات المدققة للسنة الماضية في الآجال (خلال 6 أشهر)',
    'تغطية الضمانات تفوق 120%']
list_garantie_fisc = [
    'أقل مستوى لرأس المال',
    'خطاب التنازل عن حقوق سابقة',
    'هامش ضمان الجدية',
    'خطاب دمج الحسابات'
]
list_autre_term = [
    'رهن الحصص والاسهم',
    'رهن حسابات جارية/لأجل'
]
list_poste = [
    'الاطارات',
    'التقنيين',
    'التنفيذ'
]

list_siege = [
    'المقر الاجتماعي',
    'المقرات الثانوية 01',
    'المقرات الثانوية 02',
    'المقرات الثانوية 03'
]

list_situation = [
    'السنة',
    'حقوق الملكية',
    'مجموع الميزانية',
    'رقم الأعمال',
    'صافي الارباح'
]

list_bilan = [
    ('1', 'حقوق الملكية'),
    ('1', 'رأس المال'),
    ('1', 'الاحتياطات'),
    ('1', 'الارباح المتراكمة (محتجزة+محققة)'),
    ('1', 'حقوق الملكية / مجموع الميزانية'),
    ('1', 'صافي رأس المال العامل'),
    ('1', 'احتياجات رأس المال العامل'),
    ('1', 'FR/BFR'),
    ('2', 'مجموع المطلوبات (الديون)'),
    ('2', 'التزامات بنكية'),
    ('2', 'تسهيلات الموردين'),
    ('2', 'ضرائب مستحقة غير مدفوعة'),
    ('2', 'مطلوبات أخرى متداولة'),
    ('2', 'نسبة المديونية Leverage'),
    ('2', 'الالتزامات تجاه البنوك / حقوق'),
    ('3', 'مجموع الميزانية'),
    ('3', '(المبيعات، الايرادات)'),
    ('3', 'EBITDA'),
    ('3', 'صافي الارباح'),
    ('3', 'صافي الارباح/المبيعات ROS'),
    ('3', 'معدل العائد على الموجودات ROA'),
    ('3', 'معدل العائد على حقوق الملكية ROE'),
    ('4', 'التدفقات النقدية التشغيلية'),
    ('4', 'نسبة التداول (السيولة)'),
    ('4', 'نسبة السيولة السريعة'),
    ('5', 'حقوق عند الزبائن'),
    ('5', 'المخزون'),
    ('5', 'متوسط دوران المخزون (يوم)'),
    ('5', 'متوسط فترة التحصيل (يوم)'),
    ('5', 'متوسط مدة تسهيلات الموردين  (يوم)'),
]

tcr_list = [
    (1, 'Chiffre d`affaire', 'رقم الأعمال - المبيعات'),
    (2, 'Revente en l`état', 'إعادة البيع على الحالة'),
    (3, 'Production vendue', 'الإنتاج المثبت'),
    (4, 'Travaux', 'الاشغال'),
    (5, 'Service', 'خدمات'),
    (6, 'Achats consommés', 'مشتريات مستهلكة'),
    (7, 'Autres charges externes', 'خدمات خارجية ومشتريات أخرى'),
    (8, "Valeur ajoutée d'exploitation", 'القيمة المضافة للاستغلال'),
    (9, "Charges de personnel", 'أعباء المستخدمين'),
    (10, "Impôts, taxes et versements assimilés", 'الضرائب والرسوم والمدفوعات المماثلة'),
    (11, "Excédent Brut d'Exploitation", 'إجمالي فائض الاستغلال'),
    (12, "Autres produits opérationnels", 'المنتجات العملياتية الأخرى'),
    (13, "Autres charges opérationnelles", 'الأعباء العملياتية الأخرى'),
    (14, "Dotations aux amortissements", 'مخصصات الاستهلاك ،المؤونات وخسائر القيمة'),
    (15, "Résultat Opérationnel", 'النتيجة العملياتية'),
    (16, "Charges financières", 'الأعباء المالية'),
    (17, "Produit financier", 'المنتجات المالية'),
    (18, "Résultat Ordinaire Avant Impôts", 'النتيجة العادية قبل الضرائب'),
    (19, "Impôts sur les bénéfices", 'الضرائب الواجب دفعها على النتائج العادية'),
    (20, "Résultat Net", 'النتيجة الصافية للنشاطات العادية'),
]

actif_list = [
    (1, 'Immobilisations Incorporelles', 'اصول معنوية'),
    (2, 'Immobilisations Corporelles', 'اصول عينية'),
    (3, 'Terrains', 'أراضي'),
    (4, 'Bâtiments', 'مباني'),
    (5, 'Autres immobilisations corporelles', 'اصول عينية أخرى'),
    (6, 'Immobilisations en concession', 'اصول ممنوح امتيازها'),
    (7, 'Immobilisations encours', 'اصول يجري إنجازها'),
    (8, 'Immobilisations financières', 'اصول مالية'),
    (9, 'Titres mis en équivalence', 'سندات موضوعة موضع معادلة'),
    (10, 'Autres participations et créances rattachées', 'مساهمات أخرى وحسابات دائنة ملحقة بها'),
    (11, 'Autres titres immobilisés', 'سندات أخرى مثبتة'),
    (12, 'Prêts et autres actifs financiers non courants', 'قروض وأصول مالية أخرى غير جارية'),
    (13, 'Impôts différés actif', 'ضرائب مؤجلة على الأصل'),
    (14, 'TOTAL ACTIF NON COURANT', 'مجــموع الأصـول غيـر الجـارية'),
    (15, 'Stocks et encours', 'مخزونات ومنتجات قيد التنفيذ'),
    (16, 'Créances et emplois assimilés', 'حسابات دائنة واستخدامات مماثلة'),
    (17, 'Clients', 'الزبائن'),
    (18, 'Autres débiteurs', 'المدينون الآخرون'),
    (19, 'Impôts et assimilés', 'الضرائب وما شابهها'),
    (20, 'Autres créances et emplois assimilés', 'حسابات دائنة أخرى واستخدامات مماثلة'),
    (21, 'Disponibilités et assimilés', 'الأموال الموظفة والأصول المالية الأخرى'),
    (22, 'Trésorerie', 'الخزينة'),
    (23, 'TOTAL ACTIF COURANT', 'مجــموع الأصــول الجـارية'),
    (24, 'Total Actif', 'المجمــوع العـام للأصــول')
]

passif_list = [
    (1, 'Capital émis', 'رأس مال تم إصداره'),
    (2, 'Capital non appelé', 'رأس مال غير مستعان به'),
    (3, 'Primes et réserves -  Réserves consolidés', 'علاوات واحتياطات – احتياطات مدمجة'),
    (4, 'Ecart de réévaluation', 'فوارق إعادة التقييم'),
    (5, "Ecart d'équivalence", 'فارق المعادلة'),
    (6, 'Résultat net - Résultat net du groupe', 'نتيجة صافية /  نتيجة صافية حصة المجمع'),
    (7, 'Autres capitaux propores - Report à nouveau', 'رؤوس أموال خاصة أخرى / ترحيل من جديد'),
    (8, 'TOTAL CAPITAUX PROPRES (I)', 'إجمالي حقوق الملكية'),
    (9, 'Emprunts et dettes financières', 'قروض وديون مالية'),
    (10, 'Impôts (différés et provisionnés)', 'ضرائب (مؤجلة ومرصود لها)'),
    (11, 'Autres dettes non courantes', 'ديون أخرى غير جارية'),
    (12, "Provisions et produits constatés d'avance", 'مؤونات ومنتجات ثابتة مسبقا'),
    (13, "TOTAL (II)", 'مجموع الخصوم غير الجـارية II'),
    (14, "Fournisseurs et comptes rattachés", 'مـوردون وحسابات ملحـقة'),
    (15, "Impôts", 'ضرائـب'),
    (16, "Autres dettes", 'ديـون أخرى'),
    (17, "Trésorerie passif", 'خزيـنة الخصوم'),
    (18, "TOTAL PASSIFS NON COURANT (III)", 'مجـــموع الخــصوم الجـارية III'),
    (19, "Total Passif", 'المجمــوع العـام للخــصوم'),
]

LIST = [('1', 'طلب التسهيلات ممضي من طرف المفوض القانوني عن الشركة'),
          ('2', 'الميزانيات لثلاث سنوات السابقة مصادق عليها من طرف المدقق المحاس'),
          ('3',
           ' الميزانية الافتتاحية و الميزانية المتوقعة للسنة المراد تمويلها موقعة من طرف الشركة (حديثة النشأة)'),
          ('4', 'مخطط تمويل الاستغلال مقسم الى أرباع السنة للسنة المراد تمويلها'),
          ('5',
           ' المستندات و الوثائق المتعلقة بنشاط الشركة ( عقود، صفقات ،  طلبيات ، ... )'),
          ('6', 'محاضر الجمعيات العادية و الغير العادية للأشخاص المعنويين'),
          ('7', 'نسخة مصادق عليها من السجل التجاري'),
          ('8', 'نسخة مصادق عليها من القانون الأساسي للشركة'),
          ('9', 'مداولة الشركاء أو مجلس الإدارة لتفويض المسير لطلب القروض البنكية'),
          ('10', 'نسخة مصادق عليها من النشرة الرسمية للإعلانات القانونية'),
          ('11', 'نسخة طبق الأصل لعقد ملكية أو استئجار المحلات ذات الاستعمال المهني'),
          ('12',
           ' نسخة طبق الأصل للشهادات الضريبية و شبه الضريبية حديثة (أقل من ثلاثة أشهر)'),
          ('13', 'استمارة كشف مركزية المخاطر ممضية من طرف ممثل الشركة (نموذج مرفق)'),
          ('14', 'آخر تقرير مدقق الحسابات'),
          ('15', 'Actif, Passif, TCR (N, N-1)'),
          ('16', 'Actif, Passif, TCR (N-2, N-3)')
          ]

class Etape(models.Model):
    _name = 'wk.etape'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    workflow = fields.Many2one('wk.workflow.dashboard', ondelete="cascade")
    etape = fields.Many2one('wk.state.principal', string='Etape')
    name = fields.Char(string='Nom', related='etape.name', store=True)
    sequence = fields.Integer(string='Sequence', related='etape.sequence')

    assigned_to_agence = fields.Many2one('res.users', string='المكلف بالملف',
                                         default=lambda self: self.env.user,
                                         domain=lambda self: [('groups_id', 'in', self.env.ref('dept_wk.dept_wk_group_agent_agence').id)])
    state = fields.Selection(related='workflow.state', store=True)
    state_branch = fields.Selection([('branch_1', 'الفرع'),
                                     ('branch_2', 'مدير الفرع'),
                                     ('branch_3', ' الفرع'),
                                     ('branch_4', 'مدير الفرع'),
                                     ('branch_5', 'انتهاء التحليل'),
                                     ('branch_rejected', 'طلب مرفوض'),
                                     ], track_visibility='always', string='وضعية الملف')
    # fields of branch
    nom_client = fields.Many2one('res.partner', string='اسم المتعامل', domain=lambda self: [('branche', '=', self.env.user.partner_id.branche.id), ('is_client', '=', True)], related='workflow.nom_client')
    branche = fields.Many2one('wk.agence', string='الفرع', related='nom_client.branche')
    num_compte = fields.Char(string='رقم الحساب', related='nom_client.num_compte', store=True)
    demande = fields.Many2one('wk.type.demande', string='الطلب', related='workflow.demande')
    classification = fields.Many2one('wk.classification', string="تصنيف الشركة", related='nom_client.classification')
    adress_siege = fields.Char(string='عنوان المقر الاجتماعي', related='nom_client.adress_siege')
    wilaya = fields.Many2one('wk.wilaya', string='الولاية', related='nom_client.wilaya')
    nif = fields.Char(string='NIF', related='nom_client.nif')
    num_registre_commerce = fields.Char(string='رقم السجل التجاري', related='nom_client.rc')
    date_ouverture_compte = fields.Date(string='تاريخ فتح الحساب', related='nom_client.date_ouverture_compte')
    date_inscription = fields.Date(string='تاريخ القيد في السجل التجاري', related='nom_client.date_inscription')
    date_debut_activite = fields.Date(string='تاريخ بداية النشاط', related='nom_client.date_debut_activite')
    activite = fields.Many2one('wk.activite', string='النشاط الرئيسي حسب بنك الجزائر', related='nom_client.activite')
    activite_second = fields.Many2one('wk.secteur', string='النشاط الثانوي حسب السجل التجاري', related='nom_client.activite_second')
    activite_sec = fields.Char( string='النشاط الثانوي حسب السجل التجاري', related='nom_client.activite_sec')
    activity_code = fields.Char(string='رمز النشاط حسب السجل التجاري', related='nom_client.activity_code')
    activity_description = fields.Char(string='النشاط حسب السجل التجاري', related='nom_client.activity_description')
    phone = fields.Char(string='الهاتف', related='nom_client.mobile')
    email = fields.Char(string='البريد الإلكتروني', related='nom_client.email')
    siteweb = fields.Char(string='الموقع الالكتروني للشركة', related='nom_client.website')
    gerant = fields.Many2one('res.partner', string='المسير',
                             domain="[('parent_id', '=', nom_client),('is_company', '=', False)]")
    partner_id = fields.Many2one('res.partner', string='المسير',related='gerant', store=True)
    phone_gerant = fields.Char(string='الهاتف', related='gerant.mobile')
    email_gerant = fields.Char(string='البريد الإلكتروني', related='gerant.email')
    email_to = fields.Char(string='البريد الإلكتروني', store=True)
    email_from = fields.Char(string='البريد الإلكتروني', related='user_id.partner_id.email', store=True)
    author_id = fields.Many2one('res.partner', related='user_id.partner_id', store=True)

    unit_prod = fields.Html(string='وحدات الانتاج')
    stock = fields.Html(string='المخازن')
    prod_company = fields.Html(string='منتوجات الشركة')
    prod_company_related = fields.Html(string='منتوجات الشركة', related='prod_company')
    politique_comm = fields.Html(string='السياسة التسويقية')
    cycle_exploit = fields.Html(string='دورة الاستغلال')
    concurrence = fields.Html(string='المنافسة و دراسة السوق')
    program_invest = fields.Html(string='البرنامج الاستثماري /المشاريع التطويرية')
    result_visit = fields.Html(string='نتائج الزيارة')
    description_company = fields.Html(string='تعريف الشركة')
    history_relation = fields.Html(string='تاريخ العلاقة وحالة التحصيل')
    recommendation_visit = fields.Html(string='توصية الفرع',)
    recommendation_responsable_agence = fields.Html(string='توصية مدير الفرع', )
    images = fields.One2many('wk.documents', 'etape_id', string='الصور المرفقة')

    company_id = fields.Many2one('res.company', default=lambda self: self.env.company)
    forme_jur = fields.Many2one('wk.forme.jur', string='الشكل القانوني', related='nom_client.forme_jur')
    currency_id = fields.Many2one('res.currency', related='company_id.currency_id')
    chiffre_affaire = fields.Monetary(string='راس المال الشركة KDA', currency_field='currency_id',
                                      related='nom_client.chiffre_affaire')

    doc_checked = fields.Boolean(string="أؤكد المستندات")
    doc_checked_vis = fields.Boolean(string="أؤكد المستندات", related='doc_checked')
    documents = fields.One2many('wk.document.check', 'etape_id', string='التاكد من الوثائق المرفقة')

    kyc = fields.One2many('wk.kyc.details', 'etape_id')
    apropos = fields.One2many('wk.partenaire', 'etape_id', string='نبذة عن المتعامل')
    gestion = fields.One2many('wk.gestion', 'etape_id', string='فريق التسيير')
    employees = fields.One2many('wk.nombre.employee', 'etape_id', string='عدد العمال (حسب الفئة المهنية)')
    sieges = fields.One2many('wk.siege', 'etape_id', string='مقرات تابعة للشركة')
    tailles = fields.One2many('wk.taille', 'etape_id', string='حجم و هيكل التمويلات المطلوبة')
    situations = fields.One2many('wk.situation', 'etape_id', string='التمويل لدى البنوك الاخرى')
    situations_fin = fields.One2many('wk.situation.fin', 'etape_id',
                                     string='البيانات المالية المدققة للثلاث سنوات الأخيرة KDA')

    fournisseur = fields.One2many('wk.fournisseur', 'etape_id', string='الموردين')
    client = fields.One2many('wk.client', 'etape_id', string='الزبائن')

    # Financial fields
    state_finance = fields.Selection([
                                    ('finance_1', 'مدير التمويلات'),
                                    ('finance_2', 'المحلل المالي'),
                                    ('finance_3', 'مدير التمويلات'),
                                    ('finance_5', 'في انتظار مديرية الاعمال التجارية و المخاطر'),
                                    ('finance_6', 'في انتظار مديرية الاعمال التجارية'),
                                    ('finance_7', 'في انتظار ادارة المخاطر'),
                                    ('finance_4', 'انتهاء التحليل'),
                                    ('finance_rejected', 'طلب مرفوض'),
                                    ], track_visibility='always', string='وضعية الملف')

    assigned_to_finance = fields.Many2one('res.users', string='المحلل المالي',
                                domain=lambda self: [('groups_id', 'in', self.env.ref('dept_wk.dept_wk_group_analyste').id)],
                                track_visibility='always')

    taux_change = fields.Float(string='1$ = ?DA: سعر الصرف', default=1)
    annee_fiscal = fields.Integer(string='السنة المالية N', compute='change_annee',)
    annee_fiscal_list = fields.Many2one('wk.year', string='السنة المالية N',)
    facilite_accorde = fields.One2many('wk.facilite.accorde', 'etape_id',
                                       string='تفاصيل التسهيلات الممنوحة (KDA)')
    detail_garantie_actuel_ids = fields.One2many('wk.detail.garantie', 'etape_id', string='الضمانات العقارية الحالية')
    garantie_actuel_comment = fields.Html(string='تعليق')
    detail_garantie_propose_ids = fields.One2many('wk.detail.garantie.propose', 'etape_id',
                                                  string='الضمانات العقارية المقترحة')
    garantie_propose_comment = fields.Html(string='تعليق')
    garantie_conf = fields.One2many('wk.garantie.conf', 'etape_id',
                                    string='الشروط السابقة/المقترحة و الموافق عليها من لجان التمويل')
    garantie_fin = fields.One2many('wk.garantie.fin', 'etape_id', string='الشروط المالية')
    garantie_autres = fields.One2many('wk.garantie.autres', 'etape_id', string='الشروط الاخرى')
    risque_central = fields.One2many('wk.risque.line', 'etape_id', string='مركزية المخاطر')
    compute_risque = fields.Float(string='compute field', compute='compute_risk')
    risque_date = fields.Date(string='مركزية المخاطر بتاريخ')
    nbr_banque = fields.Integer(string='عدد البنوك المصرحة')
    comment_risk_central = fields.Html(string='تعليق')
    capture_filename = fields.Char(default='ملف مركزية المخاطر')
    risk_capture = fields.Binary(string='ملف مركزية المخاطر')
    position_tax = fields.One2many('wk.position', 'etape_id', string='الوضعية الجبائية والشبه جبائية')
    mouvement = fields.One2many('wk.mouvement', 'etape_id',
                                string='الحركة والأعمال الجانبية للحساب مع مصرف السلام الجزائر (KDA)')
    detail_mouvement = fields.Html(string='التوطين البنكي')
    computed_field = fields.Boolean(compute='compute_ratio', store=True)

    @api.depends('mouvement')
    def compute_ratio(self):
        for rec in self:
            first = rec.mouvement.filtered(lambda l: l.sequence == 1)
            second = rec.mouvement.filtered(lambda l: l.sequence == 2)
            third = rec.mouvement.filtered(lambda l: l.sequence == 3)
            if first and second:
                third.n_dz = first.n_dz / second.n_dz if second.n_dz != 0 else 0
                third.n1_dz = first.n1_dz / second.n1_dz if second.n1_dz != 0 else 0
                third.n2_dz = first.n2_dz / second.n2_dz if second.n2_dz != 0 else 0
                third.n3_dz = first.n3_dz / second.n3_dz if second.n3_dz != 0 else 0
            rec.computed_field = True

    companies = fields.One2many('wk.companies', 'etape_id')
    companies_fisc = fields.One2many('wk.companies.fisc', 'etape_id')
    tcr_group = fields.Many2one('import.ocr.tcr', string='TCR')
    passif_group = fields.Many2one('import.ocr.passif', string='Passif')
    actif_group = fields.Many2one('import.ocr.actif', string='Actif')

    comment_fisc = fields.Html(string='تعليق')
    visualisation2 = fields.Binary(string='visualisation')

    facitlite_existante = fields.One2many('wk.facilite.existante', 'etape_id')
    risque_ids = fields.One2many('risk.scoring', 'etape_id', string='الشركات ذات الصلة',
                                 domain="[('is_initial', '=', False)]")
    mouvement_group = fields.One2many('wk.mouvement.group', 'etape_id',
                                      string='الحركة والأعمال الجانبية للمجموعة مع مصرف السلام الجزائر (KDA)')
    tcr_id = fields.Many2one('import.ocr.tcr', string='TCR')
    passif_id = fields.Many2one('import.ocr.passif', string='Passif')
    actif_id = fields.Many2one('import.ocr.actif', string='Actif')

    tcr1_id = fields.Many2one('import.ocr.tcr', string='TCR')
    passif1_id = fields.Many2one('import.ocr.passif', string='Passif')
    actif1_id = fields.Many2one('import.ocr.actif', string='Actif')

    bilan_id = fields.One2many('wk.bilan', 'etape_id')
    bilan1_id = fields.One2many('wk.bilan.cat1', 'etape_id')
    comment_cat1 = fields.Html(string='تعليق')
    bilan2_id = fields.One2many('wk.bilan.cat2', 'etape_id')
    comment_cat2 = fields.Html(string='تعليق')
    bilan3_id = fields.One2many('wk.bilan.cat3', 'etape_id')
    comment_cat3 = fields.Html(string='تعليق')
    bilan4_id = fields.One2many('wk.bilan.cat4', 'etape_id')
    comment_cat4 = fields.Html(string='تعليق')
    bilan5_id = fields.One2many('wk.bilan.cat5', 'etape_id')
    comment_cat5 = fields.Html(string='تعليق')

    recap_ids = fields.One2many('wk.recap', 'etape_id')
    var_ids = fields.One2many('wk.variable', 'etape_id')

    visualisation1 = fields.Binary(string='visualisation')

    weakness_ids = fields.One2many('wk.swot.weakness', 'etape_id')
    strength_ids = fields.One2many('wk.swot.strength', 'etape_id')
    threat_ids = fields.One2many('wk.swot.threat', 'etape_id')
    opportunitie_ids = fields.One2many('wk.swot.opportunitie', 'etape_id')

    description_prjt_invest = fields.Html(string='وصف المشروع الاستثماري')
    actif_invest = fields.Html(string='الاصول المراد الاستثمار بها')
    pays_prod = fields.Many2many('res.country', string='بلد المنشأ')
    valeur_total = fields.Float(string='القيمة الإجمالية للاستثمار')
    auto_financement = fields.Float(string='التمويل الذاتي')
    financement_demande = fields.Float(string='التمويل المطلوب')
    duree_financement = fields.Integer(string='مدة التمويل')
    invest_id = fields.Many2one('tcr.analysis.import', string='Investissement')
    avis_invest = fields.Html(string='راي المحلل المالي عن التمويل الاستثماري ')
    recommandation_analyste_fin = fields.Html(string='توصية المحلل المالي', )
    facilite_propose = fields.One2many('wk.facilite.propose', 'etape_id', string='التسهيلات المقترحة')
    garantie_ids = fields.Many2many('wk.garanties', string='الضمانات المقترحة')
    garanties = fields.Html(string='الضمانات المقترحة', )
    garanties_demande = fields.Many2many('wk.garanties', 'garantie_demande_rel', string='الضمانات')
    comite = fields.Many2one('wk.comite', string='اللجنة')
    recommandation_dir_fin = fields.Html(string='راي مدير ادارة التمويلات', )
    montant_demande = fields.Float(string='المبلغ المطلوب')
    montant_propose = fields.Float(string='المبلغ المقترح')
    company_ids = fields.Integer(string='', compute="_compute_company_fisc", )
    date_situation_comptable = fields.Date(string='اخر تاريخ للوضعية المحاسبية')
    visualisation_situation = fields.Binary(string='Chart')
    file_tcr = fields.Binary(string='الملف')
    file_tcr_name = fields.Char(string='الملف', default='الوضعية المحاسبية')
    tcr_situation = fields.One2many('wk.tcr', 'etape_id', domain="[('type', '=', 1)]")
    actif_situation = fields.One2many('wk.actif', 'etape_id', domain="[('type', '=', 1)]")
    passif_situation = fields.One2many('wk.passif', 'etape_id', domain="[('type', '=', 1)]")
    file_tcr_estim = fields.Binary(string='الملف')
    file_tcr_name_estim = fields.Char(string='الملف', default='الوضعية التقديرية')
    tcr_situation_estim = fields.One2many('wk.tcr.estim', 'etape_id')
    actif_situation_estim = fields.One2many('wk.actif.estim', 'etape_id')
    passif_situation_estim = fields.One2many('wk.passif.estim', 'etape_id')
    commentaire_situation = fields.Html(string='تعليق')
    exception_ids = fields.Many2many('wk.exception', string='الاستثناءات مع سياسة الائتمان')
    date_final = fields.Date(string='تاريخ قرار لجنة التسهيلات' ,default=fields.Date.today)
    # Commercial fields
    assigned_to_commercial = fields.Many2one('res.users', string='المكلف بالاعمال التجارية',
                                          domain=lambda self: [
                                              ('groups_id', 'in', self.env.ref('dept_wk.dept_wk_group_charge_commercial').id)],
                                          track_visibility='always')

    state_commercial = fields.Selection([('commercial_1', 'مدير الاعمال التجارية'),
                                         ('commercial_2', 'مديرية الاعمال التجارية'),
                                         ('commercial_3', 'مدير الاعمال التجارية'),
                                         ('commercial_4', 'انتهاء التحليل'),
                                         ],  track_visibility='always', string='وضعية الملف')

    visualisation = fields.Binary(string='visualisation')
    analyse_secteur_act = fields.Html(string='تحليل قطاع عمل العميل')
    analyse_concurrence = fields.Html(string='تحليل المنافسة')
    ampleur_benefice = fields.Float(string='حجم الارباح PNB المتوقعة')
    analyse_relation = fields.Html(string='تحليل اهمية العلاقة على المدى المتوسط')
    recommendation_dir_commercial = fields.Html(string='توصية مدير إدارة الاعمال التجارية', )
    recommendation_commercial = fields.Html(string='توصية المكلف بالاعمال التجارية', )

    # Risk fields
    risk_scoring = fields.Many2one('risk.scoring', string='إدارة المخاطر')
    resultat_scoring = fields.Integer(string='التنقيط الاجمالي', related='risk_scoring.resultat_scoring')
    state_risque = fields.Selection([('risque_1', 'مدير المخاطر'),
                                     ('risque_3', 'المكلف بادارة المخاطر'),
                                     ('risque_4', 'مدير المخاطر'),
                                     ('risque_2', 'انتهاء التحليل'),
                                     ], track_visibility='always', string='وضعية الملف')
    assigned_to_risque = fields.Many2one('res.users', string='المكلف بادارة المخاطر',
                                             domain=lambda self: [
                                                 ('groups_id', 'in',
                                                  self.env.ref('dept_wk.dept_wk_group_responsable_credit').id)],
                                             track_visibility='always')
    validator_risque = fields.Many2one('res.users', string='مراجع من طرف',
                                             domain=lambda self: [
                                                 ('groups_id', 'in',
                                                  self.env.ref('dept_wk.dept_wk_group_responsable_risque').id)],
                                             track_visibility='always')
    recommandation_dir_risque = fields.Html(string='توصية مدير إدارة المخاطر',
                                            )

    # Vice president fields
    state_vice = fields.Selection([('vice_1', 'نائب المدير العام'),
                                     ('vice_2', 'انتهاء التحليل'),
                                     ('vice_rejected', 'طلب مرفوض'),
                                     ], track_visibility='always',default='vice_1', string='وضعية الملف')
    assigned_to_vice = fields.Many2one('res.users', string='نائب المدير العام', track_visibility='always')
    recommandation_tresorerie = fields.Html(string='رأي مسؤول الخزينة', track_visibility='always',)
    recommandation_vice_dir_fin = fields.Html(string='توصية/قرار نائب المدير العام', )
    state_dg = fields.Selection([('dg_1', 'المدير العام'),
                                   ('dg_2', 'انتهاء التحليل'),
                                   ('dg_rejected', 'طلب مرفوض')], default='dg_1', track_visibility='always', string='وضعية الملف')
    recommandation_dg = fields.Html(string='توصية/قرار المدير العام',)
    state_tres = fields.Selection([('tres_1', 'رئيس قطاع الخزينة'),
                                   ('tres_2', 'انتهاء التحليل')], track_visibility='always', default='tres_1', string='وضعية الملف')
    # Comite fields
    state_comite = fields.Selection([('comite_1', 'لجنة التسهيلات'),
                                     ('comite_2', 'انتهاء التحليل'),
                                     ('comite_rejected', 'طلب مرفوض'),
                                     ], track_visibility='always', string='وضعية الملف')

    recommandation_fin_comite = fields.Html(string='توصية/قرار لجنة التسهيلات')
    recommandation_comite3 = fields.Html(string='توصية/قرار لجنة التمويلات')
    recommandation_comite4 = fields.Html(string='قرار مجلس الادارة')

    # Extra fields
    raison_a_revoir = fields.Html(string='سبب طلب المراجعة')
    state = fields.Selection(string='حالة الملف', related='workflow.state')
    state_compute = fields.Float(string='Pourcentage', compute='compute_pourcentage_state')
    user_id = fields.Many2one('res.users', string='المكلف بالملف', compute='compute_user')
    reference = fields.Char(string='Réference', related='workflow.name')
    date = fields.Date(string='تاريخ', related='workflow.date')
    tracking_state = fields.Many2one('wk.tracking', compute='_compute_track', store=True)
    dossier_verouiller = fields.Boolean(string='Dossier Verrouiller', track_visibility='always')
    active = fields.Boolean(default=True)
    can_edit = fields.Boolean(string='', compute='compute_readonly')
    can_edit_finance = fields.Boolean(string='', compute='compute_readonly_finance')

    '''template_situation = fields.Binary(string='Template situation comptable', compute='compute_template_xls')
    template_name = fields.Char(default='Template situation comptable.xls')

    def compute_template_xls(self):
        for rec in self:
            data1 = []
            for line in tcr_list:
                data1.append(line[1])
            data2 = []
            for line in actif_list:
                data2.append(line[1])
            data3 = []
            for line in passif_list:
                data3.append(line[1])
            df1 = pd.DataFrame({
                "Poste comptable": data1,
                "Montant": [0] * len(data1)
            })
            df2 = pd.DataFrame({
                "Poste comptable": data2,
                "Montant": [0] * len(data2)
            })
            df3 = pd.DataFrame({
                "Poste comptable": data3,
                "Montant": [0] * len(data3)
            })
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                df1.to_excel(writer, index=False, sheet_name='TCR')
                df2.to_excel(writer, index=False, sheet_name='Actif')
                df3.to_excel(writer, index=False, sheet_name='Passif')
            buffer.seek(0)
            file_base64 = base64.b64encode(buffer.read()).decode('utf-8')
            rec.template_situation = file_base64'''

    company_distribution_file = fields.Binary(string="البيانات")
    
    
    def upload_company_distribution(self):
        print('######### Starting Upload company_distribution #########')
        self.ensure_one()

        try:
            file_content = base64.b64decode(self.company_distribution_file)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            
            self.apropos.unlink()
            count = 0

            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                age = datetime.datetime.strptime(record[1], '%d-%m-%Y').strftime('%Y-%m-%d')
                try:
                    print(f'Processing record {count}: {record}')
                    self.env['wk.partenaire'].create({
                        'etape_id': self.id,
                        'nom_partenaire': record[0],
                        'age': age,
                        'statut_partenaire': record[2],
                        'nationalite': self.env['res.country'].search([('name', '=', record[3])], limit=1).id,
                        'pourcentage': record[4],
                    })
                    print('************************* Creation done *************************')
                except Exception as e:
                    print(f'Error in record {count}: {e}')
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            print(f'Failed to load or process file: {e}')
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

        
    def download_company_distribution(self):
        self.ensure_one()

        # Générer le fichier Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Ajout des en-têtes des colonnes
        headers = ['اسم الشريك/المالك', 'تاريخ التاسيس/الميلاد', 'صفة الشريك', 'الجنسية','نسبة الحصة']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)

        # Ajout des données du champ one2many
        row = 1
        for line in self.apropos:
            worksheet.write(row, 0, line.nom_partenaire)
            worksheet.write(row, 1, line.age.strftime('%d-%m-%Y'))
            worksheet.write(row, 2, line.statut_partenaire)
            worksheet.write(row, 3, line.nationalite.name)
            worksheet.write(row, 4, line.pourcentage)
            row += 1

        # Clôture du fichier Excel
        workbook.close()
        output.seek(0)

        # Encoder le fichier en base64
        result = base64.b64encode(output.read())

        # Obtenir l'URL de base
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')

        # Créer une pièce jointe
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "توزيع راس مال الشركة.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })

        # Préparer l'URL de téléchargement
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'

        # Retourner l'action pour télécharger le fichier
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  
    
    management_team_file1 = fields.Binary(string="البيانات")

    def upload_management_team_file1(self):
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.management_team_file1)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.gestion.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    self.env['wk.gestion'].create({
                        'etape_id': self.id,
                        'name': record[0],
                        'job': record[1],
                        'niveau_etude': record[2],
                        'age': record[3],
                        'experience': record[4],
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

        
    def download_management_team_file1(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['السيد(ة)', 'المهنة', 'المستوى الدراسي', 'السن','الخبرة المهنية']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.gestion:
            worksheet.write(row, 0, line.name)
            worksheet.write(row, 1, line.job)
            worksheet.write(row, 2, line.niveau_etude)
            worksheet.write(row, 3, line.age)
            worksheet.write(row, 4, line.experience)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "فريق التسيير1.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  
        
    
    required_funds = fields.Binary(string="البيانات")
    
    def upload_required_funds(self):
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.required_funds)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.tailles.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    self.env['wk.taille'].create({
                        'etape_id': self.id,
                        'type_demande': self.env['wk.product'].search([('name', '=', record[0])], limit=1).id,
                        'montant': record[1],
                        'raison': record[2],
                        'preg': record[3],
                        'duree': record[4],
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

    def download_required_funds(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['نوع التسهيلات', 'المبلغ المطلوب', 'الغرض من التمويل', 'هامش الجدية','المدة (الايام)']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.tailles:
            worksheet.write(row, 0, line.type_demande.name)
            worksheet.write(row, 1, line.montant)
            worksheet.write(row, 2, line.raison)
            worksheet.write(row, 3, line.preg)
            worksheet.write(row, 4, line.duree)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "حجم و هيكل التمويلات المطلوبة.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  

    headquarter = fields.Binary(string="البيانات")
    
    def upload_headquarter(self):
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.headquarter)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.sieges.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    self.env['wk.siege'].create({
                        'etape_id': self.id,
                        'name': record[0],
                        'adresse': record[1],
                        'nature': self.env['wk.nature.juridique'].search([('name', '=', record[2])], limit=1).id,
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

    def download_headquarter(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['', 'العنوان', 'الطبيعة القانونية']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.sieges:
            worksheet.write(row, 0, line.name)
            worksheet.write(row, 1, line.adresse)
            worksheet.write(row, 2, line.nature.name)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "مقرات تابعة للشركة.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  
    
    financial_situation_file1 = fields.Binary(string="البيانات")
    
    def upload_financial_situation_file1(self):
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.financial_situation_file1)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.situations.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    self.env['wk.situation'].create({
                        'etape_id': self.id,
                        'banque': self.env['wk.banque'].search([('name', '=', record[0])], limit=1).id,
                        'type_fin': self.env['wk.fin.banque'].search([('name', '=', record[1])], limit=1).id,
                        'montant': record[2],
                        'encours': record[3],
                        'garanties': record[4],
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

        
    def download_financial_situation_file1(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['البنك', 'نوع التمويل', 'المبلغ KDA', 'المبلغ المستغل KDA','الضمانات الممنوحة']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.situations:
            worksheet.write(row, 0, line.banque.name)
            worksheet.write(row, 1, line.type_fin.name)
            worksheet.write(row, 2, line.montant)
            worksheet.write(row, 3, line.encours)
            worksheet.write(row, 4, line.garanties)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "الوضعية المالية1.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  
    
    suppliers_clients_file1 = fields.Binary(string="الموردون ")
    suppliers_clients_file2 = fields.Binary(string=" الزبائن")
    
    def upload_suppliers_clients_file1(self):
        print('*********************************************************************')
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.suppliers_clients_file1)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.fournisseur.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    type_payment_ids = self.env['wk.type.payment'].search([('name', 'in', record[2].split(', '))]).ids
                    self.env['wk.fournisseur'].create({
                        'etape_id': self.id,
                        'name': record[0],
                        'country': self.env['res.country'].search([('name', '=', record[1])], limit=1).id,
                        'type_payment': [(6, 0, type_payment_ids)], 
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e: 
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

        
    def download_suppliers_clients_file1(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['الاسم', 'البلد', 'طريقة السداد']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.fournisseur:
            payment_types = ', '.join(line.type_payment.mapped('name'))
            worksheet.write(row, 0, line.name)
            worksheet.write(row, 1, line.country.name)
            worksheet.write(row, 2, payment_types)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "الموردون.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  
    
    def upload_suppliers_clients_file2(self):
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.suppliers_clients_file2)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.client.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    type_payment_ids = self.env['wk.type.payment'].search([('name', 'in', record[2].split(', '))]).ids
                    self.env['wk.client'].create({
                        'etape_id': self.id,
                        'name': record[0],
                        'country': self.env['res.country'].search([('name', '=', record[1])], limit=1).id,
                        'type_payment': type_payment_ids
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

        
    def download_suppliers_clients_file2(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['الاسم', 'البلد', 'طريقة السداد']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.client:
            payment_types = ', '.join(line.type_payment.mapped('name'))
            worksheet.write(row, 0, line.name)
            worksheet.write(row, 1, line.country.name)
            worksheet.write(row, 2, payment_types)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "الزبائن.xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }  
    
    kda = fields.Binary(string="البيانات")
    
    def upload_kda(self):
        self.ensure_one()
        try:
            file_content = base64.b64decode(self.kda)
            wb = openpyxl.load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            self.companies.unlink()
            count = 0
            for record in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    self.env['wk.companies'].create({
                        'etape_id': self.id,
                        'name': record[0],
                        'date_creation': record[1],
                        'activite': self.env['wk.activite'].search([('name', '=', record[2])], limit=1).id,
                        'chiffre_affaire': record[3],
                        'n1_num_affaire': record[4],
                        'n_num_affaire': record[5],
                    })
                except Exception as e:
                    raise UserError('Veuillez verifier le fichier attaché')
        except Exception as e:
            raise UserError('Failed to process the file. Please ensure it is correctly formatted.')

        
    def download_kda(self):
        self.ensure_one()
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        headers = ['الشركة', 'سنة التاسيس', 'النشاط الرئيسي', 'راس المال', 'رقم الاعمال N-1', 'رقم الاعمال N']
        for i, header in enumerate(headers):
            worksheet.write(0, i, header)
        row = 1
        for line in self.companies:
            worksheet.write(row, 0, line.name)
            worksheet.write(row, 1, line.date_creation)
            worksheet.write(row, 2, line.activite.name)
            worksheet.write(row, 3, line.chiffre_affaire)
            worksheet.write(row, 4, line.n1_num_affaire)
            worksheet.write(row, 5, line.n_num_affaire)
            row += 1
        workbook.close()
        output.seek(0)
        result = base64.b64encode(output.read())
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        attachment_obj = self.env['ir.attachment']
        attachment_id = attachment_obj.create({
            'name': "معلومات حول الشركات ذات الصلة (KDA).xlsx",
            'datas': result,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        download_url = '/web/content/' + str(attachment_id.id) + '?download=true'
        return {
            "type": "ir.actions.act_url",
            "url": str(base_url) + str(download_url),
            "target": "new",
        }
    
    
    
    
    @api.depends('risque_ids')
    def _compute_company_fisc(self):
        for rec in self:
            if rec.sequence == 2:
                rec.company_ids = len(rec.risque_ids)
                etape1 = rec.workflow.states.filtered(lambda l:l.sequence == 1)
                scoring = etape1.risk_scoring
                for company in rec.risque_ids:
                    if company != scoring:
                        company.scoring_id = scoring.id
                        company.parent_id = rec.workflow.id
                    else:
                        company.scoring_id = False
                        company.etape_id = False
                        company.is_initial = True
            else:
                rec.company_ids = 0

    @api.depends('risque_ids')
    def _compute_company_fisc(self):
        for rec in self:
            if rec.sequence == 2:
                rec.company_ids = len(rec.risque_ids)
                etape1 = rec.workflow.states.filtered(lambda l:l.sequence == 1)
                scoring = etape1.risk_scoring
                for company in rec.risque_ids:
                    if company != scoring:
                        company.scoring_id = scoring.id
                        company.parent_id = rec.workflow.id
                    else:
                        company.scoring_id = False
                        company.etape_id = False
                        company.is_initial = True
            else:
                rec.company_ids = 0


    @api.model
    def create(self, vals):
        res = super(Etape, self).create(vals)
        if (res.demande == self.env.ref('dept_wk.type_demande_1') or res.demande == self.env.ref('dept_wk.type_demande_2') or res.demande ==self.env.ref('dept_wk.type_demande_3')) and not res.workflow.workflow_old:
            if res.etape.sequence == 1:
                for index, item in LIST:
                    doc = self.env['wk.document.check'].create({'list_document': index,
                                                                'list_doc': item,
                                                                'etape_id': res.id})
                for item in List_items:
                    line = self.env['wk.kyc.details'].create({'info': item, 'etape_id': res.id})
                for item in list_poste:
                    line = self.env['wk.nombre.employee'].create({'name': item,
                                                                  'etape_id': res.id})
                for item in list_siege:
                    line = self.env['wk.siege'].create({'name': item,
                                                        'etape_id': res.id})
                count = 0
                for item in list_situation:
                    line = self.env['wk.situation.fin'].create({'type': item,
                                                                'sequence': count,
                                                                'etape_id': res.id})
                    count += 1
                res.situations_fin.filtered(lambda l: l.sequence == 0).write({
                    'year1': res.annee_fiscal,
                    'year2': res.annee_fiscal - 1,
                    'year3': res.annee_fiscal - 2,
                })
            elif res.etape.sequence == 2:
                for item in list_garantie:
                    line = self.env['wk.garantie.conf'].create({'info': item, 'etape_id': res.id})
                for item in list_garantie_fisc:
                    line = self.env['wk.garantie.fin'].create({'info': item, 'etape_id': res.id})
                for item in list_autre_term:
                    line = self.env['wk.garantie.autres'].create({'info': item, 'etape_id': res.id})
                for item in List_risque:
                    line = self.env['wk.risque.line'].create({'declaration': item, 'etape_id': res.id})
                for item in List_position:
                    line = self.env['wk.position'].create({'name': item, 'etape_id': res.id})
                count = 0
                for item in list_mouvement:
                    line = self.env['wk.mouvement'].create({'mouvement': item,
                                                            'etape_id': res.id,
                                                            'sequence': count})
                    count += 1
                res.mouvement.filtered(lambda l: l.sequence == 0).write({
                    'n_dz': res.annee_fiscal,
                    'n1_dz': res.annee_fiscal - 1,
                    'n2_dz': res.annee_fiscal - 2,
                    'n3_dz': res.annee_fiscal - 3,
                })
                self.env['wk.mouvement.group'].create({'company': 'السنة',
                                                       'etape_id': res.id,
                                                       'n_dz': res.annee_fiscal,
                                                       'n1_dz': res.annee_fiscal - 1,
                                                       'n2_dz': res.annee_fiscal - 2,
                                                       'sequence': 0})
                count = 0
                for item in list_fisc:
                    line = self.env['wk.companies.fisc'].create({'declaration': item,
                                                                 'sequence': count,
                                                                 'etape_id': res.id})
                    count += 1
                res.companies_fisc.filtered(lambda l: l.sequence == 0).write({
                    'year_4': res.annee_fiscal,
                    'year_3': res.annee_fiscal - 1,
                    'year_2': res.annee_fiscal - 2,
                    'year_1': res.annee_fiscal - 3,
                })
                vals = {'etape_id': res.id,
                        'sequence': 0,
                        'declaration': 'السنة',
                        'year_1': res.annee_fiscal - 3,
                        'year_2': res.annee_fiscal - 2,
                        'year_3': res.annee_fiscal - 1,
                        'year_4': res.annee_fiscal,
                        }
                self.env['wk.bilan.cat1'].create(vals)
                self.env['wk.bilan.cat2'].create(vals)
                self.env['wk.bilan.cat3'].create(vals)
                self.env['wk.bilan.cat4'].create(vals)
                self.env['wk.bilan.cat5'].create(vals)
                count = 1
                for index, item in list_bilan:
                    line = self.env['wk.bilan'].create({'declaration': item,
                                                        'categorie': index,
                                                        'etape_id': res.id,
                                                        'sequence': count})
                    count += 1
                count = 1
                for item in list_recap:
                    line = self.env['wk.recap'].create({'declaration': item, 'etape_id': res.id, 'sequence': count})
                    count += 1
                count = 1
                for item in list_var:
                    line = self.env['wk.variable'].create({'var': item, 'etape_id': res.id, 'sequence': count})
                    count += 1
            years = self.env['wk.year'].search([])
            if not years:
                start = datetime.date.today().year - 2
                stop =  datetime.date.today().year + 10
                for i in range(start, stop):
                    self.env['wk.year'].create({'name': str(i)})
        return res

    @api.onchange('annee_fiscal_list')
    def change_annee(self):
        for rec in self:
            rec.annee_fiscal = int(rec.annee_fiscal_list.name)
            if rec.annee_fiscal_list:
                rec.situations_fin.filtered(lambda l: l.sequence == 0).write({
                    'year1': rec.annee_fiscal,
                    'year2': rec.annee_fiscal - 1,
                    'year3': rec.annee_fiscal - 2,
                })
                rec.mouvement.filtered(lambda l: l.sequence == 0).write({
                    'n_dz': rec.annee_fiscal,
                    'n1_dz': rec.annee_fiscal - 1,
                    'n2_dz': rec.annee_fiscal - 2,
                    'n3_dz': rec.annee_fiscal - 3,
                })
                rec.mouvement_group.filtered(lambda l: l.company == 'السنة').write({'n_dz': rec.annee_fiscal,
                                                       'n1_dz': rec.annee_fiscal - 1,
                                                       'n2_dz': rec.annee_fiscal - 2,
                                                       'sequence': 0})
                rec.companies_fisc.filtered(lambda l: l.sequence == 0).write({
                    'year_4': rec.annee_fiscal,
                    'year_3': rec.annee_fiscal - 1,
                    'year_2': rec.annee_fiscal - 2,
                    'year_1': rec.annee_fiscal - 3,
                })
                rec.bilan1_id.filtered(lambda l:l.declaration == 'السنة').write({
                        'year_1': rec.annee_fiscal - 3,
                        'year_2': rec.annee_fiscal - 2,
                        'year_3': rec.annee_fiscal - 1,
                        'year_4': rec.annee_fiscal,})
                rec.bilan2_id.filtered(lambda l:l.declaration == 'السنة').write({
                        'year_1': rec.annee_fiscal - 3,
                        'year_2': rec.annee_fiscal - 2,
                        'year_3': rec.annee_fiscal - 1,
                        'year_4': rec.annee_fiscal})
                rec.bilan3_id.filtered(lambda l:l.declaration == 'السنة').write({
                        'year_1': rec.annee_fiscal - 3,
                        'year_2': rec.annee_fiscal - 2,
                        'year_3': rec.annee_fiscal- 1,
                        'year_4': rec.annee_fiscal})
                rec.bilan4_id.filtered(lambda l:l.declaration == 'السنة').write({
                        'year_1': rec.annee_fiscal- 3,
                        'year_2': rec.annee_fiscal - 2,
                        'year_3': rec.annee_fiscal - 1,
                        'year_4': rec.annee_fiscal})
                rec.bilan5_id.filtered(lambda l:l.declaration == 'السنة').write({
                        'year_1': rec.annee_fiscal - 3,
                        'year_2': rec.annee_fiscal - 2,
                        'year_3': rec.annee_fiscal - 1,
                        'year_4': rec.annee_fiscal })

    def verrouiller_dossier(self):
        for rec in self:
            if self.env.user.has_group('dept_wk.dept_wk_group_analyste'):
                view_id = self.env.ref('dept_wk.confirmation_etape_wizard_form').id
                context = dict(self.env.context or {})
                context['etape'] = rec.id
                context['verrouiller'] = True
                return {
                    'name': 'تأكيد',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'etape.wizard',
                    'view_id': view_id,
                    'target': 'new',
                    'context': context,
                }

    def verrouiller_dossier_function(self):
        for rec in self:
            if rec.state_finance == 'finance_4':
                etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                exist_com = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                exist_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
                if not exist_com:
                    vals = {
                        'workflow': rec.workflow.id,
                        'nom_client': etape_1.nom_client.id,
                        'branche': etape_1.branche.id,
                        'num_compte': etape_1.num_compte,
                        'demande': etape_1.demande.id,
                        'etape': self.env.ref('dept_wk.princip_3').id,
                        'state_commercial': 'commercial_1',
                        'gerant': etape_1.gerant.id,
                        'unit_prod': etape_1.unit_prod,
                        'stock': etape_1.stock,
                        'prod_company': etape_1.prod_company,
                        'politique_comm': etape_1.politique_comm,
                        'cycle_exploit': etape_1.cycle_exploit,
                        'concurrence': etape_1.concurrence,
                        'program_invest': etape_1.program_invest,
                        'result_visit': etape_1.result_visit,
                        'description_company': etape_1.description_company,
                        'recommendation_visit': etape_1.recommendation_visit,
                        'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                    }
                    etape = self.env['wk.etape'].create(vals)
                    for doc in etape_1.documents:
                        self.env['wk.document.check'].create({
                            'list_doc': doc.list_doc,
                            'document': doc.document,
                            'answer': doc.answer,
                            'note': doc.note,
                            'filename': doc.filename,
                            'etape_id': etape.id})
                    for image in etape_1.images:
                        self.env['wk.documents'].create({'picture': image.picture,
                                                         'name': image.name,
                                                         'etape_id': etape.id})
                    for kyc in etape_1.kyc:
                        self.env['wk.kyc.details'].create({'info': kyc.info,
                                                           'answer': kyc.answer,
                                                           'detail': kyc.detail,
                                                           'etape_id': etape.id})
                    for a in etape_1.apropos:
                        self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                          'age': a.age,
                                                          'pourcentage': a.pourcentage,
                                                          'statut_partenaire': a.statut_partenaire,
                                                          'nationalite': a.nationalite.id,
                                                          'etape_id': etape.id
                                                          })
                    for g in etape_1.gestion:
                        self.env['wk.gestion'].create({
                            'name': g.name,
                            'job': g.job,
                            'niveau_etude': g.niveau_etude,
                            'age': g.age,
                            'experience': g.experience,
                            'etape_id': etape.id
                        })
                    for empl in etape_1.employees:
                        self.env['wk.nombre.employee'].create({
                            'name': empl.name,
                            'poste_permanent': empl.poste_permanent,
                            'poste_non_permanent': empl.poste_non_permanent,
                            'etape_id': etape.id
                        })
                    for siege in etape_1.sieges:
                        self.env['wk.siege'].create({
                            'name': siege.name,
                            'adresse': siege.adresse,
                            'nature': siege.nature.id,
                            'etape_id': etape.id
                        })
                    for taille in etape_1.tailles:
                        self.env['wk.taille'].create({
                            'type_demande': taille.type_demande.id,
                            'montant': taille.montant,
                            'raison': taille.raison,
                            'etape_id': etape.id,
                            'garanties': taille.garanties.ids})
                    for sit in etape_1.situations:
                        self.env['wk.situation'].create({
                            'banque': sit.banque.id,
                            'type_fin': sit.type_fin.id,
                            'montant': sit.montant,
                            'garanties': sit.garanties,
                            'etape_id': etape.id
                        })
                    for sit in etape_1.situations_fin:
                        self.env['wk.situation.fin'].create({
                            'type': sit.type,
                            'sequence': sit.sequence,
                            'year1': sit.year1,
                            'year2': sit.year2,
                            'year3': sit.year3,
                            'etape_id': etape.id
                        })
                    for client in etape_1.client:
                        self.env['wk.client'].create({
                            'name': client.name,
                            'country': client.country.id,
                            'type_payment': client.type_payment.ids,
                            'etape_id': etape.id
                        })
                    for f in etape_1.fournisseur:
                        self.env['wk.fournisseur'].create({
                            'name': f.name,
                            'country': f.country.id,
                            'type_payment': f.type_payment.ids,
                            'etape_id': etape.id
                        })
                    '''if not exist_risk:
                        etape = self.env['wk.etape'].create({'workflow': rec.workflow.id,
                                                             'etape': self.env.ref('dept_wk.princip_4').id,
                                                             'risk_scoring': etape_1.risk_scoring.id,
                                                             'state_risque': 'risque_1'})'''
                if exist_com.state_commercial == 'commercial_4':
                    if not exist_risk:
                        exist_risk = self.env['wk.etape'].create({'workflow': rec.workflow.id,
                                                                  'etape': self.env.ref('dept_wk.princip_4').id,
                                                                  'risk_scoring': etape_1.risk_scoring.id,
                                                                  'state_risque': 'risque_1'})
                if exist_risk.state_risque == 'risque_2':
                    etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                    rec.workflow.state = '5'
                    vals = {
                        'workflow': rec.workflow.id,
                        'nom_client': etape_1.nom_client.id,
                        'branche': etape_1.branche.id,
                        'num_compte': etape_1.num_compte,
                        'demande': etape_1.demande.id,
                        'state_vice': 'vice_1',
                        'etape': self.env.ref('dept_wk.princip_5').id,
                        'gerant': etape_1.gerant.id,
                        'unit_prod': etape_1.unit_prod,
                        'stock': etape_1.stock,
                        'prod_company': etape_1.prod_company,
                        'politique_comm': etape_1.politique_comm,
                        'cycle_exploit': etape_1.cycle_exploit,
                        'concurrence': etape_1.concurrence,
                        'program_invest': etape_1.program_invest,
                        'result_visit': etape_1.result_visit,
                        'description_company': etape_1.description_company,
                        'recommendation_visit': etape_1.recommendation_visit,
                        'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                        'analyse_secteur_act': exist_com.analyse_secteur_act,
                        'analyse_concurrence': exist_com.analyse_concurrence,
                        'ampleur_benefice': exist_com.ampleur_benefice,
                        'analyse_relation': exist_com.analyse_relation,
                        'recommendation_dir_commercial': exist_com.recommendation_dir_commercial,
                        'recommendation_commercial': exist_com.recommendation_commercial,
                        'risk_scoring': exist_risk.risk_scoring.id,
                        'recommandation_dir_risque': exist_risk.recommandation_dir_risque,
                        'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                        'garantie_ids': etape_fin.garantie_ids.ids,
                        'exception_ids': etape_fin.exception_ids.ids,
                        'comite': etape_fin.comite.id,
                        'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
                    }
                    if etape:
                        etape.write(vals)
                        etape.documents.unlink()
                        etape.images.unlink()
                        etape.kyc.unlink()
                        etape.apropos.unlink()
                        etape.gestion.unlink()
                        etape.employees.unlink()
                        etape.sieges.unlink()
                        etape.tailles.unlink()
                        etape.situations.unlink()
                        etape.situations_fin.unlink()
                        etape.client.unlink()
                        etape.fournisseur.unlink()
                        etape.facilite_propose.unlink()
                    else:
                        etape = self.env['wk.etape'].create(vals)
                    for doc in etape_1.documents:
                        self.env['wk.document.check'].create({
                            'list_doc': doc.list_doc,
                            'document': doc.document,
                            'answer': doc.answer,
                            'note': doc.note,
                            'filename': doc.filename,
                            'etape_id': etape.id})
                    for image in etape_1.images:
                        self.env['wk.documents'].create({'picture': image.picture,
                                                         'name': image.name,
                                                         'etape_id': etape.id})
                    for kyc in etape_1.kyc:
                        self.env['wk.kyc.details'].create({'info': kyc.info,
                                                           'answer': kyc.answer,
                                                           'detail': kyc.detail,
                                                           'etape_id': etape.id})
                    for a in etape_1.apropos:
                        self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                          'age': a.age,
                                                          'pourcentage': a.pourcentage,
                                                          'statut_partenaire': a.statut_partenaire,
                                                          'nationalite': a.nationalite.id,
                                                          'etape_id': etape.id
                                                          })
                    for g in etape_1.gestion:
                        self.env['wk.gestion'].create({
                            'name': g.name,
                            'job': g.job,
                            'niveau_etude': g.niveau_etude,
                            'age': g.age,
                            'experience': g.experience,
                            'etape_id': etape.id
                        })
                    for empl in etape_1.employees:
                        self.env['wk.nombre.employee'].create({
                            'name': empl.name,
                            'poste_permanent': empl.poste_permanent,
                            'poste_non_permanent': empl.poste_non_permanent,
                            'etape_id': etape.id
                        })
                    for siege in etape_1.sieges:
                        self.env['wk.siege'].create({
                            'name': siege.name,
                            'adresse': siege.adresse,
                            'nature': siege.nature.id,
                            'etape_id': etape.id
                        })
                    for taille in etape_1.tailles:
                        self.env['wk.taille'].create({
                            'type_demande': taille.type_demande.id,
                            'montant': taille.montant,
                            'raison': taille.raison,
                            'etape_id': etape.id,
                            'garanties': taille.garanties.ids})
                    for sit in etape_1.situations:
                        self.env['wk.situation'].create({
                            'banque': sit.banque.id,
                            'type_fin': sit.type_fin.id,
                            'montant': sit.montant,
                            'garanties': sit.garanties,
                            'etape_id': etape.id
                        })
                    for sit in etape_1.situations_fin:
                        self.env['wk.situation.fin'].create({
                            'type': sit.type,
                            'sequence': sit.sequence,
                            'year1': sit.year1,
                            'year2': sit.year2,
                            'year3': sit.year3,
                            'etape_id': etape.id
                        })
                    for client in etape_1.client:
                        self.env['wk.client'].create({
                            'name': client.name,
                            'country': client.country.id,
                            'type_payment': client.type_payment.ids,
                            'etape_id': etape.id
                        })
                    for f in etape_1.fournisseur:
                        self.env['wk.fournisseur'].create({
                            'name': f.name,
                            'country': f.country.id,
                            'type_payment': f.type_payment.ids,
                            'etape_id': etape.id
                        })
                    for fac in etape_fin.facilite_propose:
                        self.env['wk.facilite.propose'].create({
                            'type_facilite': fac.type_facilite.id,
                            'type_demande_ids': fac.type_demande_ids.ids,
                            'montant_dz': fac.montant_dz,
                            'preg': fac.preg,
                            'duree': fac.duree,
                            'condition': fac.condition,
                            'etape_id': etape.id})

    def validate_information(self):
        for rec in self:
                view_id = self.env.ref('dept_wk.confirmation_etape_wizard_form').id
                context = dict(self.env.context or {})
                context['to_validate'] = True
                context['etape'] = rec.id
                return {
                    'name': 'تأكيد',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'etape.wizard',
                    'view_id': view_id,
                    'target': 'new',
                    'context': context,
                }

    def action_situation(self):
        for rec in self:
            if not rec.file_tcr:
                raise UserError(_('Attacher le fichier'))
            elif not check_if_xls_file(rec.file_tcr):
                raise UserError(_('Attacher un fichier excel'))
            else:
                if not rec.tcr_situation and len(rec.tcr_situation) == 0:
                    for index, item1, item2 in tcr_list:
                        self.env['wk.tcr'].create({'name': item1,
                                                   'name_ar': item2,
                                                   'sequence': index,
                                                   'type': 1,
                                                   'etape_id': rec.id})

                    for index, item1, item2 in actif_list:
                        self.env['wk.actif'].create({'name': item1,
                                                   'name_ar': item2,
                                                   'sequence': index,
                                                   'type': 1,
                                                   'etape_id': rec.id})
                    for index, item1, item2 in passif_list:
                        self.env['wk.passif'].create({'name': item1,
                                                   'name_ar': item2,
                                                   'sequence': index,
                                                   'type': 1,
                                                   'etape_id': rec.id})
                xls_data = BytesIO(base64.b64decode(rec.file_tcr))
                workbook = xlrd.open_workbook(file_contents=xls_data.read())
                count = 0
                for sheet_index in range(min(workbook.nsheets, 3)): # Loop through the first 3 sheets
                    sheet = workbook.sheet_by_index(sheet_index)
                    # Read data from the sheet
                    for row_idx in range(sheet.nrows):
                        row_data = sheet.row_values(row_idx)
                        if count == 0:
                            tcr = self.env['wk.tcr'].search([('etape_id', '=', rec.id),
                                                             ('name', '=', row_data[0]),
                                                             ('type', '=', 1)])
                            tcr.valeur = row_data[1]
                        if count == 1:
                            actif = self.env['wk.actif'].search([('etape_id', '=', rec.id),
                                                                 ('name', '=', row_data[0]),
                                                                 ('type', '=', 1)])
                            actif.valeur = row_data[1]
                        if count == 2:
                            passif = self.env['wk.passif'].search([('etape_id', '=', rec.id),
                                                                   ('name', '=', row_data[0]),
                                                                   ('type', '=', 1)])
                            passif.valeur = row_data[1]
                    count += 1
                label1 = 'CA'
                label2 = 'EBE'

                # Récupérer les valeurs de CA et EBE
                data1 = rec.tcr_situation.filtered(lambda l: l.sequence == 1).valeur
                data2 = rec.tcr_situation.filtered(lambda l: l.sequence == 11).valeur

                # Créer le graphique à barres
                fig, ax = plt.subplots()
                width = 0.35  # Largeur des barres
                bar1 = ax.bar(label1, data1, width, label=label1)
                bar2 = ax.bar(label2, data2, width, label=label2)

                # Ajouter des étiquettes et une légende
                ax.set_ylabel('Montant')
                ax.set_title('CA vs EBE')
                ax.legend()

                # Enregistrer le graphique dans un objet BytesIO
                buf = BytesIO()
                plt.savefig(buf, format='jpeg', dpi=100)
                buf.seek(0)

                # Convertir l'image en base64 et l'assigner à votre champ 'visualisation_situation'
                rec.visualisation_situation = base64.b64encode(buf.getvalue()).decode()
                buf.close()

    def action_situation_estimate(self):
        for rec in self:
            if not rec.file_tcr_estim:
                raise UserError(_('Attacher le fichier'))
            elif not check_if_xls_file(rec.file_tcr_estim):
                raise UserError(_('Attacher un fichier excel'))
            else:
                if not rec.tcr_situation_estim:
                    for index, item1, item2 in tcr_list:
                        self.env['wk.tcr.estim'].create({'name': item1,
                                                   'name_ar': item2,
                                                   'sequence': index,
                                                   'type': 2,
                                                   'etape_id': rec.id})

                    for index, item1, item2 in actif_list:
                        self.env['wk.actif.estim'].create({'name': item1,
                                                   'name_ar': item2,
                                                   'sequence': index,
                                                   'type': 2,
                                                   'etape_id': rec.id})
                    for index, item1, item2 in passif_list:
                        self.env['wk.passif.estim'].create({'name': item1,
                                                   'name_ar': item2,
                                                   'sequence': index,
                                                   'type': 2,
                                                   'etape_id': rec.id})
                xls_data = BytesIO(base64.b64decode(rec.file_tcr_estim))
                workbook = xlrd.open_workbook(file_contents=xls_data.read())
                count = 0
                for sheet_index in range(min(workbook.nsheets, 3)): # Loop through the first 3 sheets
                    sheet = workbook.sheet_by_index(sheet_index)
                    # Read data from the sheet
                    for row_idx in range(sheet.nrows):
                        row_data = sheet.row_values(row_idx)
                        if count == 0:
                            tcr = self.env['wk.tcr.estim'].search([('etape_id', '=', rec.id),
                                                                   ('name', '=', row_data[0]),
                                                                   ('type', '=', 2)])
                            tcr.valeur = row_data[1]
                        if count == 1:
                            actif = self.env['wk.actif.estim'].search([('etape_id', '=', rec.id),
                                                                       ('name', '=', row_data[0]),
                                                                       ('type', '=', 2)])
                            actif.valeur = row_data[1]
                        if count == 2:
                            passif = self.env['wk.passif.estim'].search([('etape_id', '=', rec.id),
                                                                         ('name', '=', row_data[0]),
                                                                         ('type', '=', 2)])
                            passif.valeur = row_data[1]
                    count += 1

    def compute_readonly(self):
        for rec in self:
            result = False
            if rec.etape.sequence == 1:
                if rec.state_branch in ['branch_2', 'branch_4'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_agence') and self.env.user.partner_id.branche == rec.branche:
                    result = True
                elif rec.state_branch in ['branch_1', 'branch_3'] and self.env.user == rec.assigned_to_agence and self.env.user.partner_id.branche == rec.branche:
                    result = True
            elif rec.etape.sequence == 2:
                if rec.state_finance in ['finance_1', 'finance_3'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_analyste'):
                    result = True
                elif rec.state_finance == 'finance_2' and self.env.user == rec.assigned_to_finance:
                    result = True
            elif rec.etape.sequence == 3:
                if rec.state_commercial in ['commercial_1', 'commercial_3'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_commercial'):
                    result = True
                elif rec.state_commercial == 'commercial_2' and self.env.user.has_group('dept_wk.dept_wk_group_charge_commercial') and self.env.user == rec.assigned_to_commercial:
                    result = True
            elif rec.etape.sequence == 4:
                if rec.state_risque == 'risque_3' and rec.assigned_to_risque == self.env.user and self.env.user.has_group('dept_wk.dept_wk_group_responsable_credit'):
                    result = True
                elif rec.state_risque in ['risque_1', 'risque_4', 'risque_2'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_risque'):
                    result = True
            elif rec.etape.sequence == 10:
                if self.env.user.has_group('dept_wk.dept_wk_group_tres'):
                    result = True
            elif rec.etape.sequence == 5:
                if self.env.user.has_group('dept_wk.dept_wk_group_dga'):
                    result = True
            elif rec.etape.sequence == 9:
                if self.env.user.has_group('dept_wk.dept_wk_group_dg'):
                    result = True
            elif rec.etape.sequence == 6:
                if self.env.user.has_group('dept_wk.dept_wk_group_comite'):
                    result = True
            elif rec.etape.sequence == 8:
                if self.env.user.has_group('dept_wk.dept_wk_group_analyste'):
                    result = True
            rec.can_edit = result

    def compute_readonly_finance(self):
        for rec in self:
            result = False
            if rec.etape.sequence == 2:
                if rec.state_finance in ['finance_1', 'finance_2'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_analyste'):
                    result = True
            rec.can_edit_finance = result

    def compute_pourcentage_state(self):
        for rec in self:
            years = self.env['wk.year'].search([])
            if not years:
                start = datetime.date.today().year - 2
                stop = datetime.date.today().year + 10
                for i in range(start, stop):
                    self.env['wk.year'].create({'name': str(i)})
            rec.active = True
            partner = rec.workflow.states.filtered(lambda l: l.sequence == 1).gerant
            if partner:
                mail_invite = self.env['mail.wizard.invite'].with_context({
                    'default_res_model': 'wk.etape',
                    'default_res_id': rec.id
                }).with_user(self.env.user).create({
                    'partner_ids': [(4, partner.id)],
                    'notify': False})
                mail_invite.add_followers()
            if rec.etape.sequence == 1:
                rec.workflow.assigned_to_agence = rec.assigned_to_agence.id
                for doc in rec.documents:
                    if doc.list_document:
                        for index, item in LIST:
                            if index == doc.list_document:
                                doc.filename = item
                                doc.list_doc = item
                    else:
                        doc.filename = doc.list_doc
                if rec.state_branch == 'branch_1':
                    rec.state_compute = 0
                    rec.workflow.state = '1'
                elif rec.state_branch == 'branch_2':
                    rec.state_compute = 0.25
                    rec.workflow.state = '1'
                elif rec.state_branch == 'branch_3':
                    rec.state_compute = 0.5
                    rec.workflow.state = '1'
                elif rec.state_branch == 'branch_4':
                    rec.state_compute = 0.75
                    rec.workflow.state = '1'
                else:
                    rec.state_compute = 1
            elif rec.etape.sequence == 2:
                rec.workflow.assigned_to_finance = rec.assigned_to_finance.id
                for doc in rec.documents:
                    if doc.list_document:
                        for index, item in LIST:
                            if index == doc.list_document:
                                doc.filename = item
                                doc.list_doc = item
                    else:
                        doc.filename = doc.list_doc
                if rec.state_finance == 'finance_1':
                    rec.state_compute = 0
                elif rec.state_finance == 'finance_2':
                    rec.state_compute = 0.33
                elif rec.state_finance == 'finance_3':
                    rec.state_compute = 0.66
                elif rec.state_finance in ['finance_4', 'finance_5', 'finance_6', 'finance_7']:
                    rec.state_compute = 1
                else:
                    rec.state_compute = 1
            elif rec.etape.sequence == 3:
                if rec.state_commercial == 'commercial_1':
                    rec.state_compute = 0
                elif rec.state_commercial == 'commercial_2':
                    rec.state_compute = 0.33
                elif rec.state_commercial == 'commercial_3':
                    rec.state_compute = 0.66
                elif rec.state_commercial == 'commercial_4':
                    rec.state_compute = 1
                rec.workflow.state_commercial = rec.state_commercial
            elif rec.etape.sequence == 4:
                if rec.state_risque == 'risque_1':
                    rec.state_compute = 0
                elif rec.state_risque == 'risque_3':
                    rec.state_compute = 0.33
                elif rec.state_risque == 'risque_4':
                    rec.state_compute = 0.66
                elif rec.state_risque == 'risque_2':
                    rec.state_compute = 1
                rec.workflow.state_risque = rec.state_risque
            elif rec.etape.sequence == 10:
                if rec.state_tres == 'tres_1':
                    rec.state_compute = 0
                elif rec.state_tres == 'tres_2':
                    rec.state_compute = 1
                else:
                    rec.state_compute = 0
            elif rec.etape.sequence == 5:
                if rec.state_vice == 'vice_1':
                    rec.state_compute = 0
                elif rec.state_vice == 'vice_2':
                    rec.state_compute = 1
                else:
                    rec.state_compute = 0
            elif rec.etape.sequence == 9:
                if rec.state_dg == 'dg_1':
                    rec.state_compute = 0
                elif rec.state_dg == 'dg_2':
                    rec.state_compute = 1
                else:
                    rec.state_compute = 0
            elif rec.etape.sequence == 6:
                if rec.state_comite == 'comite_1':
                    rec.state_compute = 0
                elif rec.state_comite == 'comite_2':
                    rec.state_compute = 1
                else:
                    rec.state_compute = 0
            elif rec.etape.sequence == 8:
                rec.state_compute = 1

    def compute_user(self):
        for rec in self:
            if rec.etape.sequence == 1:
                rec.user_id = rec.assigned_to_agence.id
            elif rec.etape.sequence == 2:
                rec.user_id = rec.assigned_to_finance.id
            elif rec.etape.sequence == 3:
                rec.user_id = rec.assigned_to_commercial.id
                rec.workflow.assigned_to_commercial = rec.assigned_to_commercial.id
            elif rec.etape.sequence == 4:
                rec.user_id = rec.assigned_to_risque.id
                rec.workflow.assigned_to_risque = rec.assigned_to_risque.id
            elif rec.etape.sequence == 5:
                rec.user_id = rec.assigned_to_vice.id
            else:
                rec.user_id = False
            etape1 = rec.workflow.states.filtered(lambda l: l.sequence == 1)
            montant = 0
            if etape1.tailles:
                montant = sum(etape1.tailles.mapped('montant'))
            rec.montant_demande = montant
            rec.workflow.montant_demande = montant
            etape2 = rec.workflow.states.filtered(lambda l: l.sequence == 2)
            rec.dossier_verouiller = etape2.dossier_verouiller
            if rec.workflow.state == '7':
                last_track = self.env['wk.tracking'].search([('workflow_id', '=', rec.workflow.id)])[-1]
                rec.workflow.date_fin = last_track.date_fin
            if etape2.facilite_propose:
                montant = sum(etape2.facilite_propose.mapped('montant_dz'))
            rec.montant_propose = montant
            if not rec.garantie_conf:
                for item in list_garantie:
                    line = self.env['wk.garantie.conf'].create({'info': item, 'etape_id': rec.id})
            if not rec.garantie_fin:
                for item in list_garantie_fisc:
                    line = self.env['wk.garantie.fin'].create({'info': item, 'etape_id': rec.id})
            if not rec.garantie_autres:
                for item in list_autre_term:
                    line = self.env['wk.garantie.autres'].create({'info': item, 'etape_id': rec.id})
            if not rec.risque_central:
                for item in List_risque:
                    line = self.env['wk.risque.line'].create({'declaration': item, 'etape_id': rec.id})
            if not rec.position_tax:
                for item in List_position:
                    line = self.env['wk.position'].create({'name': item, 'etape_id': rec.id})
            count = 0
            if not rec.mouvement:
                for item in list_mouvement:
                    line = self.env['wk.mouvement'].create({'mouvement': item,
                                                            'etape_id': rec.id,
                                                            'sequence': count})
                    count += 1
                rec.mouvement.filtered(lambda l: l.sequence == 0).write({
                    'n_dz': rec.annee_fiscal,
                    'n1_dz': rec.annee_fiscal - 1,
                    'n2_dz': rec.annee_fiscal - 2,
                    'n3_dz': rec.annee_fiscal - 3,
                })
                self.env['wk.mouvement.group'].create({'company': 'السنة',
                                                       'etape_id': rec.id,
                                                       'n_dz': rec.annee_fiscal,
                                                       'n1_dz': rec.annee_fiscal - 1,
                                                       'n2_dz': rec.annee_fiscal - 2,
                                                       'sequence': 0})
            count = 0
            if not rec.companies_fisc:
                for item in list_fisc:
                    line = self.env['wk.companies.fisc'].create({'declaration': item,
                                                                 'sequence': count,
                                                                 'etape_id': rec.id})
                    count += 1
                rec.companies_fisc.filtered(lambda l: l.sequence == 0).write({
                    'year_4': rec.annee_fiscal,
                    'year_3': rec.annee_fiscal - 1,
                    'year_2': rec.annee_fiscal - 2,
                    'year_1': rec.annee_fiscal - 3,
                })
            if not rec.bilan_id:
                vals = {'etape_id': rec.id,
                        'sequence': 0,
                        'declaration': 'السنة',
                        'year_1': rec.annee_fiscal - 3,
                        'year_2': rec.annee_fiscal - 2,
                        'year_3': rec.annee_fiscal - 1,
                        'year_4': rec.annee_fiscal,
                        }
                self.env['wk.bilan.cat1'].create(vals)
                self.env['wk.bilan.cat2'].create(vals)
                self.env['wk.bilan.cat3'].create(vals)
                self.env['wk.bilan.cat4'].create(vals)
                self.env['wk.bilan.cat5'].create(vals)
                count = 1
                for index, item in list_bilan:
                    line = self.env['wk.bilan'].create({'declaration': item,
                                                        'categorie': index,
                                                        'etape_id': rec.id,
                                                        'sequence': count})
                    count += 1
                count = 1
                for item in list_recap:
                    line = self.env['wk.recap'].create({'declaration': item, 'etape_id': rec.id, 'sequence': count})
                    count += 1
                count = 1
                for item in list_var:
                    line = self.env['wk.variable'].create({'var': item, 'etape_id': rec.id, 'sequence': count})
                    count += 1
            years = self.env['wk.year'].search([])
            if not years:
                start = datetime.date.today().year - 2
                stop =  datetime.date.today().year + 10
                for i in range(start, stop):
                    self.env['wk.year'].create({'name': str(i)})


    def action_get_view(self):
        for rec in self:
            view_id = self.env.ref('dept_wk.view_wk_etape_form').id
            return {
                'type': 'ir.actions.act_window',
                'name': _('الفرع'),
                'view_mode': 'form',
                'res_model': 'wk.etape',
                'res_id': rec.id,
                'views': [[view_id, 'form']],
            }

    def validate_information_function(self):
        for rec in self:
            if rec.etape.sequence == 1:
                if rec.state_branch == 'branch_1':
                    if self.env.user.has_group('dept_wk.dept_wk_group_agent_agence'):
                        rec.state_branch = 'branch_2'
                        rec.raison_a_revoir = False
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser à valider cette etape'))
                elif rec.state_branch == 'branch_2':
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_agence'):
                        rec.state_branch = 'branch_3'
                        rec.raison_a_revoir = False
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser à valider cette etape'))
                elif rec.state_branch == 'branch_3':
                    if self.env.user.has_group('dept_wk.dept_wk_group_agent_agence'):
                        rec.state_branch = 'branch_4'
                        rec.raison_a_revoir = False
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser à valider cette etape'))
                else:
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_agence'):
                        rec.state_branch = 'branch_5'
                        folder = self.env['documents.folders'].search([('branch', '=', rec.branche.id),
                                                                       ('client', '=', rec.nom_client.id)])
                        if not folder:
                            folder_branch = self.env['documents.folders'].search([('branch', '=', rec.branche.id),
                                                                                  ('client', '=', False)])
                            if not folder_branch:
                                folder_branch = self.env['documents.folders'].create({'branch': rec.branche.id,
                                                                                      'name': rec.branche.ref})
                            folder = self.env['documents.folders'].create({'branch': rec.branche.id,
                                                                          'name': rec.num_compte or 'غير معين',
                                                                          'parent_folder_id': folder_branch.id,
                                                                          'client': rec.nom_client.id})
                        etape_created = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                        if not etape_created:
                            etape = self.env['wk.etape'].create({'workflow': rec.workflow.id,
                                                                 'etape': self.env.ref('dept_wk.princip_2').id,
                                                                 'state_finance': 'finance_1',
                                                                 'annee_fiscal_list': rec.annee_fiscal_list.id,
                                                                 'assigned_to_agence': rec.assigned_to_agence.id})
                            for doc in rec.documents:
                                if doc.document:
                                    self.env['wk.document.check'].create({
                                                                          'list_doc': doc.list_doc,
                                                                      'document': doc.document,
                                                                      'answer': doc.answer,
                                                                      'note': doc.note,
                                                                      'filename': doc.filename,
                                                                      'etape_id': etape.id})

                                    doc_attached = self.env['ir.attachment'].search([('doc', '=', doc.id),
                                                                                     ('folder_id', '=', folder.id)])
                                    if not doc_attached:
                                        self.env['ir.attachment'].create({'folder_id': folder.id,
                                                                          'datas': doc.document,
                                                                          'doc': doc.id,
                                                                          'create_uid': self.env.user,
                                                                          'name': doc.filename if doc.filename else doc.list_doc})
                        else:
                            if etape_created.state_finance != 'finance_4':
                                etape_created.documents.unlink()
                                for doc in rec.documents:
                                    if doc.document:
                                        self.env['wk.document.check'].create({
                                                                              'list_doc': doc.list_doc,
                                                                              'document': doc.document,
                                                                              'answer': doc.answer,
                                                                              'note': doc.note,
                                                                              'filename': doc.filename,
                                                                              'etape_id': etape_created.id})
                        etape_revision = rec.workflow.states.filtered(lambda l: l.etape.sequence == 8)
                        vals = {
                            'etape': self.env.ref('dept_wk.princip_8').id,
                            'assigned_to_agence': rec.assigned_to_agence.id,
                            'workflow': rec.workflow.id,
                            'nom_client': rec.nom_client.id,
                            'branche': rec.branche.id,
                            'gerant': rec.gerant.id,
                            'annee_fiscal_list': rec.annee_fiscal_list.id,
                            'unit_prod': rec.unit_prod,
                            'stock': rec.stock,
                            'prod_company': rec.prod_company,
                            'politique_comm': rec.politique_comm,
                            'cycle_exploit': rec.cycle_exploit,
                            'concurrence': rec.concurrence,
                            'program_invest': rec.program_invest,
                            'result_visit': rec.result_visit,
                            'description_company': rec.description_company,
                            'recommendation_visit': rec.recommendation_visit,
                            'recommendation_responsable_agence': rec.recommendation_responsable_agence,
                        }
                        if not etape_revision:
                            etape_revision = self.env['wk.etape'].create(vals)
                        else:
                            etape_revision.write(vals)
                            etape_revision.documents.unlink()
                            etape_revision.images.unlink()
                            etape_revision.kyc.unlink()
                            etape_revision.apropos.unlink()
                            etape_revision.gestion.unlink()
                            etape_revision.employees.unlink()
                            etape_revision.sieges.unlink()
                            etape_revision.tailles.unlink()
                            etape_revision.situations.unlink()
                            etape_revision.situations_fin.unlink()
                            etape_revision.client.unlink()
                            etape_revision.fournisseur.unlink()
                            etape_revision.facilite_propose.unlink()
                        for doc in rec.documents:
                            self.env['wk.document.check'].create({
                                                                  'list_doc': doc.list_doc,
                                                                  'document': doc.document,
                                                                  'answer': doc.answer,
                                                                  'note': doc.note,
                                                                  'filename': doc.filename,
                                                                  'etape_id': etape_revision.id})
                        for image in rec.images:
                            self.env['wk.documents'].create({'picture': image.picture,
                                                             'name': image.name,
                                                             'etape_id': etape_revision.id})
                        for kyc in rec.kyc:
                            self.env['wk.kyc.details'].create({'info': kyc.info,
                                                               'answer': kyc.answer,
                                                               'detail': kyc.detail,
                                                               'etape_id': etape_revision.id})
                        for a in rec.apropos:
                            self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                              'age': a.age,
                                                              'pourcentage': a.pourcentage,
                                                              'statut_partenaire': a.statut_partenaire,
                                                              'nationalite': a.nationalite.id,
                                                              'etape_id': etape_revision.id
                                                              })
                        for g in rec.gestion:
                            self.env['wk.gestion'].create({
                                'name': g.name,
                                'job': g.job,
                                'niveau_etude': g.niveau_etude,
                                'age': g.age,
                                'experience': g.experience,
                                'etape_id': etape_revision.id
                            })
                        for empl in rec.employees:
                            self.env['wk.nombre.employee'].create({
                                'name': empl.name,
                                'poste_permanent': empl.poste_permanent,
                                'poste_non_permanent': empl.poste_non_permanent,
                                'etape_id': etape_revision.id
                            })
                        for siege in rec.sieges:
                            self.env['wk.siege'].create({
                                'name': siege.name,
                                'adresse': siege.adresse,
                                'nature': siege.nature.id,
                                'etape_id': etape_revision.id
                            })
                        for taille in rec.tailles:
                            self.env['wk.taille'].create({
                                'type_demande': taille.type_demande.id,
                                'montant': taille.montant,
                                'raison': taille.raison,
                                'etape_id': etape_revision.id,
                                'garanties': taille.garanties.ids})
                        for sit in rec.situations:
                            self.env['wk.situation'].create({
                                'banque': sit.banque.id,
                                'type_fin': sit.type_fin.id,
                                'montant': sit.montant,
                                'garanties': sit.garanties,
                                'etape_id': etape_revision.id
                            })
                        for sit in rec.situations_fin:
                            self.env['wk.situation.fin'].create({
                                'type': sit.type,
                                'sequence': sit.sequence,
                                'year1': sit.year1,
                                'year2': sit.year2,
                                'year3': sit.year3,
                                'etape_id': etape_revision.id
                            })
                        for client in rec.client:
                            self.env['wk.client'].create({
                                'name': client.name,
                                'country': client.country.id,
                                'type_payment': client.type_payment.ids,
                                'etape_id': etape_revision.id
                            })
                        for f in rec.fournisseur:
                            self.env['wk.fournisseur'].create({
                                'name': f.name,
                                'country': f.country.id,
                                'type_payment': f.type_payment.ids,
                                'etape_id': etape_revision.id
                            })

                            rec.workflow.state = '2'
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser à valider cette etape'))
            elif rec.etape.sequence == 2:
                if rec.state_finance == 'finance_1':
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_analyste'):
                        rec.state_finance = 'finance_2'
                        rec.workflow.assigned_to_finance = rec.assigned_to_finance.id
                        rec.raison_a_revoir = False
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser'))
                elif rec.state_finance == 'finance_2':
                    rec.state_finance = 'finance_3'
                    rec.raison_a_revoir = False
                elif rec.state_finance == 'finance_3':
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_analyste'):
                        etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                        etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                        rec.state_finance = 'finance_4'
                        self.verrouiller_dossier_function()
                        rec.workflow.state = '3'
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser'))
            elif rec.etape.sequence == 3:
                if rec.state_commercial == 'commercial_1':
                    rec.state_commercial = 'commercial_2'
                    rec.raison_a_revoir = False
                elif rec.state_commercial == 'commercial_2':
                    rec.state_commercial = 'commercial_3'
                    rec.raison_a_revoir = False
                elif rec.state_commercial == 'commercial_3':
                    etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                    rec.state_commercial = 'commercial_4'
                    rec.workflow.state = '4'
                    if etape_fin.state_finance == 'finance_6':
                        etape_fin.state_finance = 'finance_4'
                    etape_1 = rec.workflow.states.filtered(lambda l: l.sequence == 1)
                    rec.raison_a_revoir = False
                    etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                    if not etape_risk:
                        etape_risk = self.env['wk.etape'].create({'workflow': rec.workflow.id,
                                                                  'etape': self.env.ref('dept_wk.princip_4').id,
                                                                  'risk_scoring': etape_1.risk_scoring.id,
                                                                  'state_risque': 'risque_1'})

                    rec.raison_a_revoir = False
            elif rec.etape.sequence == 4:
                if rec.state_risque == 'risque_1':
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_risque'):
                        rec.state_risque = 'risque_3'
                        rec.workflow.assigned_to_risque = rec.assigned_to_risque.id
                        rec.raison_a_revoir = False
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser'))
                elif rec.state_risque == 'risque_3':
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_credit'):
                        rec.state_risque = 'risque_4'
                        rec.raison_a_revoir = False
                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser'))
                elif rec.state_risque == 'risque_4':
                    if self.env.user.has_group('dept_wk.dept_wk_group_responsable_risque'):
                        rec.state_risque = 'risque_2'
                        rec.raison_a_revoir = False
                        etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                        # etape_fin.state_finance = 'finance_4'
                        etape_fin.raison_a_revoir = False
                        if etape_fin.comite == self.env.ref('dept_wk.pouvoir_1'):
                            rec.workflow.state = '10'
                            etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                            etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                            etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                            etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 10)
                            vals = {
                                'workflow': rec.workflow.id,
                                'nom_client': etape_1.nom_client.id,
                                'branche': etape_1.branche.id,
                                'num_compte': etape_1.num_compte,
                                'demande': etape_1.demande.id,
                                'state_tres': 'tres_1',
                                'etape': self.env.ref('dept_wk.princip_10').id,
                                'gerant': etape_1.gerant.id,
                                'unit_prod': etape_1.unit_prod,
                                'stock': etape_1.stock,
                                'prod_company': etape_1.prod_company,
                                'politique_comm': etape_1.politique_comm,
                                'cycle_exploit': etape_1.cycle_exploit,
                                'concurrence': etape_1.concurrence,
                                'program_invest': etape_1.program_invest,
                                'result_visit': etape_1.result_visit,
                                'description_company': etape_1.description_company,
                                'recommendation_visit': etape_1.recommendation_visit,
                                'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                                'analyse_secteur_act': etape_comm.analyse_secteur_act,
                                'analyse_concurrence': etape_comm.analyse_concurrence,
                                'ampleur_benefice': etape_comm.ampleur_benefice,
                                'analyse_relation': etape_comm.analyse_relation,
                                'recommendation_dir_commercial': etape_comm.recommendation_dir_commercial,
                                'recommendation_commercial': etape_comm.recommendation_commercial,
                                'risk_scoring': etape_risk.risk_scoring.id,
                                'recommandation_dir_risque': etape_risk.recommandation_dir_risque,
                                'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                                'garantie_ids': etape_fin.garantie_ids.ids,
                                'exception_ids': etape_fin.exception_ids.ids,
                                'comite': etape_fin.comite.id,
                                'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
                            }
                            if etape:
                                etape.write(vals)
                                etape.documents.unlink()
                                etape.images.unlink()
                                etape.kyc.unlink()
                                etape.apropos.unlink()
                                etape.gestion.unlink()
                                etape.employees.unlink()
                                etape.sieges.unlink()
                                etape.tailles.unlink()
                                etape.situations.unlink()
                                etape.situations_fin.unlink()
                                etape.client.unlink()
                                etape.fournisseur.unlink()
                                etape.facilite_propose.unlink()
                            else:
                                etape = self.env['wk.etape'].create(vals)
                            for doc in etape_1.documents:
                                self.env['wk.document.check'].create({
                                    'list_doc': doc.list_doc,
                                    'document': doc.document,
                                    'answer': doc.answer,
                                    'note': doc.note,
                                    'filename': doc.filename,
                                    'etape_id': etape.id})
                            for image in etape_1.images:
                                self.env['wk.documents'].create({'picture': image.picture,
                                                                 'name': image.name,
                                                                 'etape_id': etape.id})
                            for kyc in etape_1.kyc:
                                self.env['wk.kyc.details'].create({'info': kyc.info,
                                                                   'answer': kyc.answer,
                                                                   'detail': kyc.detail,
                                                                   'etape_id': etape.id})
                            for a in etape_1.apropos:
                                self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                                  'age': a.age,
                                                                  'pourcentage': a.pourcentage,
                                                                  'statut_partenaire': a.statut_partenaire,
                                                                  'nationalite': a.nationalite.id,
                                                                  'etape_id': etape.id
                                                                  })
                            for g in etape_1.gestion:
                                self.env['wk.gestion'].create({
                                    'name': g.name,
                                    'job': g.job,
                                    'niveau_etude': g.niveau_etude,
                                    'age': g.age,
                                    'experience': g.experience,
                                    'etape_id': etape.id
                                })
                            for empl in etape_1.employees:
                                self.env['wk.nombre.employee'].create({
                                    'name': empl.name,
                                    'poste_permanent': empl.poste_permanent,
                                    'poste_non_permanent': empl.poste_non_permanent,
                                    'etape_id': etape.id
                                })
                            for siege in etape_1.sieges:
                                self.env['wk.siege'].create({
                                    'name': siege.name,
                                    'adresse': siege.adresse,
                                    'nature': siege.nature.id,
                                    'etape_id': etape.id
                                })
                            for taille in etape_1.tailles:
                                self.env['wk.taille'].create({
                                    'type_demande': taille.type_demande.id,
                                    'montant': taille.montant,
                                    'raison': taille.raison,
                                    'etape_id': etape.id,
                                    'garanties': taille.garanties.ids})
                            for sit in etape_1.situations:
                                self.env['wk.situation'].create({
                                    'banque': sit.banque.id,
                                    'type_fin': sit.type_fin.id,
                                    'montant': sit.montant,
                                    'garanties': sit.garanties,
                                    'etape_id': etape.id
                                })
                            for sit in etape_1.situations_fin:
                                self.env['wk.situation.fin'].create({
                                    'type': sit.type,
                                    'sequence': sit.sequence,
                                    'year1': sit.year1,
                                    'year2': sit.year2,
                                    'year3': sit.year3,
                                    'etape_id': etape.id
                                })
                            for client in etape_1.client:
                                self.env['wk.client'].create({
                                    'name': client.name,
                                    'country': client.country.id,
                                    'type_payment': client.type_payment.ids,
                                    'etape_id': etape.id
                                })
                            for f in etape_1.fournisseur:
                                self.env['wk.fournisseur'].create({
                                    'name': f.name,
                                    'country': f.country.id,
                                    'type_payment': f.type_payment.ids,
                                    'etape_id': etape.id
                                })
                            for fac in etape_fin.facilite_propose:
                                self.env['wk.facilite.propose'].create({
                                    'type_facilite': fac.type_facilite.id,
                                    'type_demande_ids': fac.type_demande_ids.ids,
                                    'montant_dz': fac.montant_dz,
                                    'preg': fac.preg,
                                    'duree': fac.duree,
                                    'condition': fac.condition,
                                    'etape_id': etape.id})
                        else:
                            rec.workflow.state = '5'
                            etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                            etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                            etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                            etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
                            vals = {
                                'workflow': rec.workflow.id,
                                'nom_client': etape_1.nom_client.id,
                                'branche': etape_1.branche.id,
                                'num_compte': etape_1.num_compte,
                                'demande': etape_1.demande.id,
                                'state_vice': 'vice_1',
                                'etape': self.env.ref('dept_wk.princip_5').id,
                                'gerant': etape_1.gerant.id,
                                'unit_prod': etape_1.unit_prod,
                                'stock': etape_1.stock,
                                'prod_company': etape_1.prod_company,
                                'politique_comm': etape_1.politique_comm,
                                'cycle_exploit': etape_1.cycle_exploit,
                                'concurrence': etape_1.concurrence,
                                'program_invest': etape_1.program_invest,
                                'result_visit': etape_1.result_visit,
                                'description_company': etape_1.description_company,
                                'recommendation_visit': etape_1.recommendation_visit,
                                'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                                'analyse_secteur_act': etape_comm.analyse_secteur_act,
                                'analyse_concurrence': etape_comm.analyse_concurrence,
                                'ampleur_benefice': etape_comm.ampleur_benefice,
                                'analyse_relation': etape_comm.analyse_relation,
                                'recommendation_dir_commercial': etape_comm.recommendation_dir_commercial,
                                'recommendation_commercial': etape_comm.recommendation_commercial,
                                'risk_scoring': etape_risk.risk_scoring.id,
                                'recommandation_dir_risque': etape_risk.recommandation_dir_risque,
                                'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                                'garantie_ids': etape_fin.garantie_ids.ids,
                                'exception_ids': etape_fin.exception_ids.ids,
                                'comite': etape_fin.comite.id,
                                'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
                            }
                            if etape:
                                etape.write(vals)
                                etape.documents.unlink()
                                etape.images.unlink()
                                etape.kyc.unlink()
                                etape.apropos.unlink()
                                etape.gestion.unlink()
                                etape.employees.unlink()
                                etape.sieges.unlink()
                                etape.tailles.unlink()
                                etape.situations.unlink()
                                etape.situations_fin.unlink()
                                etape.client.unlink()
                                etape.fournisseur.unlink()
                                etape.facilite_propose.unlink()
                            else:
                                etape = self.env['wk.etape'].create(vals)
                            for doc in etape_1.documents:
                                self.env['wk.document.check'].create({
                                    'list_doc': doc.list_doc,
                                    'document': doc.document,
                                    'answer': doc.answer,
                                    'note': doc.note,
                                    'filename': doc.filename,
                                    'etape_id': etape.id})
                            for image in etape_1.images:
                                self.env['wk.documents'].create({'picture': image.picture,
                                                                 'name': image.name,
                                                                 'etape_id': etape.id})
                            for kyc in etape_1.kyc:
                                self.env['wk.kyc.details'].create({'info': kyc.info,
                                                                   'answer': kyc.answer,
                                                                   'detail': kyc.detail,
                                                                   'etape_id': etape.id})
                            for a in etape_1.apropos:
                                self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                                  'age': a.age,
                                                                  'pourcentage': a.pourcentage,
                                                                  'statut_partenaire': a.statut_partenaire,
                                                                  'nationalite': a.nationalite.id,
                                                                  'etape_id': etape.id
                                                                  })
                            for g in etape_1.gestion:
                                self.env['wk.gestion'].create({
                                    'name': g.name,
                                    'job': g.job,
                                    'niveau_etude': g.niveau_etude,
                                    'age': g.age,
                                    'experience': g.experience,
                                    'etape_id': etape.id
                                })
                            for empl in etape_1.employees:
                                self.env['wk.nombre.employee'].create({
                                    'name': empl.name,
                                    'poste_permanent': empl.poste_permanent,
                                    'poste_non_permanent': empl.poste_non_permanent,
                                    'etape_id': etape.id
                                })
                            for siege in etape_1.sieges:
                                self.env['wk.siege'].create({
                                    'name': siege.name,
                                    'adresse': siege.adresse,
                                    'nature': siege.nature.id,
                                    'etape_id': etape.id
                                })
                            for taille in etape_1.tailles:
                                self.env['wk.taille'].create({
                                    'type_demande': taille.type_demande.id,
                                    'montant': taille.montant,
                                    'raison': taille.raison,
                                    'etape_id': etape.id,
                                    'garanties': taille.garanties.ids})
                            for sit in etape_1.situations:
                                self.env['wk.situation'].create({
                                    'banque': sit.banque.id,
                                    'type_fin': sit.type_fin.id,
                                    'montant': sit.montant,
                                    'garanties': sit.garanties,
                                    'etape_id': etape.id
                                })
                            for sit in etape_1.situations_fin:
                                self.env['wk.situation.fin'].create({
                                    'type': sit.type,
                                    'sequence': sit.sequence,
                                    'year1': sit.year1,
                                    'year2': sit.year2,
                                    'year3': sit.year3,
                                    'etape_id': etape.id
                                })
                            for client in etape_1.client:
                                self.env['wk.client'].create({
                                    'name': client.name,
                                    'country': client.country.id,
                                    'type_payment': client.type_payment.ids,
                                    'etape_id': etape.id
                                })
                            for f in etape_1.fournisseur:
                                self.env['wk.fournisseur'].create({
                                    'name': f.name,
                                    'country': f.country.id,
                                    'type_payment': f.type_payment.ids,
                                    'etape_id': etape.id
                                })
                            for fac in etape_fin.facilite_propose:
                                self.env['wk.facilite.propose'].create({
                                    'type_facilite': fac.type_facilite.id,
                                    'type_demande_ids': fac.type_demande_ids.ids,
                                    'montant_dz': fac.montant_dz,
                                    'preg': fac.preg,
                                    'duree': fac.duree,
                                    'condition': fac.condition,
                                    'etape_id': etape.id})

                    else:
                        raise ValidationError(_('Vous n\'etes pas autoriser'))
            elif rec.etape.sequence == 10:
                if rec.state_tres == 'tres_1' and self.env.user.has_group('dept_wk.dept_wk_group_tres'):
                    rec.state_tres = 'tres_2'
                    rec.raison_a_revoir = False
                    etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                    # etape_fin.state_finance = 'finance_4'
                    etape_fin.raison_a_revoir = False
                    rec.workflow.state = '5'
                    etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                    etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                    etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                    etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
                    vals = {
                        'workflow': rec.workflow.id,
                        'nom_client': etape_1.nom_client.id,
                        'branche': etape_1.branche.id,
                        'num_compte': etape_1.num_compte,
                        'demande': etape_1.demande.id,
                        'state_vice': 'vice_1',
                        'etape': self.env.ref('dept_wk.princip_5').id,
                        'gerant': etape_1.gerant.id,
                        'unit_prod': etape_1.unit_prod,
                        'stock': etape_1.stock,
                        'prod_company': etape_1.prod_company,
                        'politique_comm': etape_1.politique_comm,
                        'cycle_exploit': etape_1.cycle_exploit,
                        'concurrence': etape_1.concurrence,
                        'program_invest': etape_1.program_invest,
                        'result_visit': etape_1.result_visit,
                        'description_company': etape_1.description_company,
                        'recommendation_visit': etape_1.recommendation_visit,
                        'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                        'analyse_secteur_act': etape_comm.analyse_secteur_act,
                        'analyse_concurrence': etape_comm.analyse_concurrence,
                        'ampleur_benefice': etape_comm.ampleur_benefice,
                        'analyse_relation': etape_comm.analyse_relation,
                        'recommendation_dir_commercial': etape_comm.recommendation_dir_commercial,
                        'recommendation_commercial': etape_comm.recommendation_commercial,
                        'risk_scoring': etape_risk.risk_scoring.id,
                        'recommandation_dir_risque': etape_risk.recommandation_dir_risque,
                        'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                        'garantie_ids': etape_fin.garantie_ids.ids,
                        'exception_ids': etape_fin.exception_ids.ids,
                        'comite': etape_fin.comite.id,
                        'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
                    }
                    if etape:
                        etape.write(vals)
                        etape.documents.unlink()
                        etape.images.unlink()
                        etape.kyc.unlink()
                        etape.apropos.unlink()
                        etape.gestion.unlink()
                        etape.employees.unlink()
                        etape.sieges.unlink()
                        etape.tailles.unlink()
                        etape.situations.unlink()
                        etape.situations_fin.unlink()
                        etape.client.unlink()
                        etape.fournisseur.unlink()
                        etape.facilite_propose.unlink()
                    else:
                        etape = self.env['wk.etape'].create(vals)
                    for doc in etape_1.documents:
                        self.env['wk.document.check'].create({
                            'list_doc': doc.list_doc,
                            'document': doc.document,
                            'answer': doc.answer,
                            'note': doc.note,
                            'filename': doc.filename,
                            'etape_id': etape.id})
                    for image in etape_1.images:
                        self.env['wk.documents'].create({'picture': image.picture,
                                                         'name': image.name,
                                                         'etape_id': etape.id})
                    for kyc in etape_1.kyc:
                        self.env['wk.kyc.details'].create({'info': kyc.info,
                                                           'answer': kyc.answer,
                                                           'detail': kyc.detail,
                                                           'etape_id': etape.id})
                    for a in etape_1.apropos:
                        self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                          'age': a.age,
                                                          'pourcentage': a.pourcentage,
                                                          'statut_partenaire': a.statut_partenaire,
                                                          'nationalite': a.nationalite.id,
                                                          'etape_id': etape.id
                                                          })
                    for g in etape_1.gestion:
                        self.env['wk.gestion'].create({
                            'name': g.name,
                            'job': g.job,
                            'niveau_etude': g.niveau_etude,
                            'age': g.age,
                            'experience': g.experience,
                            'etape_id': etape.id
                        })
                    for empl in etape_1.employees:
                        self.env['wk.nombre.employee'].create({
                            'name': empl.name,
                            'poste_permanent': empl.poste_permanent,
                            'poste_non_permanent': empl.poste_non_permanent,
                            'etape_id': etape.id
                        })
                    for siege in etape_1.sieges:
                        self.env['wk.siege'].create({
                            'name': siege.name,
                            'adresse': siege.adresse,
                            'nature': siege.nature.id,
                            'etape_id': etape.id
                        })
                    for taille in etape_1.tailles:
                        self.env['wk.taille'].create({
                            'type_demande': taille.type_demande.id,
                            'montant': taille.montant,
                            'raison': taille.raison,
                            'etape_id': etape.id,
                            'garanties': taille.garanties.ids})
                    for sit in etape_1.situations:
                        self.env['wk.situation'].create({
                            'banque': sit.banque.id,
                            'type_fin': sit.type_fin.id,
                            'montant': sit.montant,
                            'garanties': sit.garanties,
                            'etape_id': etape.id
                        })
                    for sit in etape_1.situations_fin:
                        self.env['wk.situation.fin'].create({
                            'type': sit.type,
                            'sequence': sit.sequence,
                            'year1': sit.year1,
                            'year2': sit.year2,
                            'year3': sit.year3,
                            'etape_id': etape.id
                        })
                    for client in etape_1.client:
                        self.env['wk.client'].create({
                            'name': client.name,
                            'country': client.country.id,
                            'type_payment': client.type_payment.ids,
                            'etape_id': etape.id
                        })
                    for f in etape_1.fournisseur:
                        self.env['wk.fournisseur'].create({
                            'name': f.name,
                            'country': f.country.id,
                            'type_payment': f.type_payment.ids,
                            'etape_id': etape.id
                        })
                    for fac in etape_fin.facilite_propose:
                        self.env['wk.facilite.propose'].create({
                            'type_facilite': fac.type_facilite.id,
                            'type_demande_ids': fac.type_demande_ids.ids,
                            'montant_dz': fac.montant_dz,
                            'preg': fac.preg,
                            'duree': fac.duree,
                            'condition': fac.condition,
                            'etape_id': etape.id})
                else:
                    raise ValidationError(_('Vous n\'etes pas autoriser'))
            elif rec.etape.sequence == 5:
                if rec.state_vice == 'vice_1':
                    rec.state_vice = 'vice_2'
                    etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                    rec.workflow.state = '6'
                    etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                    etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                    etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                    etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 6)
                    vals = {'workflow': rec.workflow.id,
                            'etape': self.env.ref('dept_wk.princip_6').id,
                            'state_comite': 'comite_1',
                            'nom_client': etape_1.nom_client.id,
                            'gerant': etape_1.gerant.id,
                            'recommendation_visit': etape_1.recommendation_visit,
                            'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                            'analyse_concurrence': etape_comm.analyse_concurrence,
                            'ampleur_benefice': etape_comm.ampleur_benefice,
                            'analyse_relation': etape_comm.analyse_relation,
                            'recommendation_dir_commercial': etape_comm.recommendation_dir_commercial,
                            'recommendation_commercial': etape_comm.recommendation_commercial,
                            'risk_scoring': etape_risk.risk_scoring.id,
                            'recommandation_dir_risque': etape_risk.recommandation_dir_risque,
                            'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                            'garantie_ids': etape_fin.garantie_ids.ids,
                            'exception_ids': etape_fin.exception_ids.ids,
                            'comite': etape_fin.comite.id,
                            'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
                            'recommandation_vice_dir_fin': etape_fin.recommandation_vice_dir_fin,
                            'recommandation_dg': rec.recommandation_dg,
                            }
                    if not etape:
                        etape = self.env['wk.etape'].create(vals)
                    else:
                        etape.write(vals)
                    etape.facilite_propose.unlink()
                    for fac in etape_fin.facilite_propose:
                        self.env['wk.facilite.propose'].create({
                            'type_facilite': fac.type_facilite.id,
                            'type_demande_ids': fac.type_demande_ids.ids,
                            'montant_dz': fac.montant_dz,
                            'preg': fac.preg,
                            'duree': fac.duree,
                            'condition': fac.condition,
                            'etape_id': etape.id})
            elif rec.etape.sequence == 9:
                if rec.state_dg == 'dg_1':
                    rec.state_dg = 'dg_2'
                    etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                    if etape_fin.comite == self.env.ref('dept_wk.pouvoir_1'):
                        rec.workflow.state = '7'
                        rec.workflow.date_fin = datetime.date.today()
                    else:
                        rec.workflow.state = '6'
                        etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                        etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                        etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                        etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 6)
                        vals = {'workflow': rec.workflow.id,
                                     'etape': self.env.ref('dept_wk.princip_6').id,
                                     'state_comite': 'comite_1',
                                     'nom_client': etape_1.nom_client.id,
                                     'gerant': etape_1.gerant.id,
                                     'recommendation_visit': etape_1.recommendation_visit,
                                     'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                                        'analyse_concurrence': etape_comm.analyse_concurrence,
                                        'ampleur_benefice': etape_comm.ampleur_benefice,
                                        'analyse_relation': etape_comm.analyse_relation,
                                        'recommendation_dir_commercial': etape_comm.recommendation_dir_commercial,
                                        'recommendation_commercial': etape_comm.recommendation_commercial,
                                        'risk_scoring': etape_risk.risk_scoring.id,
                                     'recommandation_dir_risque': etape_risk.recommandation_dir_risque,
                                     'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                                     'garantie_ids': etape_fin.garantie_ids.ids,
                                     'exception_ids': etape_fin.exception_ids.ids,
                                     'comite': etape_fin.comite.id,
                                     'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
                                     'recommandation_vice_dir_fin': etape_fin.recommandation_vice_dir_fin,
                                     'recommandation_dg': rec.recommandation_dg,
                                     }
                        if not etape:
                            etape = self.env['wk.etape'].create(vals)
                        else:
                            etape.write(vals)
                        etape.facilite_propose.unlink()
                        for fac in etape_fin.facilite_propose:
                            self.env['wk.facilite.propose'].create({
                                'type_facilite': fac.type_facilite.id,
                                'type_demande_ids': fac.type_demande_ids.ids,
                                'montant_dz': fac.montant_dz,
                                'preg': fac.preg,
                                'duree': fac.duree,
                                'condition': fac.condition,
                                'etape_id': etape.id})
            elif rec.etape.sequence == 6:
                if rec.state_comite == 'comite_1':
                    rec.state_comite = 'comite_2'
                    rec.workflow.state = '7'
                    rec.workflow.date_fin = datetime.date.today()
            email_template = self.env.ref('dept_wk.notification_mail_template')
            email_values = {
                'email_to': self.get_mail_to(),
            }
            email_template.send_mail(self.id, force_send=True, email_values=email_values)


    def a_revoir(self, one_step=False):
        for rec in self:
            if rec.etape.sequence == 1:
                if rec.state_branch == 'branch_2':
                    rec.state_branch = 'branch_1'
                elif rec.state_branch == 'branch_3':
                    rec.state_branch = 'branch_2'
                elif rec.state_branch == 'branch_4':
                    rec.state_branch = 'branch_3'
                elif rec.state_branch == 'branch_5':
                    rec.state_branch = 'branch_4'
            elif rec.etape.sequence == 2:
                if rec.state_finance == 'finance_1':
                    if not rec.dossier_verouiller:
                        rec.workflow.state = '1'
                        etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                        etape.state_branch = 'branch_4'
                    else:
                        raise ValidationError(_('لا يمكنكم طلب المراجعة الملف مقبول'))
                elif rec.state_finance == 'finance_2':
                    if not rec.dossier_verouiller and not one_step:
                        rec.state_finance = 'finance_1'
                        etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                        etape.state_branch = 'branch_4'
                    elif not rec.dossier_verouiller and one_step:
                        rec.state_finance = 'finance_1'
                        rec.dossier_verouiller = False
                elif rec.state_finance == 'finance_3':
                    rec.state_finance = 'finance_2'
                elif rec.state_finance in ['finance_4','finance_5', 'finance_6', 'finance_7']:
                    rec.state_finance = 'finance_3'
            elif rec.etape.sequence == 3:
                if rec.state_commercial == 'commercial_2':
                    rec.state_commercial = 'commercial_1'
                elif rec.state_commercial == 'commercial_3':
                    rec.state_commercial = 'commercial_2'
            elif rec.etape.sequence == 4:
                if rec.state_risque == 'risque_2':
                    rec.state_risque = 'risque_4'
                elif rec.state_risque == 'risque_3':
                    rec.state_risque = 'risque_1'
                elif rec.state_risque == 'risque_4':
                    rec.state_risque = 'risque_3'
            elif rec.etape.sequence == 10:
                if rec.state_tres == 'tres_1':
                    etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                    etape.state_risque = 'risque_4'
                    rec.workflow.state = '4'
                elif rec.state_tres == 'tres_2':
                    rec.state_tres = 'tres_1'
                elif rec.state_vice == 'vice_2':
                    rec.state_vice = 'vice_1'
            elif rec.etape.sequence == 5:
                if rec.state_vice == 'vice_1':
                    etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 10)
                    etape.state_tres = 'tres_1'
                    rec.workflow.state = '10'
                elif rec.state_vice == 'vice_2':
                    rec.state_vice = 'vice_1'
            elif rec.etape.sequence == 6:
                if rec.state_comite == 'comite_1':
                    etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
                    etape.state_vice = 'vice_1'
                    rec.workflow.state = '5'
                elif rec.state_comite == 'comite_2':
                    rec.state_comite = 'comite_1'
            elif rec.etape.sequence == 9:
                if rec.state_dg == 'dg_1':
                    etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
                    etape.state_vice = 'vice_1'
                    rec.workflow.state = '5'
                elif rec.state_dg == 'dg_2':
                    rec.state_dg = 'dg_1'

    def a_revoir_2(self, etat, raison):
        for rec in self:
            if etat == '1':
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
                etape.state_branch = 'branch_4'
                etape.raison_a_revoir = raison
                etape.workflow.state = '1'
            if etat == '2':
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
                etape.state_finance = 'finance_3'
                etape.raison_a_revoir = raison
                etape.workflow.state = '2'
            if etat == '3':
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
                etape.state_commercial = 'commercial_3'
                etape.raison_a_revoir = raison
                etape.workflow.state = '3'
            if etat == '4':
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
                etape.state_risque = 'risque_4'
                etape.raison_a_revoir = raison
                etape.workflow.state = '4'
            if etat == '5':
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
                etape.state_vice = 'vice_1'
                etape.raison_a_revoir = raison
                etape.workflow.state = '5'
            if etat == '10':
                etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 10)
                etape.state_tres = 'tres_1'
                etape.raison_a_revoir = raison
                etape.workflow.state = '10'

    def revoir_action(self):
        for rec in self:
            view_id = self.env.ref('dept_wk.retour_wizard_form').id
            not_one_step = False
            if not rec.sequence == 2 and not rec.state_finance == 'finance_2':
                not_one_step = True
            return {
                'name': 'سبب طلب المراجعة',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'wk.wizard.retour',
                'view_id': view_id,
                'target': 'new',
                'context': {'not_one_step': not_one_step}
            }

    def open_tracking(self):
        self.ensure_one()
        view_id = self.env.ref('dept_wk.view_wk_tracking_tree').id
        return {
                'name': "تتبع",
                'res_model': 'wk.tracking',
                'view_mode': 'tree',
                'view_id': view_id,
                'domain': [('workflow_id', '=', self.workflow.id)],
                'type': 'ir.actions.act_window',
                'context': {'create': False,
                            'edit': False,
                            'delete': False},
            }

    def create_pouvoir(self):
        for rec in self:
            etape_fin = rec.workflow.states.filtered(lambda l: l.etape.sequence == 2)
            # etape_fin.state_finance = 'finance_4'
            etape_fin.raison_a_revoir = False
            rec.workflow.state = '5'
            etape_1 = rec.workflow.states.filtered(lambda l: l.etape.sequence == 1)
            etape_comm = rec.workflow.states.filtered(lambda l: l.etape.sequence == 3)
            etape_risk = rec.workflow.states.filtered(lambda l: l.etape.sequence == 4)
            etape = rec.workflow.states.filtered(lambda l: l.etape.sequence == 5)
            vals = {
                'nom_client': etape_1.nom_client.id,
                'branche': etape_1.branche.id,
                'num_compte': etape_1.num_compte,
                'demande': etape_1.demande.id,
                'etape': self.env.ref('dept_wk.princip_5').id,
                'gerant': etape_1.gerant.id,
                'unit_prod': etape_1.unit_prod,
                'stock': etape_1.stock,
                'prod_company': etape_1.prod_company,
                'politique_comm': etape_1.politique_comm,
                'cycle_exploit': etape_1.cycle_exploit,
                'concurrence': etape_1.concurrence,
                'program_invest': etape_1.program_invest,
                'result_visit': etape_1.result_visit,
                'description_company': etape_1.description_company,
                'recommendation_visit': etape_1.recommendation_visit,
                'recommendation_responsable_agence': etape_1.recommendation_responsable_agence,
                'analyse_secteur_act': etape_comm.analyse_secteur_act,
                'analyse_concurrence': etape_comm.analyse_concurrence,
                'ampleur_benefice': etape_comm.ampleur_benefice,
                'analyse_relation': etape_comm.analyse_relation,
                'recommendation_dir_commercial': etape_comm.recommendation_dir_commercial,
                'recommendation_commercial': etape_comm.recommendation_commercial,
                'risk_scoring': etape_risk.risk_scoring.id,
                'recommandation_dir_risque': etape_risk.recommandation_dir_risque,
                'recommandation_analyste_fin': etape_fin.recommandation_analyste_fin,
                'garantie_ids': etape_fin.garantie_ids.ids,
                'exception_ids': etape_fin.exception_ids.ids,
                'comite': etape_fin.comite.id,
                'recommandation_dir_fin': etape_fin.recommandation_dir_fin,
            }
            if etape:
                etape.write(vals)
                etape.documents.unlink()
                etape.images.unlink()
                etape.kyc.unlink()
                etape.apropos.unlink()
                etape.gestion.unlink()
                etape.employees.unlink()
                etape.sieges.unlink()
                etape.tailles.unlink()
                etape.situations.unlink()
                etape.situations_fin.unlink()
                etape.client.unlink()
                etape.fournisseur.unlink()
                etape.facilite_propose.unlink()
            else:
                etape = self.env['wk.etape'].create(vals)
            for doc in etape_1.documents:
                self.env['wk.document.check'].create({
                    'list_doc': doc.list_doc,
                    'document': doc.document,
                    'answer': doc.answer,
                    'note': doc.note,
                    'filename': doc.filename,
                    'etape_id': etape.id})
            for image in etape_1.images:
                self.env['wk.documents'].create({'picture': image.picture,
                                                 'name': image.name,
                                                 'etape_id': etape.id})
            for kyc in etape_1.kyc:
                self.env['wk.kyc.details'].create({'info': kyc.info,
                                                   'answer': kyc.answer,
                                                   'detail': kyc.detail,
                                                   'etape_id': etape.id})
            for a in etape_1.apropos:
                self.env['wk.partenaire'].create({'nom_partenaire': a.nom_partenaire,
                                                  'age': a.age,
                                                  'pourcentage': a.pourcentage,
                                                  'statut_partenaire': a.statut_partenaire,
                                                  'nationalite': a.nationalite.id,
                                                  'etape_id': etape.id
                                                  })
            for g in etape_1.gestion:
                self.env['wk.gestion'].create({
                    'name': g.name,
                    'job': g.job,
                    'niveau_etude': g.niveau_etude,
                    'age': g.age,
                    'experience': g.experience,
                    'etape_id': etape.id
                })
            for empl in etape_1.employees:
                self.env['wk.nombre.employee'].create({
                    'name': empl.name,
                    'poste_permanent': empl.poste_permanent,
                    'poste_non_permanent': empl.poste_non_permanent,
                    'etape_id': etape.id
                })
            for siege in etape_1.sieges:
                self.env['wk.siege'].create({
                    'name': siege.name,
                    'adresse': siege.adresse,
                    'nature': siege.nature.id,
                    'etape_id': etape.id
                })
            for taille in etape_1.tailles:
                self.env['wk.taille'].create({
                    'type_demande': taille.type_demande.id,
                    'montant': taille.montant,
                    'raison': taille.raison,
                    'etape_id': etape.id,
                    'garanties': taille.garanties.ids})
            for sit in etape_1.situations:
                self.env['wk.situation'].create({
                    'banque': sit.banque.id,
                    'type_fin': sit.type_fin.id,
                    'montant': sit.montant,
                    'garanties': sit.garanties,
                    'etape_id': etape.id
                })
            for sit in etape_1.situations_fin:
                self.env['wk.situation.fin'].create({
                    'type': sit.type,
                    'sequence': sit.sequence,
                    'year1': sit.year1,
                    'year2': sit.year2,
                    'year3': sit.year3,
                    'etape_id': etape.id
                })
            for client in etape_1.client:
                self.env['wk.client'].create({
                    'name': client.name,
                    'country': client.country.id,
                    'type_payment': client.type_payment.ids,
                    'etape_id': etape.id
                })
            for f in etape_1.fournisseur:
                self.env['wk.fournisseur'].create({
                    'name': f.name,
                    'country': f.country.id,
                    'type_payment': f.type_payment.ids,
                    'etape_id': etape.id
                })
            for fac in etape_fin.facilite_propose:
                self.env['wk.facilite.propose'].create({
                    'type_facilite': fac.type_facilite.id,
                    'type_demande_ids': fac.type_demande_ids.ids,
                    'montant_dz': fac.montant_dz,
                    'preg': fac.preg,
                    'duree': fac.duree,
                    'condition': fac.condition,
                    'etape_id': etape.id})
            email_template = self.env.ref('dept_wk.notification_dec_mail_template')
            email_template.send_mail(rec.id, force_send=True)

    def open_revision_branch(self):
        self.ensure_one()
        view_id = self.env.ref('dept_wk.view_wk_etape_form').id
        #etape = self.workflow.states.filtered(lambda l: l.etape.sequence == 8)
        etape = self.env['wk.etape'].search([('workflow', '=', self.workflow.id),
                                             ('etape', '=', self.env.ref('dept_wk.princip_8').id)
                                             ])[-1]

        return {
                'name': "مراجعة بيانات الفرع",
                'res_model': 'wk.etape',
                'res_id': etape.id,
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'create': False,
                            'delete': False},
            }

    def reject_request(self):
        view_id = self.env.ref('dept_wk.retour_wizard_form').id
        ctx = {}
        ctx['refus'] = True
        return {
            'name': 'سبب رفض الطلب',
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'wk.wizard.retour',
            'view_id': view_id,
            'target': 'new',
            'context': ctx
        }
    def reject_request_function(self):
        for rec in self:
            if rec.etape.sequence == 1:
                if rec.state_branch in ['branch_2', 'branch_4'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_agence'):
                    rec.state_branch = 'branch_rejected'
                    rec.workflow.state = '8'
            if rec.etape.sequence == 2:
                if rec.state_finance in ['finance_1', 'finance_3'] and self.env.user.has_group('dept_wk.dept_wk_group_responsable_analyste'):
                    rec.state_finance = 'finance_rejected'
                    rec.workflow.state = '8'
            if rec.etape.sequence == 5:
                if rec.state_vice in ['vice_1']:
                    rec.state_vice = 'vice_rejected'
                    rec.workflow.state = '8'
            if rec.etape.sequence == 6:
                if rec.state_comite in ['comite_1']:
                    rec.state_comite = 'comite_rejected'
                    rec.workflow.state = '8'
            email_template = self.env.ref('dept_wk.notification_refus_mail_template')
            email_values = {
                'email_to': [rec.assigned_to_agence.partner_id.email, rec.assigned_to_finance.partner_id.email],
            }
            email_template.send_mail(rec.id, force_send=True, email_values=email_values)

    def create_tcr_group(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_tcr_view_form').id
            return {
                'name': 'TCR',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.tcr',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 1, 'is_group': True}
            }

    def create_actif_group(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_actif_view_form').id
            return {
                'name': 'Actif',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.actif',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 1, 'is_group': True}
            }

    def create_passif_group(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_passif_view_form').id
            return {
                'name': 'Passif',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.passif',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 1, 'is_group': True}
            }

    def import_data_group(self):
        for rec in self:
            if rec.tcr_group.state not in ['valide', 'modified'] or rec.actif_group.state not in ['valide', 'modified'] or rec.passif_group.state not in ['valide', 'modified']:
                raise ValidationError("Vous devriez d'abord valider les bilans")
            else:
                bilan_1 = rec.companies_fisc.filtered(lambda r: r.sequence == 1)
                # total I حقوق الملكية
                passif_1 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 12)
                bilan_1.write({'year_4': passif_1.montant_n,
                               'year_3': passif_1.montant_n1})

                bilan_2 = rec.companies_fisc.filtered(lambda r: r.sequence == 2)
                # مجموع الديون
                passif_2 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 14)
                passif_3 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                bilan_2.write({'year_4': passif_2.montant_n + passif_3.montant_n,
                               'year_3': passif_2.montant_n1 + passif_3.montant_n1})
                #سبة المديونية leverage
                bilan_3 = rec.companies_fisc.filtered(lambda r: r.sequence == 3)
                actif_1 = rec.actif_group.actif_lines.filtered(lambda r: r.rubrique.sequence == 26)
                passif_5 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                passif_6 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 14)
                bilan_3.write({'year_4': (passif_6.montant_n + passif_5.montant_n - actif_1.montant_n) / passif_1.montant_n if passif_1.montant_n != 0 else 0,
                                'year_3': (passif_6.montant_n1 + passif_5.montant_n1 - actif_1.montant_n1) / passif_1.montant_n1 if passif_1.montant_n1 != 0 else 0})

                #Gearing
                bilan_4 = rec.companies_fisc.filtered(lambda r: r.sequence == 4)
                bilan_4.write({'year_4': ((passif_6.montant_n + passif_5.montant_n) / passif_1.montant_n) * 100 if passif_1.montant_n != 0 else 0,
                                'year_3': ((passif_6.montant_n1 + passif_5.montant_n1) / passif_1.montant_n1) * 100 if passif_1.montant_n1 != 0 else 0})

                #رقم الاعمال
                bilan_5 = rec.companies_fisc.filtered(lambda r: r.sequence == 5)
                tcr_1 = rec.tcr_group.tcr_lines.filtered(lambda r: r.rubrique.sequence == 7)
                bilan_5.write({'year_4': tcr_1.montant_n,
                                'year_3': tcr_1.montant_n1,
                                })
                # EBIDTA
                tcr_2 = rec.tcr_group.tcr_lines.filtered(lambda r: r.rubrique.sequence == 33)
                bilan_6 = rec.companies_fisc.filtered(lambda r: r.sequence == 6)
                bilan_6.write({'year_4': tcr_2.montant_n,
                                'year_3': tcr_2.montant_n1})

                #صافي الربح
                bilan_7 = rec.companies_fisc.filtered(lambda r: r.sequence == 7)
                tcr_3 = rec.tcr_group.tcr_lines.filtered(lambda r: r.rubrique.sequence == 50)
                bilan_7.write({'year_4': tcr_3.montant_n,
                               'year_3': tcr_3.montant_n1,})

                # راس المال العامل
                bilan_8 = rec.companies_fisc.filtered(lambda r: r.sequence == 8)
                actif_2 = rec.actif_group.actif_lines.filtered(lambda r: r.rubrique.sequence == 16)
                passif_18 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                bilan_8.write({'year_4': passif_1.montant_n - actif_2.montant_n,
                                'year_3': passif_1.montant_n1 - actif_2.montant_n1
                                })

                #احتياجات راس المال العامل
                bilan_9 = rec.companies_fisc.filtered(lambda r: r.sequence == 9)
                passif_4_1 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                passif_20 = rec.passif_group.passif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                actif_18 = rec.actif_group.actif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                actif_20 = rec.actif_group.actif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                bilan_9.write({'year_4': actif_18.montant_n + actif_20.montant_n - passif_20.montant_n,
                               'year_3': actif_18.montant_n1 + actif_20.montant_n1 - passif_20.montant_n1
                                })

    def create_viz3(self):
        for rec in self:
            line1 =  rec.companies_fisc.filtered(lambda r: r.sequence == 7)
            line2 = rec.companies_fisc.filtered(lambda r: r.sequence == 5)
            line3 = rec.companies_fisc.filtered(lambda r: r.sequence == 6)
            data1 = [line1.year_1, line1.year_2, line1.year_3, line1.year_4]
            data2 = [line2.year_1, line2.year_2, line2.year_3, line2.year_4]
            data3 = [line3.year_1, line3.year_2, line3.year_3, line3.year_4]
            label1 = 'NRC'
            label2 = 'CA'
            year = ["N-3", "N-2", "N-1", "N"]
            fig, ax = plt.subplots()
            width = 0.12
            X_axis = np.arange(len(year))
            rects1 = ax.bar(X_axis - width, data1, width, color="yellow", label=label1)
            rects2 = ax.bar(X_axis, data2, width, color="orange", label=label2)
            rects3 = ax.bar(X_axis + width, data3, width, color="red", label="EBITDA")
            ax.set_ylabel('Montant')
            ax.set_title('Montant par année')
            ax.set_xticks(X_axis + width, year)
            ax.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))
            fig.tight_layout()
            buf = BytesIO()
            plt.savefig(buf, format='jpeg', dpi=100)
            buf.seek(0)
            rec.visualisation2 = base64.b64encode(buf.getvalue())
            buf.close()

    def action_create_tcr(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_tcr_view_form').id
            return {
                'name': 'TCR',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.tcr',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 1}
            }

    def action_open_risk(self):
        for rec in self:
            view_id = self.env.ref('dept_wk.view_risk_scoring_form').id
            rec.risk_scoring = rec.workflow.risk_scoring.id
            return {
                'name': 'ادارة المخاطر',
                'res_model': 'risk.scoring',
                'view_mode': 'form',
                'res_id': rec.risk_scoring.id,
                'view_id': view_id,
                'type': 'ir.actions.act_window',
            }

    def action_create_tcr1(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_tcr_view_form').id
            return {
                'name': 'TCR',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.tcr',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 2}
            }

    def action_create_actif(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_actif_view_form').id
            return {
                'name': 'Actif',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.actif',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 1}
            }

    def action_create_actif1(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_actif_view_form').id
            return {
                'name': 'Actif',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.actif',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 2}
            }

    def action_create_passif(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_passif_view_form').id
            return {
                'name': 'Passif',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.passif',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 1}
            }

    def action_create_passif1(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.import_ocr_passif_view_form').id
            return {
                'name': 'Passif',
                'domain': [('parent_id', '=', rec.id)],
                'res_model': 'import.ocr.passif',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'parent_id': rec.id, 'year': 2}
            }

    def import_data(self):
        for rec in self:
            if rec.tcr_id.state not in ['valide', 'modified'] or rec.actif_id.state not in ['valide', 'modified'] or rec.passif_id.state not in ['valide', 'modified']:
                raise ValidationError("Vous devriez d'abord valider les bilans")
            else:
                to_delete = rec.bilan_id.filtered(lambda r: r.declaration in ['ACTIF NET IMMOBILISE CORPOREL',
                                                                              'الات ومعدات و عتاد نقل',
                                                                              'إهتلاكات المعدات',
                                                                              'اهتلاكات / آلات و معدات و عتاد نقل'])
                if to_delete:
                    to_delete1 = rec.bilan1_id.filtered(lambda r: r.sequence in [6, 7, 8, 9])
                    to_delete.unlink()
                    to_delete1.unlink()
                    other_bilan = rec.bilan_id.filtered(lambda r: r.sequence >= 6)
                    count = 6
                    for item in other_bilan:
                        in_view = self.env[f'wk.bilan.cat{item.categorie}'].search([('bilan', '=', item.id)])
                        item.sequence = in_view.sequence = count
                        count += 1
                    print(other_bilan)
                bilan_1 = rec.bilan_id.filtered(lambda r: r.sequence == 1)
                # total I حقوق الملكية
                passif_1 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 12)
                passif1_1 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 12)
                bilan_1.write({'year_4': passif_1.montant_n,
                               'year_3': passif_1.montant_n1,
                               'year_2': passif1_1.montant_n,
                               'year_1': passif1_1.montant_n1})

                # capital emis رأس المال
                bilan_2 = rec.bilan_id.filtered(lambda r: r.sequence == 2)
                passif_2 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 2)
                passif1_2 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 2)
                bilan_2.write({'year_4': passif_2.montant_n,
                               'year_3': passif_2.montant_n1,
                               'year_2': passif1_2.montant_n,
                               'year_1': passif1_2.montant_n1,
                               })

                # Passif - Autres capitaux propres - report à nouveau الاحتياطات
                bilan_3 = rec.bilan_id.filtered(lambda r: r.sequence == 3)
                passif_3 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 4)
                passif1_3 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 4)
                bilan_3.write({'year_4': passif_3.montant_n,
                               'year_3': passif_3.montant_n1,
                               'year_2': passif1_3.montant_n,
                               'year_1': passif1_3.montant_n1})

                # الارباح المتراكمة
                bilan_4 = rec.bilan_id.filtered(lambda r: r.sequence == 4)
                passif_3 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 8)
                passif1_3 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 8)
                tcr_3 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 50)
                tcr1_3 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 50)
                bilan_4.write({'year_4': passif_3.montant_n + tcr_3.montant_n,
                               'year_3': passif_3.montant_n1 + tcr_3.montant_n1,
                               'year_2': passif1_3.montant_n + tcr1_3.montant_n,
                               'year_1': passif1_3.montant_n1 + tcr1_3.montant_n1})

                #حقوق الملكية / مجموع الميزانية
                bilan_5 = rec.bilan_id.filtered(lambda r: r.sequence == 5)
                passif_12 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 25)
                passif1_12 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 25)
                bilan_5.write({'year_4': (passif_1.montant_n / passif_12.montant_n) * 100 if passif_12.montant_n != 0 else 0,
                               'year_3': (passif_1.montant_n1 / passif_12.montant_n1) * 100 if passif_12.montant_n1 != 0 else 0,
                               'year_2': (passif1_1.montant_n / passif1_12.montant_n) * 100 if passif1_12.montant_n != 0 else 0,
                               'year_1': (passif1_1.montant_n1 / passif1_12.montant_n1) * 100 if passif1_12.montant_n1 != 0 else 0})
                if passif_12.montant_n == 0:
                    bilan_5.is_null_4 = True
                else:
                    bilan_5.is_null_4 = False
                if passif_12.montant_n1 == 0:
                    bilan_5.is_null_3 = True
                else:
                    bilan_5.is_null_3 = False
                if passif1_12.montant_n == 0:
                    bilan_5.is_null_2 = True
                else:
                    bilan_5.is_null_2 = False
                if passif1_12.montant_n1 == 0:
                    bilan_5.is_null_1 = True
                else:
                    bilan_5.is_null_1 = False
                '''# Actif net immobilisé corporel
                bilan_6 = rec.bilan_id.filtered(lambda r: r.sequence == 6)
                actif_1 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 4)
                actif1_1 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 4)
                bilan_6.write({'year_4': actif_1.montant_n,
                               'year_3': actif_1.montant_n1,
                               'year_2': actif1_1.montant_n,
                               'year_1': actif1_1.montant_n1})

                # الات ومعدات وعتاد نقل
                bilan_7 = rec.bilan_id.filtered(lambda r: r.sequence == 7)
                actif_7 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 7)
                actif1_7 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 7)
                bilan_7.write({'year_4': actif_7.montant_n,
                               'year_3': actif_7.montant_n1,
                               'year_2': actif1_7.montant_n,
                               'year_1': actif1_7.montant_n1})

                #اهتلاكات المعدات
                bilan_8 = rec.bilan_id.filtered(lambda r: r.sequence == 8)
                bilan_8.write({'year_4': actif_7.montant_2n,
                               'year_3': 0,
                               'year_2': 0,
                               'year_1': 0})
                # اهتلاكات / الات ومعدات وعتاد نقل
                bilan_9 = rec.bilan_id.filtered(lambda r: r.sequence == 9)
                bilan_9.write({'year_4': (bilan_8.year_4 / bilan_7.year_4) * 100 if bilan_7.year_4 != 0 else 0,
                               'year_3': (bilan_8.year_3 / bilan_7.year_3) * 100 if bilan_7.year_3 != 0 else 0,
                               'year_2': (bilan_8.year_2 / bilan_7.year_2) * 100 if bilan_7.year_2 != 0 else 0,
                               'year_1': (bilan_8.year_1 / bilan_7.year_1) * 100 if bilan_7.year_1 != 0 else 0})
                if bilan_7.year_4 == 0:
                    bilan_9.is_null_4 = True
                if bilan_7.year_3 == 0:
                    bilan_9.is_null_3 = True
                if bilan_7.year_2 == 0:
                    bilan_9.is_null_2 = True
                if bilan_7.year_1 == 0:
                    bilan_9.is_null_1 = True'''
                # Passif (Total I + Total II) - Total actif non courant   صافي رأس المال العامل
                actif_2 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 16)
                actif1_2 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 16)
                actif_27 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 27)
                actif1_27 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 27)
                passif_24 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                passif1_24 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                bilan_10 = rec.bilan_id.filtered(lambda r: r.sequence == 6)
                print(bilan_10.declaration)
                bilan_10.write({'year_4': actif_27.montant_n - passif_24.montant_n,
                                'year_3': actif_27.montant_n1 - passif_24.montant_n1,
                                'year_2': actif1_27.montant_n - passif1_24.montant_n,
                                'year_1': actif1_27.montant_n1 - passif1_24.montant_n1,
                                })

                # احتياجات رأس المال العامل
                passif_4_1 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                passif1_4_1 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                actif_3 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 27)
                actif1_3 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 27)

                passif_5 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                passif1_5 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                bilan_11 = rec.bilan_id.filtered(lambda r: r.sequence == 7)
                passif_20 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                passif1_20 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                actif_18 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                actif1_18 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                actif_20 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                actif1_20 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                actif_126 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 26)
                actif1_126 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 26)
                bilan_11.write({'year_4': actif_27.montant_n - actif_126.montant_n - (passif_24.montant_n - passif_5.montant_n),
                                'year_3': actif_27.montant_n1 - actif_126.montant_n1 - (passif_24.montant_n1 - passif_5.montant_n1),
                                'year_2': actif1_27.montant_n - actif1_126.montant_n - (passif1_24.montant_n - passif1_5.montant_n),
                                'year_1': actif1_27.montant_n1 - actif1_126.montant_n1 - (passif1_24.montant_n1 - passif1_5.montant_n1),
                                })

                # FR / BFR Passif (Total I + Total II) - Actif (Total actif non courant)  / Actif (Stock et encours + Créances et emploi assimili + Disponibilité et assimilé) - Passif (Total III)
                bilan_12 = rec.bilan_id.filtered(lambda r: r.sequence == 8)
                actif_4 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                actif1_4 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                actif_12 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 19)
                actif1_12 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 19)
                actif_13 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                actif1_13 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 24)

                if bilan_11.year_4 == 0:
                    bilan_12.is_null_4 = True
                if bilan_11.year_4 == 0:
                    bilan_12.is_null_3 = True
                if bilan_11.year_4 == 0:
                    bilan_12.is_null_2 = True
                if bilan_11.year_4 == 0:
                    bilan_12.is_null_1 = True
                bilan_12.write({'year_4': (bilan_10.year_4 / bilan_11.year_4) * 100 if bilan_11.year_4 != 0 else 0,
                                'year_3': (bilan_10.year_3 / bilan_11.year_3) * 100 if bilan_11.year_3 != 0 else 0,
                                'year_2': (bilan_10.year_2 / bilan_11.year_2) * 100 if bilan_11.year_2 != 0 else 0,
                                'year_1': (bilan_10.year_1 / bilan_11.year_1) * 100 if bilan_11.year_1 != 0 else 0,
                                })
                # مجموع المطلوبات Passif - Total II + Total III
                bilan_13 = rec.bilan_id.filtered(lambda r: r.sequence == 9)
                passif_5 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                passif1_5 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                passif_6 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 14)
                passif1_6 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 14)
                passif_22 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 22)
                passif1_22 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 22)
                bilan_13.write({'year_4': passif_5.montant_n + passif_6.montant_n + passif_20.montant_n + passif_22.montant_n,
                                'year_3': passif_5.montant_n1 + passif_6.montant_n1 + passif_20.montant_n1 + passif_22.montant_n1,
                                'year_2': passif1_5.montant_n + passif1_6.montant_n + passif1_20.montant_n + passif1_22.montant_n,
                                'year_1': passif1_5.montant_n1 + passif1_6.montant_n1 + passif1_20.montant_n1 + passif1_22.montant_n1,
                                })

                # التزامات بنكية
                bilan_14 = rec.bilan_id.filtered(lambda r: r.sequence == 10)
                passif_5 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                passif1_5 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 23)
                passif_6 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 14)
                passif1_6 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 14)
                bilan_14.write({'year_4': passif_5.montant_n + passif_6.montant_n,
                                'year_3': passif_5.montant_n1 + passif_6.montant_n1,
                                'year_2': passif1_5.montant_n + passif1_6.montant_n,
                                'year_1': passif1_5.montant_n1 + passif1_6.montant_n1,
                                })
                # تسهيلات الموردين
                bilan_15 = rec.bilan_id.filtered(lambda r: r.sequence == 11)
                passif_7 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                passif1_7 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                var_5 = rec.var_ids.filtered(lambda r: r.sequence == 5)
                bilan_15.write({'year_4': passif_7.montant_n,
                               'year_3': passif_7.montant_n1,
                               'year_2': passif1_7.montant_n,
                               'year_1': passif1_7.montant_n1})

                var_5.write({'montant': passif_7.montant_n})

                # Passif - Impôts مستحقات ضرائب
                bilan_16 = rec.bilan_id.filtered(lambda r: r.sequence == 12)
                passif_8 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 21)
                passif1_8 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 21)
                bilan_16.write({'year_4': passif_8.montant_n,
                               'year_3': passif_8.montant_n1,
                               'year_2': passif1_8.montant_n,
                               'year_1': passif1_8.montant_n1,
                               })

                #Passif - Autres dettes + fournisseur  مطلوبات أخرى متداولة
                bilan_17 = rec.bilan_id.filtered(lambda r: r.sequence == 13)
                passif_8 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 22)
                passif1_8 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 22)
                bilan_17.write({'year_4': passif_8.montant_n,
                                'year_3': passif_8.montant_n1,
                                'year_2': passif1_8.montant_n,
                                'year_1': passif1_8.montant_n1,
                                })

                # (Emprunts et dettes financières passif + Trésorerie passif - Trésorerie coté actif ) / Total I coté passif نسبة المديونية Leverage
                bilan_18 = rec.bilan_id.filtered(lambda r: r.sequence == 14)
                passif_18 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                passif1_18 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 18)
                passif_24 = rec.passif_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                passif1_24 = rec.passif1_id.passif_lines.filtered(lambda r: r.rubrique.sequence == 24)
                bilan_18.write({'year_4': (passif_18.montant_n + passif_24.montant_n) / passif_1.montant_n if passif_1.montant_n != 0 else 0,
                               'year_3': (passif_18.montant_n1 + passif_24.montant_n1) / passif_1.montant_n1 if passif_1.montant_n1 != 0 else 0,
                               'year_2': (passif1_18.montant_n + passif1_24.montant_n) / passif1_1.montant_n if passif1_1.montant_n != 0 else 0,
                               'year_1': (passif1_18.montant_n1 + passif1_24.montant_n1) / passif1_1.montant_n1 if passif1_1.montant_n1 != 0 else 0,
                               })
                if passif_1.montant_n == 0:
                    bilan_18.is_null_4 = True
                else:
                    bilan_18.is_null_4 = False
                if passif_1.montant_n1 == 0:
                    bilan_18.is_null_3 = True
                else:
                    bilan_18.is_null_3 = False
                if passif1_1.montant_n == 0:
                    bilan_18.is_null_2 = True
                else:
                    bilan_18.is_null_2 = False
                if passif1_1.montant_n1 == 0:
                    bilan_18.is_null_1 = True
                else:
                    bilan_18.is_null_1 = False

                #الالتزامات اتجاه البنوك / الحقوق
                bilan_19 = rec.bilan_id.filtered(lambda r: r.sequence == 15)
                bilan_19.write({'year_4': bilan_14.year_4 / passif_1.montant_n if passif_1.montant_n != 0 else 0,
                                'year_3': bilan_14.year_3 / passif_1.montant_n1 if passif_1.montant_n1 != 0 else 0,
                                'year_2': bilan_14.year_2 / passif1_1.montant_n if passif1_1.montant_n != 0 else 0,
                                'year_1': bilan_14.year_1 / passif1_1.montant_n1 if passif1_1.montant_n1 != 0 else 0,
                                })

                if passif_1.montant_n == 0:
                    bilan_19.is_null_4 = True
                else:
                    bilan_19.is_null_4 = False
                if passif_1.montant_n1 == 0:
                    bilan_19.is_null_3 = True
                else:
                    bilan_19.is_null_3 = False
                if passif1_1.montant_n == 0:
                    bilan_19.is_null_2 = True
                else:
                    bilan_19.is_null_2 = False
                if passif1_1.montant_n1 == 0:
                    bilan_19.is_null_1 = True
                else:
                    bilan_19.is_null_1 = False

                # مجموع الميزانية
                bilan_20 = rec.bilan_id.filtered(lambda r: r.sequence == 16)
                bilan_20.write({'year_4': passif_12.montant_n,
                                'year_3': passif_12.montant_n1,
                                'year_2': passif1_12.montant_n,
                                'year_1': passif1_12.montant_n1
                                })
                # المبيعات ، الايرادات
                bilan_21 = rec.bilan_id.filtered(lambda r: r.sequence == 17)
                tcr_1 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 7)
                tcr1_1 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 7)
                var_1 = rec.var_ids.filtered(lambda r: r.sequence == 1)
                bilan_21.write({'year_4': tcr_1.montant_n,
                                'year_3': tcr_1.montant_n1,
                                'year_2': tcr1_1.montant_n,
                                'year_1': tcr1_1.montant_n1
                                })
                var_1.write({'montant': tcr_1.montant_n})
                # TCR- Excédent brut d`exploitation   EBITDA
                bilan_22 = rec.bilan_id.filtered(lambda r: r.sequence == 18)
                tcr_2 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 33)
                tcr1_2 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 33)
                bilan_22.write({'year_4': tcr_2.montant_n,
                                'year_3': tcr_2.montant_n1,
                                'year_2': tcr1_2.montant_n,
                                'year_1': tcr1_2.montant_n1})

                # TCR - Résultat net de l`exercice   صافي الأرباح
                bilan_23 = rec.bilan_id.filtered(lambda r: r.sequence == 19)
                tcr_3 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 50)
                tcr1_3 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 50)
                bilan_23.write({'year_4': tcr_3.montant_n,
                                'year_3': tcr_3.montant_n1,
                                'year_2': tcr1_3.montant_n,
                                'year_1': tcr1_3.montant_n1})

                bilan_24 = rec.bilan_id.filtered(lambda r: r.sequence == 20)
                # صافي الأرباح/المبيعات
                bilan_24.write({'year_4': (tcr_3.montant_n / tcr_1.montant_n) * 100 if tcr_1.montant_n != 0 else 0,
                                'year_3': (tcr_3.montant_n1 / tcr_1.montant_n1) * 100 if tcr_1.montant_n1 != 0 else 0,
                                'year_2': (tcr1_3.montant_n / tcr1_1.montant_n) * 100 if tcr1_1.montant_n != 0 else 0,
                                'year_1': (tcr1_3.montant_n1 / tcr1_1.montant_n1) * 100 if tcr1_1.montant_n1 != 0 else 0})

                if tcr_1.montant_n == 0:
                    bilan_24.is_null_4 = True
                else:
                    bilan_24.is_null_4 = False
                if tcr_1.montant_n1 == 0:
                    bilan_24.is_null_3 = True
                else:
                    bilan_24.is_null_3 = False
                if tcr1_1.montant_n == 0:
                    bilan_24.is_null_2 = True
                else:
                    bilan_24.is_null_2 = False
                if tcr1_1.montant_n1 == 0:
                    bilan_24.is_null_1 = True
                else:
                    bilan_24.is_null_1 = False
                # معدل العائد على الموجودات ROA
                bilan_25 = rec.bilan_id.filtered(lambda r: r.sequence == 21)
                bilan_25.write({'year_4': (tcr_3.montant_n / actif_2.montant_n) * 100 if actif_2.montant_n != 0 else 0,
                                'year_3': (tcr_3.montant_n1 / actif_2.montant_n1) * 100 if actif_2.montant_n1 != 0 else 0,
                                'year_2': (tcr1_3.montant_n / actif1_2.montant_n) * 100 if actif1_2.montant_n != 0 else 0,
                                'year_1': (tcr1_3.montant_n1 / actif1_2.montant_n1) * 100 if actif1_2.montant_n1 != 0 else 0})

                if actif_2.montant_n == 0:
                    bilan_25.is_null_4 = True
                else:
                    bilan_25.is_null_4 = False
                if actif_2.montant_n1 == 0:
                    bilan_25.is_null_3 = True
                else:
                    bilan_25.is_null_3 = False
                if actif1_2.montant_n == 0:
                    bilan_25.is_null_2 = True
                else:
                    bilan_25.is_null_2 = False
                if actif1_2.montant_n1 == 0:
                    bilan_25.is_null_1 = True
                else:
                    bilan_25.is_null_1 = False
                # معدل العائد على حقوق الملكية ROE
                bilan_26 = rec.bilan_id.filtered(lambda r: r.sequence == 22)
                bilan_26.write({'year_4': (tcr_3.montant_n / passif_1.montant_n) * 100 if passif_1.montant_n != 0 else 0,
                                'year_3': (tcr_3.montant_n1 / passif_1.montant_n1) * 100 if passif_1.montant_n1 != 0 else 0,
                                'year_2': (tcr1_3.montant_n / passif1_1.montant_n) * 100 if passif1_1.montant_n != 0 else 0,
                                'year_1': (tcr1_3.montant_n1 / passif1_1.montant_n1) * 100 if passif1_1.montant_n1 != 0 else 0})

                if passif_1.montant_n == 0:
                    bilan_26.is_null_4 = True
                else:
                    bilan_26.is_null_4 = False
                if passif_1.montant_n1 == 0:
                    bilan_26.is_null_3 = True
                else:
                    bilan_26.is_null_3 = False
                if passif1_1.montant_n == 0:
                    bilan_26.is_null_2 = True
                else:
                    bilan_26.is_null_2 = False
                if passif1_1.montant_n1 == 0:
                    bilan_26.is_null_1 = True
                else:
                    bilan_26.is_null_1 = False
                # التدفقات النقدية التشغيلية
                bilan_27 = rec.bilan_id.filtered(lambda r: r.sequence == 23)
                tcr_36 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 36)
                tcr1_36 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 36)
                bilan_27.write(
                    {'year_4': tcr_3.montant_n + tcr_36.montant_n,
                     'year_3': tcr_3.montant_n1 + tcr_36.montant_n1,
                     'year_2': tcr1_3.montant_n + tcr1_36.montant_n,
                     'year_1': tcr1_3.montant_n1 + tcr1_36.montant_n1})

                # نسبة التداول (السيولة)
                actif_26 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 26)
                actif1_26 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 26)
                bilan_28 = rec.bilan_id.filtered(lambda r: r.sequence == 24)
                bilan_28.write({'year_4': (actif_26.montant_n + actif_18.montant_n + actif_20.montant_n) / (passif_5.montant_n + passif_20.montant_n) if passif_5.montant_n + passif_20.montant_n != 0 else 0,
                                'year_3': (actif_26.montant_n1 + actif_18.montant_n1 + actif_20.montant_n1) / (passif_5.montant_n1 + passif_20.montant_n1) if passif_5.montant_n1 + passif_20.montant_n1 != 0 else 0,
                                'year_2': (actif1_26.montant_n + actif1_18.montant_n + actif1_20.montant_n) / (passif1_5.montant_n + passif1_20.montant_n) if passif1_5.montant_n + passif1_20.montant_n != 0 else 0,
                                'year_1': (actif1_26.montant_n1 + actif1_18.montant_n1 + actif1_20.montant_n1) / (passif1_5.montant_n1 + passif1_20.montant_n1) if passif1_5.montant_n1 + passif1_20.montant_n1 != 0 else 0})

                #نسبة السيولة السريعة
                bilan_29 = rec.bilan_id.filtered(lambda r: r.sequence == 25)
                bilan_29.write({'year_4': (actif_26.montant_n) / (passif_5.montant_n + passif_20.montant_n) if passif_5.montant_n + passif_20.montant_n != 0 else 0,
                                'year_3': (actif_26.montant_n) / (passif_5.montant_n1 + passif_20.montant_n1) if passif_5.montant_n1 + passif_20.montant_n1 != 0 else 0,
                                'year_2': (actif1_26.montant_n) / (passif1_5.montant_n + passif1_20.montant_n) if passif1_5.montant_n + passif1_20.montant_n != 0 else 0,
                                'year_1': (actif1_26.montant_n1) / (passif1_5.montant_n1 + passif1_20.montant_n1) if passif1_5.montant_n1 + passif1_20.montant_n1 != 0 else 0,
                                })

                if (passif_5.montant_n + passif_20.montant_n) == 0:
                    bilan_28.is_null_4 = True
                else:
                    bilan_29.is_null_4 = False
                if (passif_5.montant_n1 + passif_20.montant_n1) == 0:
                    bilan_28.is_null_3 = True
                else:
                    bilan_29.is_null_3 = False
                if (passif1_5.montant_n + passif1_20.montant_n) == 0:
                    bilan_28.is_null_2 = True
                else:
                    bilan_29.is_null_2 = False
                if (passif1_5.montant_n1 + passif1_20.montant_n1) == 0:
                    bilan_28.is_null_1 = True
                else:
                    bilan_29.is_null_1 = False
                #حقوق عند الزبائن
                bilan_30 = rec.bilan_id.filtered(lambda r: r.sequence == 26)
                actif_5 = rec.actif_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                actif1_5 = rec.actif1_id.actif_lines.filtered(lambda r: r.rubrique.sequence == 20)
                var_3 = rec.var_ids.filtered(lambda r: r.sequence == 3)
                bilan_30.write({'year_4': actif_5.montant_n,
                                'year_3': actif_5.montant_n1,
                                'year_2': actif1_5.montant_n,
                                'year_1': actif1_5.montant_n1,
                                })
                var_3.write({'montant': actif_5.montant_n})

                # المخزون
                bilan_31 = rec.bilan_id.filtered(lambda r: r.sequence == 27)
                bilan_31.write({'year_4': actif_4.montant_n,
                                'year_3': actif_4.montant_n1,
                                'year_2': actif1_4.montant_n,
                                'year_1': actif1_4.montant_n1,
                                })
                var_4 = rec.var_ids.filtered(lambda r: r.sequence == 4)
                var_4.write({'montant': actif_4.montant_n})

                #متوسط دوران المخزون (يوم)
                bilan_32 = rec.bilan_id.filtered(lambda r: r.sequence == 28)
                tcr_5 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 12)
                tcr1_5 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 12)
                tcr_6 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 13)
                tcr1_6 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 13)
                tcr_14 = rec.tcr_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 14)
                tcr1_14 = rec.tcr1_id.tcr_lines.filtered(lambda r: r.rubrique.sequence == 14)
                var_2 = rec.var_ids.filtered(lambda r: r.sequence == 2)
                bilan_32.write({'year_4': (actif_4.montant_n * 360) / (tcr_5.montant_n + tcr_6.montant_n + tcr_14.montant_n) if (tcr_5.montant_n + tcr_6.montant_n + tcr_14.montant_n) != 0 else 0,
                                'year_3': (actif_4.montant_n1 * 360) / (tcr_5.montant_n1 + tcr_6.montant_n1 + tcr_14.montant_n1) if (tcr_5.montant_n1 + tcr_6.montant_n1 + tcr_14.montant_n1) != 0 else 0,
                                'year_2': (actif1_4.montant_n * 360) / (tcr1_5.montant_n + tcr1_6.montant_n + tcr1_14.montant_n) if (tcr1_5.montant_n + tcr1_6.montant_n + tcr1_14.montant_n) != 0 else 0,
                                'year_1': (actif1_4.montant_n1 * 360) / (tcr1_5.montant_n1 + tcr1_6.montant_n1 + tcr1_14.montant_n1) if (tcr1_5.montant_n1 + tcr1_6.montant_n1 + tcr1_14.montant_n1) != 0 else 0,
                                })

                if (tcr_5.montant_n + tcr_6.montant_n + tcr_14.montant_n) == 0:
                    bilan_32.is_null_4 = True
                else:
                    bilan_32.is_null_4 = False
                if (tcr_5.montant_n1 + tcr_6.montant_n1 + tcr_14.montant_n1) == 0:
                    bilan_32.is_null_3 = True
                else:
                    bilan_32.is_null_3 = False
                if (tcr1_5.montant_n + tcr1_6.montant_n + tcr1_14.montant_n) == 0:
                    bilan_32.is_null_2 = True
                else:
                    bilan_32.is_null_2 = False
                if (tcr1_5.montant_n1 + tcr1_6.montant_n1 + tcr1_14.montant_n1) == 0:
                    bilan_32.is_null_1 = True
                else:
                    bilan_32.is_null_1 = False

                var_2.write({'montant': (tcr_5.montant_n + tcr_6.montant_n)})
                recap_2 = rec.recap_ids.filtered(lambda r: r.sequence == 2)
                recap_2.write({'montant': bilan_32.year_4})
                #متوسط فترة التحصيل (يوم)

                bilan_33 = rec.bilan_id.filtered(lambda r: r.sequence == 29)
                bilan_33.write({'year_4': (actif_5.montant_n * 360) / tcr_1.montant_n if tcr_1.montant_n != 0 else 0,
                                'year_3': (actif_5.montant_n1 * 360) / tcr_1.montant_n1 if tcr_1.montant_n1 != 0 else 0,
                                'year_2': (actif1_5.montant_n * 360) / tcr1_1.montant_n if tcr1_1.montant_n != 0 else 0,
                                'year_1': (actif1_5.montant_n1 * 360) / tcr1_1.montant_n1 if tcr1_1.montant_n1 != 0 else 0})
                if tcr_1.montant_n == 0:
                    bilan_33.is_null_4 = True
                else:
                    bilan_33.is_null_4 = False
                if tcr_1.montant_n1 == 0:
                    bilan_33.is_null_3 = True
                else:
                    bilan_33.is_null_3 = False
                if tcr1_1.montant_n == 0:
                    bilan_33.is_null_2 = True
                else:
                    bilan_33.is_null_2 = False
                if tcr1_1.montant_n1 == 0:
                    bilan_33.is_null_1 = True
                else:
                    bilan_33.is_null_1 = False
                recap_1 = rec.recap_ids.filtered(lambda r: r.sequence == 1)
                recap_1.write({'montant': bilan_33.year_4})

                #متوسط مدة تسهيلات الموردين (يوم)
                bilan_34 = rec.bilan_id.filtered(lambda r: r.sequence == 30)
                bilan_34.write({'year_4': (passif_7.montant_n * 360) / (tcr_5.montant_n + tcr_6.montant_n) if (tcr_5.montant_n + tcr_6.montant_n) != 0 else 0,
                                'year_3': (passif_7.montant_n1 * 360) / (tcr_5.montant_n1 + tcr_6.montant_n1) if (tcr_5.montant_n1 + tcr_6.montant_n1) != 0 else 0,
                                'year_2': (passif1_7.montant_n * 360) / (tcr1_5.montant_n + tcr1_6.montant_n) if (tcr1_5.montant_n + tcr1_6.montant_n) != 0 else 0,
                                'year_1': (passif1_7.montant_n1 * 360) / (tcr1_5.montant_n1 + tcr1_6.montant_n1) if (tcr1_5.montant_n1 + tcr1_6.montant_n1) != 0 else 0,
                                })
                if (tcr_5.montant_n + tcr_6.montant_n) == 0:
                    bilan_34.is_null_4 = True
                else:
                    bilan_34.is_null_4 = False
                if (tcr_5.montant_n1 + tcr_6.montant_n1) == 0:
                    bilan_34.is_null_3 = True
                else:
                    bilan_34.is_null_3 = False
                if (tcr1_5.montant_n + tcr1_6.montant_n) == 0:
                    bilan_34.is_null_2 = True
                else:
                    bilan_34.is_null_2 = False
                if (tcr1_5.montant_n1 + tcr1_6.montant_n1) == 0:
                    bilan_34.is_null_1 = True
                else:
                    bilan_34.is_null_1 = False

                recap_3 = rec.recap_ids.filtered(lambda r: r.sequence == 3)
                recap_3.write({'montant': bilan_34.year_4})
                recap_4 = rec.recap_ids.filtered(lambda r: r.sequence == 4)
                recap_4.write({'montant': (bilan_10.year_4 * 360) / tcr_1.montant_n if tcr_1.montant_n != 0 else 0})
                recap_5 = rec.recap_ids.filtered(lambda r: r.sequence == 5)
                recap_5.write({'montant': passif_5.montant_n})
                recap_to_delete = rec.recap_ids.filtered(lambda r: r.sequence in [6, 7])
                recap_to_delete.unlink()

    def create_viz2(self):
        for rec in self:
            line1 = rec.bilan_id.filtered(lambda l: l.sequence == 23)
            line2 = rec.bilan_id.filtered(lambda l: l.sequence == 21)
            line3 = rec.bilan_id.filtered(lambda l: l.sequence == 22)
            data1 = [line1.year_1, line1.year_2, line1.year_3, line1.year_4]
            data2 = [line2.year_1, line2.year_2, line2.year_3, line2.year_4]
            data3 = [line3.year_1, line3.year_2, line3.year_3, line3.year_4]

            label1 = 'NRC'
            label2 = 'CA'
            year = ["N-3", "N-2", "N-1", "N"]
            fig, ax = plt.subplots()
            width = 0.12
            X_axis = np.arange(len(year))
            rects1 = ax.bar(X_axis - width, data1, width, color="yellow", label=label1)
            rects2 = ax.bar(X_axis, data2, width, color="orange", label=label2)
            rects3 = ax.bar(X_axis + width, data3, width, color="red", label="EBITDA")
            ax.set_ylabel('Montant')
            ax.set_title('Montant par année')
            ax.set_xticks(X_axis + width, year)
            ax.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))
            fig.tight_layout()
            buf = BytesIO()
            plt.savefig(buf, format='jpeg', dpi=100)
            buf.seek(0)
            rec.visualisation1 = base64.b64encode(buf.getvalue())
            buf.close()

    def action_create_risk(self):
        for rec in self:
            view_id = self.env.ref('dept_wk.view_risk_scoring_form').id
            return {
                'name': 'Risk Scoring',
                'domain': [('parent_id', '=', rec.workflow.id)],
                'res_model': 'risk.scoring',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'default_parent_id': rec.workflow.id,
                            'parent_id': rec.workflow.id,
                            'default_partner_id': rec.workflow.nom_client.id,
                            'default_tcr_id': rec.tcr_id.id,
                            'default_actif_id': rec.actif_id.id,
                            'default_passif_id': rec.passif_id.id}
            }

    def action_create_invest(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.tcr_analysis_import_view_form').id
            return {
                'name': "Analyse d'investissement",
                'res_model': 'tcr.analysis.import',
                'view_mode': 'form',
                'view_id': view_id,
                'type': 'ir.actions.act_window',
                'context': {'default_parent_id': rec.id,
                            'etape_id': rec.id}
            }

    def create_viz(self):
        for rec in self:
            etape = rec.workflow.states.filtered(lambda l: l.sequence == 1)
            line1 = etape.situations_fin.filtered(lambda l: l.sequence == 4)
            line2 = etape.situations_fin.filtered(lambda l: l.sequence == 3)
            data1 = [line1.year3, line1.year2, line1.year1]
            data2 = [line2.year3, line2.year2, line2.year1]
            rec.visualisation = view_viz(data1, data2)

    @api.depends('tailles')
    def compute_montant_demande(self):
        for rec in self:
            etape1 = rec.workflow.states.filtered(lambda l: l.sequence ==1)
            montant = 0
            if etape1.tailles:
                montant = sum(etape1.tailles.mapped('montant'))
            rec.montant_demande = montant

    @api.onchange('risque_central')
    def compute_risk(self):
        for rec in self:
            if rec.risque_central:
                total = rec.risque_central.filtered(lambda r: r.declaration == 'الاجمالي')
                items = rec.risque_central.filtered(lambda r: r.declaration != 'الاجمالي')
                total.montant_esalam_dz_donne = sum(items.mapped('montant_esalam_dz_donne'))
                total.montant_esalam_dollar_donne = sum(items.mapped('montant_esalam_dollar_donne'))
                total.montant_esalam_dz_used = sum(items.mapped('montant_esalam_dz_used'))
                total.montant_esalam_dollar_used = sum(items.mapped('montant_esalam_dollar_used'))
                total.montant_other_dz_donne = sum(items.mapped('montant_other_dz_donne'))
                total.montant_other_dollar_donne = sum(items.mapped('montant_other_dollar_donne'))
                total.montant_other_dz_used = sum(items.mapped('montant_other_dz_used'))
                total.montant_other_dollar_used = sum(items.mapped('montant_other_dollar_used'))
                total.montant_total_dz_donne = sum(items.mapped('montant_total_dz_donne'))
                total.montant_total_dollar_donne = sum(items.mapped('montant_total_dollar_donne'))
                total.montant_total_dz_used = sum(items.mapped('montant_total_dz_used'))
                total.montant_total_dollar_used = sum(items.mapped('montant_total_dollar_used'))

    @api.depends('state_finance',
                 'state_branch',
                 'state_commercial',
                 'state_risque')
    def _compute_track(self):
        for rec in self:
            if rec.sequence == 1:
                track = self.env['wk.tracking'].search([('workflow_id', '=', rec.workflow.id),
                                                        ('etape_id', '=', rec.id),('state', '=', 'branch_1')])
                if (track and rec.state_branch == 'branch_1') or  rec.state_branch != 'branch_1':
                    track = self.env['wk.tracking'].search([('workflow_id', '=', rec.workflow.id),
                                                            ('etape_id', '=', rec.id)])[-1]
                    track.date_fin = datetime.datetime.today()
                if rec.state_branch not in ['branch_rejected', 'branch_5']:
                    self.env['wk.tracking'].create({'workflow_id': rec.workflow.id,
                                                    'etape_id': rec.id,
                                                    'state': rec.state_branch,
                                                    'date_debut': datetime.datetime.today(),
                                                    'is_revision': True if rec.raison_a_revoir else False,
                                                    'comment': rec.raison_a_revoir})
                    
            if rec.sequence == 2:
                if rec.state_finance != 'finance_1':
                    track = self.env['wk.tracking'].search([('workflow_id', '=', rec.workflow.id),
                                                            ('etape_id', '=', rec.id)])[-1]
                    track.date_fin = datetime.datetime.today()
                if rec.state_finance not in ['finance_rejected', 'finance_4']:
                    self.env['wk.tracking'].create({'workflow_id': rec.workflow.id,
                                                    'etape_id': rec.id,
                                                    'state': rec.state_finance,
                                                    'date_debut': datetime.datetime.today(),
                                                    'is_revision': True if rec.raison_a_revoir else False,
                                                    'comment': rec.raison_a_revoir})
            if rec.sequence == 3:
                if rec.state_commercial != 'commercial_1':
                    track = self.env['wk.tracking'].search([('workflow_id', '=', rec.workflow.id),
                                                            ('etape_id', '=', rec.id)])[-1]
                    track.date_fin = datetime.datetime.today()
                if rec.state_commercial not in ['commercial_4']:
                    self.env['wk.tracking'].create({'workflow_id': rec.workflow.id,
                                                    'etape_id': rec.id,
                                                    'state': rec.state_commercial,
                                                    'date_debut': datetime.datetime.today(),
                                                    'is_revision': True if rec.raison_a_revoir else False,
                                                    'comment': rec.raison_a_revoir})
            if rec.sequence == 4:
                if rec.state_risque != 'risque_1':
                    track = self.env['wk.tracking'].search([('workflow_id', '=', rec.workflow.id),
                                                            ('etape_id', '=', rec.id)])[-1]
                    track.date_fin = datetime.datetime.today()
                if rec.state_risque not in ['risque_2']:
                    self.env['wk.tracking'].create({'workflow_id': rec.workflow.id,
                                                    'etape_id': rec.id,
                                                    'state': rec.state_risque,
                                                    'date_debut': datetime.datetime.today(),
                                                    'is_revision': True if rec.raison_a_revoir else False,
                                                    'comment': rec.raison_a_revoir})

    def open_messages(self):
        for rec in self:
            view_id = self.env.ref('mail.view_message_tree').id
            return {
                'name': "Messages",
                'res_model': 'mail.message',
                'view_mode': 'tree',
                'view_id': view_id,
                'domain': [('res_id', 'in', rec.workflow.states.ids + [rec.workflow.id]),
                           ('message_type', '=', 'comment'),
                           ('model', 'in', ['wk.etape', 'wk.workflow.dashboard'])],
                'type': 'ir.actions.act_window',
            }

    def get_mail_to(self):
        for rec in self:
            partner_ids = []
            list_final = ''
            admin_group = self.env.ref('dept_wk.dept_wk_group_administration')
            if rec.sequence == 1:
                if rec.state_branch in ['branch_1', 'branch_3']:
                    partner_ids = rec.assigned_to_agence.partner_id.email
                    list_final = partner_ids
                elif rec.state_branch in ['branch_2', 'branch_4']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_agence').users.filtered(
                        lambda l: l.branche == rec.branche and admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                elif rec.state_branch in ['branch_4']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                else:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 2:
                if rec.state_finance == 'finance_3':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                elif rec.state_finance == 'finance_2':
                    partner_ids = rec.assigned_to_finance.partner_id.email
                    list_final = partner_ids
                elif rec.state_finance == 'finance_4':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_risque').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 3:
                if rec.state_commercial == 'commercial_2':
                    partner_ids = rec.assigned_to_commercial.partner_id.email
                    list_final = partner_ids
                elif rec.state_commercial in ['commercial_3']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_commercial').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                else:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 4:
                if rec.state_risque == 'risque_3':
                    partner_ids = rec.assigned_to_risque.partner_id.email
                    list_final = partner_ids
                elif rec.state_risque == 'risque_4':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_risque').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                elif rec.state_risque == 'risque_2':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_tres').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 10:
                if rec.state_tres == 'tres_2':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_dga').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 5:
                if rec.state_vice == 'vice_2':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_dg').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            return list_final

    def get_mail_to_revoir(self):
        for rec in self:
            partner_ids = []
            list_final = ''
            admin_group = self.env.ref('dept_wk.dept_wk_group_administration')
            if rec.sequence == 1:
                if rec.state_branch in ['branch_1', 'branch_3']:
                    partner_ids = rec.assigned_to_agence.partner_id.email
                    list_final = partner_ids
                elif rec.state_branch in ['branch_2','branch_4']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_agence').users.filtered(
                        lambda l: l.branche == rec.branche and admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 2:
                if rec.state_finance in  ['finance_3']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.filtered(
                        lambda l: admin_group not in l.groups_id).mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                elif rec.state_finance == 'finance_2':
                    partner_ids = rec.assigned_to_finance.partner_id.email
                    list_final = partner_ids
                elif rec.state_finance == 'finance_1':
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 3:
                if rec.state_commercial == 'commercial_2':
                    partner_ids = rec.assigned_to_commercial.partner_id.email
                    list_final = partner_ids
                elif rec.state_commercial in ['commercial_3']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_commercial').users.mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
                else:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_analyste').users.mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            if rec.sequence == 4:
                if rec.state_risque == 'risque_3':
                    partner_ids = rec.assigned_to_risque.partner_id.email
                    list_final = partner_ids
                elif rec.state_risque == ['risque_4', 'risque_1']:
                    user_ids = self.env.ref('dept_wk.dept_wk_group_responsable_risque').users.mapped('partner_id')
                    partner_ids = user_ids.mapped('email')
                    list_final = ', '.join(partner_ids)
            return list_final

def check_if_xls_file(file):
    if file:
        mime = magic.Magic(mime=True)
        mime_type = mime.from_buffer(file)

        if mime_type == 'application/vnd.ms-excel' or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            # Do something if the file is XLS
            return True
    return False

def view_viz(data1, data2):
    year = ["N-2", "N-1", "N"]
    fig, ax = plt.subplots()
    width = 0.25
    X_axis = np.arange(len(year))
    label1 = 'NRC'
    label2 = 'CA'
    plt.rcParams['font.family'] = 'DejaVu Sans'
    rects1 = ax.bar(X_axis - (width / 2), data1, width, color="yellow", label=label1)
    rects2 = ax.bar(X_axis + (width / 2), data2, width, color="orange", label=label2)
    ax.set_ylabel('Montant')
    ax.set_title('Montant par année')
    ax.set_xticks(X_axis + width, year)
    ax.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))
    fig.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='jpeg', dpi=100)
    buf.seek(0)
    imageBase64 = base64.b64encode(buf.getvalue())
    buf.close()
    return imageBase64
