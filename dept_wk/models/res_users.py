from odoo import models, api
from openpyxl import load_workbook
import os

class ResUsers(models.Model):
    
    _inherit = 'res.users'
    
    @api.model
    def assign_arabic_names_from_excel(self):
        print('##########################################################################')
        module_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        excel_path = os.path.join(module_path, 'static/files', 'user.xlsx')

        print('Loading .................')
        try:
            workbook = load_workbook(excel_path)
            sheet = workbook.active 
        except Exception as e:
            print(f"Failed to read Excel file: {str(e)}")
            return False

        print('########### Iteration #############')
        for row in sheet.iter_rows(min_row=2, values_only=True):
            arabic_name = row[0]  
            email = row[1] 

            if not email or not arabic_name:
                print(f"Missing data. Email or Arabic name is missing.")
                continue 

            print('Search by email .................')
            user = self.search([('email', '=', email)], limit=1)

            if user:
                try:
                    user.partner_id.write({'nom_arabe': arabic_name})
                    print(f"Updated user {email} with Arabic name: {arabic_name}")
                except Exception as e:
                    print(f"Failed to update user {email}: {str(e)}")
            else:
                print(f"*** The User with email: {email} doesn't exist ***")

        return True        
