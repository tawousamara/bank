from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
import xlsxwriter
import openpyxl

import numpy as np
import matplotlib.pyplot as plt

ACTIVE_LIST = [
    ('1', "Immobilisations incorporelles"),
    ('2', "Immobilisations corporelles"),
    ('3', 'Terrain'),
    ('4', 'Batiment'),
    ('5', 'Autres immobilisations corporelles'),
    ('6', 'Immobilisations en cocession'),
    ('7', "Immobilisations en cours"),
    ('8', 'Immobilisations financières'),
    ('9', 'Titres mis en équivalence'),
    ('10', 'Autres participations'),
    ('11', 'Autres titres immobilisés'),
    ('12', 'Prêts et autres actifs financiers non courant'),
    ('13', 'Impôts différés actif'),
    ('14', 'Total Actif Non-Courant'),
    ('15', 'Stock et encours'),
    ('16', 'Créances et emplois assimilés'),
    ('17', 'Clients'),
    ('18', "Autres débiteurs"),
    ('19', "Impôts et assimilés"),
    ('20', 'Autres créances et impôts assimilés'),
    ('21', 'Disponibilités et Assimilés'),
    ('22', "Placement et autres actif financier"),
    ('23', 'Trésorerie'),
    ('24', 'Total Actif Courant'),
    ('25', 'Total General Actif'),
]
PASSIVE_LIST = [
    ('1', "Capital émis"),
    ('2', "Capital non appelé"),
    ('3', 'Prime et réserve'),
    ('4', 'Écart de réévaluation'),
    ('5', 'Écart d`équivalence'),
    ('6', 'Résultat net'),
    ('7', "Report à nouveau"),
    ('8', 'Total Capitaux Propres'),
    ('9', 'Emprunts et dette financière'),
    ('10', 'Impôts différés'),
    ('11', 'Autre dette non courant'),
    ('12', "Provisions et produits constatés d'avance"),
    ('13', 'Total Passif Non Courant'),
    ('14', 'Fournisseurs et comptes rattachés'),
    ('15', 'Impôts'),
    ('16', 'Autres dettes'),
    ('17', 'Trésorerie passif'),
    ('18', "Total Passif Courant"),
    ('19', "Total General Passif"),
]

RATIO_LIST = [
    ("1", "TB"),
    ("2", "FP/TB"),
    ("3", "BFR"),
    ("4", "BFR/CA%"),
    ("5", "BFR en jours CA"),
    ("6", "Créance en jours CA"),
    ("7", "Stock en jours d'achat"),
    ("8", "Levier"),
    ("9", "Liquidité Rapide"),
]


class BilanGeneral(models.Model):
    _name = 'bilan.general'

    name = fields.Char(string="Reference")
    date = fields.Date(string="Date")
    line_passif_ids = fields.One2many('bilan.passif', string='lignes', inverse_name='bilan_id')
    line_actif_ids = fields.One2many('bilan.actif', string='lignes', inverse_name='bilan_id')
    line_ratio_ids = fields.One2many('bilan.ratio', string='lignes', inverse_name='bilan_id')
    tcr_id = fields.Many2one('tcr.analysis.import', string="TCR")
    file_template = fields.Binary(string='Modèle Excel', compute='compute_template')
    file_template_name = fields.Char(string='fichier', default='Télécharger le modèle Excel')

    file_import_name = fields.Char(string='fichier', default='Importer le fichier Excel')
    file_import_data = fields.Binary(string='Importer le fichier Excel')

    year_prec = fields.Selection([('3', 'N-3'), ('2', 'N-2'), ('1', 'N-1'), ('0', 'N')], string='Année')
    graph_pie = fields.Binary(string='Graph', compute='compute_pie')
    graph_bar = fields.Binary(string='Graph')

    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)

    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('bilan.general.seq')
        return super(BilanGeneral, self).create(vals)

    def compute_template(self):
        result = BytesIO()
        workbook = xlsxwriter.Workbook(result)
        worksheet = workbook.add_worksheet("Actif")
        bold = workbook.add_format({'bold': True})
        worksheet.write(0, 0, 'Poste Comptable', bold)
        worksheet.write(0, 1, 'N-3', bold)
        worksheet.write(0, 2, 'N-2', bold)
        worksheet.write(0, 3, 'N-1', bold)
        worksheet.write(0, 4, 'N', bold)

        for index, entry in ACTIVE_LIST:
            bold_list = [2, 8, 14, 16, 21, 24, 25]
            if int(index) in bold_list:
                worksheet.write(int(index), 0, entry, bold)
            else:
                worksheet.write(int(index), 0, entry)
        worksheet = workbook.add_worksheet("Passif")
        bold = workbook.add_format({'bold': True})
        worksheet.write(0, 0, 'Poste Comptable', bold)
        worksheet.write(0, 1, 'N-3', bold)
        worksheet.write(0, 2, 'N-2', bold)
        worksheet.write(0, 3, 'N-1', bold)
        worksheet.write(0, 4, 'N', bold)

        for index, entry in PASSIVE_LIST:
            bold_list = [8, 13, 18, 19]
            if int(index) in bold_list:
                worksheet.write(int(index), 0, entry, bold)
            else:
                worksheet.write(int(index), 0, entry)

        workbook.close()
        buf = base64.b64encode(result.getvalue())
        self.write({'file_template': buf})
        result.close()

    def action_import_data(self):
        self.ensure_one()
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file_import_data)), read_only=True)

        self.env['bilan.actif'].search([('bilan_id', '=', self.id)]).unlink()
        self.env['bilan.passif'].search([('bilan_id', '=', self.id)]).unlink()
        count_sheet = 0
        for ws in wb:
            print('here')
            count_row = 0
            count_sheet += 1
            print(count_sheet)
            if count_sheet == 1:
                lines_ids = self.env['bilan.actif']
            else:
                lines_ids = self.env['bilan.passif']
            for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
                count_row += 1
                print(get_value(record[0]))
                lines_ids.create({
                    'bilan_id': self.id,
                    'poste_comptable': get_value(record[0], list_index=count_sheet) if get_value(record[0],
                                                                                                 list_index=count_sheet) != "" else str(
                        count_row),
                    'amount_n3': record[1],
                    'amount_n2': record[2],
                    'amount_n1': record[3],
                    'amount_n': record[4],
                })

    def action_count_ratio(self):
        print("clicked")
        for rec in self:
            self.line_ratio_ids.unlink()
            ca = self.tcr_id.line_ids.filtered(lambda r: r.poste_comptable == '1')
            tb = self.line_actif_ids.filtered(lambda r: r.poste_comptable == '25')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "1",
                                            "amount_n3": tb.amount_n3,
                                            "amount_n2": tb.amount_n2,
                                            "amount_n1": tb.amount_n1,
                                            "amount_n": tb.amount_n})
            tb = self.line_passif_ids.filtered(lambda r: r.poste_comptable == '19')
            fp = self.line_passif_ids.filtered(lambda r: r.poste_comptable == '8')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "2",
                                            "amount_n3": fp.amount_n3 / tb.amount_n3 if tb.amount_n3 > 0 else 0,
                                            "amount_n2": fp.amount_n2 / tb.amount_n2 if tb.amount_n2 > 0 else 0,
                                            "amount_n1": fp.amount_n1 / tb.amount_n1 if tb.amount_n1 > 0 else 0,
                                            "amount_n": fp.amount_n / tb.amount_n if tb.amount_n > 0 else 0})
            bfr_pass = self.line_passif_ids.filtered(lambda r: r.poste_comptable == '14')
            bfr_act = self.line_actif_ids.filtered(lambda r: r.poste_comptable in ['15', '16'])
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "3",
                                            "amount_n3": sum(bfr_act.mapped("amount_n3")) - bfr_pass.amount_n3,
                                            "amount_n2": sum(bfr_act.mapped("amount_n2")) - bfr_pass.amount_n2,
                                            "amount_n1": sum(bfr_act.mapped("amount_n1")) - bfr_pass.amount_n1,
                                            "amount_n": sum(bfr_act.mapped("amount_n")) - bfr_pass.amount_n})
            bfr = self.line_ratio_ids.filtered(lambda r: r.poste_comptable == '3')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "4",
                                            "amount_n3": bfr.amount_n3 / ca.amount_n3 if ca.amount_n3 > 0 else 0,
                                            "amount_n2": bfr.amount_n2 / ca.amount_n2 if ca.amount_n2 > 0 else 0,
                                            "amount_n1": bfr.amount_n1 / ca.amount_n1 if ca.amount_n1 > 0 else 0,
                                            "amount_n": bfr.amount_n / ca.amount_n if ca.amount_n > 0 else 0})
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "5",
                                            "amount_n3": bfr.amount_n3 * 360 / ca.amount_n3 if ca.amount_n3 > 0 else 0,
                                            "amount_n2": bfr.amount_n2 * 360 / ca.amount_n2 if ca.amount_n2 > 0 else 0,
                                            "amount_n1": bfr.amount_n1 * 360 / ca.amount_n1 if ca.amount_n1 > 0 else 0,
                                            "amount_n": bfr.amount_n * 360 / ca.amount_n if ca.amount_n > 0 else 0})
            creance = self.line_actif_ids.filtered(lambda r: r.poste_comptable == '17')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "6",
                                            "amount_n3": creance.amount_n3 * 360 / ca.amount_n3 if ca.amount_n3 > 0 else 0,
                                            "amount_n2": creance.amount_n2 * 360 / ca.amount_n2 if ca.amount_n2 > 0 else 0,
                                            "amount_n1": creance.amount_n1 * 360 / ca.amount_n1 if ca.amount_n1 > 0 else 0,
                                            "amount_n": creance.amount_n * 360 / ca.amount_n if ca.amount_n > 0 else 0})
            stck = self.line_actif_ids.filtered(lambda r: r.poste_comptable == '15')
            achat = self.tcr_id.line_ids.filtered(lambda r: r.poste_comptable == '8')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "7",
                                            "amount_n3": stck.amount_n3 * 360 / achat.amount_n3 if achat.amount_n3 > 0 else 0,
                                            "amount_n2": stck.amount_n2 * 360 / achat.amount_n2 if achat.amount_n2 > 0 else 0,
                                            "amount_n1": stck.amount_n1 * 360 / achat.amount_n1 if achat.amount_n1 > 0 else 0,
                                            "amount_n": stck.amount_n * 360 / achat.amount_n if achat.amount_n > 0 else 0})
            passif = self.line_passif_ids.filtered(lambda r: r.poste_comptable in ['9', '17'])
            passif_neg = self.line_passif_ids.filtered(lambda r: r.poste_comptable == '8')
            actif = self.line_actif_ids.filtered(lambda r: r.poste_comptable == '21')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "8",
                                            "amount_n3": (sum(passif.mapped(
                                                'amount_n3')) - actif.amount_n3) / passif_neg.amount_n3 if passif_neg.amount_n3 > 0 else 0,
                                            "amount_n2": (sum(passif.mapped(
                                                'amount_n2')) - actif.amount_n2) / passif_neg.amount_n2 if passif_neg.amount_n2 > 0 else 0,
                                            "amount_n1": (sum(passif.mapped(
                                                'amount_n1')) - actif.amount_n1) / passif_neg.amount_n1 if passif_neg.amount_n1 > 0 else 0,
                                            "amount_n": (sum(passif.mapped(
                                                'amount_n')) - actif.amount_n) / passif_neg.amount_n if passif_neg.amount_n > 0 else 0})
            act_st = self.line_actif_ids.filtered(lambda r: r.poste_comptable == '15')
            act_nd = self.line_actif_ids.filtered(lambda r: r.poste_comptable == '24')
            passif = self.line_passif_ids.filtered(lambda r: r.poste_comptable == '18')
            self.env['bilan.ratio'].create({"bilan_id": self.id,
                                            "poste_comptable": "9",
                                            "amount_n3": (
                                                                 act_nd.amount_n3 - act_st.amount_n3) / passif.amount_n3 if passif.amount_n3 > 0 else 0,
                                            "amount_n2": (
                                                                 act_nd.amount_n2 - act_st.amount_n2) / passif.amount_n2 if passif.amount_n2 > 0 else 0,
                                            "amount_n1": (
                                                                 act_nd.amount_n1 - act_st.amount_n1) / passif.amount_n1 if passif.amount_n1 > 0 else 0,
                                            "amount_n": (
                                                                act_nd.amount_n - act_st.amount_n) / passif.amount_n if passif.amount_n > 0 else 0})
            data1 = [ca.amount_n3, ca.amount_n2, ca.amount_n1, ca.amount_n]
            data2 = [bfr.amount_n3, bfr.amount_n2, bfr.amount_n1, bfr.amount_n]
            data_list = [data1, data2]
            print(data_list)
            self.graph_bar = create_bar(data_list)
            self.graph_pie = create_bar(data_list)

    @api.depends("year_prec")
    def compute_pie(self):
        if self.year_prec:
            labels = ["Stock", "Créances", "Clients"]
            if self.line_actif_ids:
                data = get_data(self.line_actif_ids, type_class=2)
                data_list = [list(data[14].values())[int(self.year_prec) + 1],
                             list(data[15].values())[int(self.year_prec) + 1],
                             list(data[16].values())[int(self.year_prec) + 1]]
            print(data_list)
            self.graph_pie = create_pie(data_list, labels)
        else:
            self.graph_pie = False

    @api.onchange("year_prec")
    def compute_pie(self):
        for rec in self:
            if self.year_prec:
                labels = ["Stock", "Créances", "Clients"]
                data = get_data(self.line_actif_ids, type_class=2)
                data_list = [list(data[14].values())[int(self.year_prec) + 1],
                             list(data[15].values())[int(self.year_prec) + 1],
                             list(data[16].values())[int(self.year_prec) + 1]]
                print(data_list)
                self.graph_pie = create_pie(data_list, labels)
            else:
                self.graph_pie = False


class BilanActif(models.Model):
    _name = 'bilan.actif'

    poste_comptable = fields.Selection(ACTIVE_LIST, string='Poste Comptable')
    amount_n = fields.Float(string='N')
    amount_n1 = fields.Float(string='N-1')
    amount_n2 = fields.Float(string='N-2')
    amount_n3 = fields.Float(string='N-3')
    bilan_id = fields.Many2one('bilan.general', string='bilan')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)


class BilanPassif(models.Model):
    _name = 'bilan.passif'

    poste_comptable = fields.Selection(PASSIVE_LIST, string='Poste Comptable')
    amount_n = fields.Float(string='N')
    amount_n1 = fields.Float(string='N-1')
    amount_n2 = fields.Float(string='N-2')
    amount_n3 = fields.Float(string='N-3')
    bilan_id = fields.Many2one('bilan.general', string='bilan')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)


class BilanRatio(models.Model):
    _name = 'bilan.ratio'

    poste_comptable = fields.Selection(RATIO_LIST, string='Poste Comptable')
    amount_n = fields.Float(string='N')
    amount_n1 = fields.Float(string='N-1')
    amount_n2 = fields.Float(string='N-2')
    amount_n3 = fields.Float(string='N-3')
    bilan_id = fields.Many2one('bilan.general', string='bilan')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)


def get_value(value, list_index=1):
    data_get = ''
    if list_index == 1:
        for index, entry in ACTIVE_LIST:
            if entry == value:
                data_get = index
    else:
        for index, entry in PASSIVE_LIST:
            if entry == value:
                data_get = index
    return data_get


def create_pie(data, labels):
    sizes = data
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    buf = BytesIO()
    plt.savefig(buf, format='jpeg', dpi=100)
    buf.seek(0)
    imageBase64 = base64.b64encode(buf.getvalue())
    buf.close()
    return imageBase64


def create_bar(data):
    data1 = data[0]
    data2 = data[1]
    year = ["N-3", "N-2", "N-1", "N"]
    fig, ax = plt.subplots()
    width = 0.5
    rects1 = ax.bar(year, data1, width, color="green", label="Chiffre d'affaire")
    rects2 = ax.bar(year, data2, width, color="yellow", bottom=np.array(data1), label="BFR")

    ax.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))
    fig.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='jpeg', dpi=100)
    buf.seek(0)
    imageBase64 = base64.b64encode(buf.getvalue())
    buf.close()
    return imageBase64


def get_data(data, type_class=1):
    recordset = []
    if type_class != 1:
        for i in data:
            element = {
                'poste_comptable': ACTIVE_LIST[int(i.poste_comptable) - 1][1],
                'amount_n': i.amount_n,
                'amount_n1': i.amount_n1,
                'amount_n2': i.amount_n2,
                'amount_n3': i.amount_n3,
            }
            recordset.append(element)
    else:
        for i in data:
            element = {
                'poste_comptable': PASSIVE_LIST[int(i.poste_comptable) - 1][1],
                'amount_n': i.amount_n,
                'amount_n1': i.amount_n1,
                'amount_n2': i.amount_n2,
                'amount_n3': i.amount_n3,
            }
            recordset.append(element)
    return recordset
