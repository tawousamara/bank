import base64
from odoo import models, fields, api, _
from datetime import datetime
from odoo.exceptions import ValidationError, UserError
import openpyxl
import re
import json
import requests

class ImportPassifOCR(models.Model):
    _name = "import.ocr.passif"
    _description = "Import Bilan Passif Data by OCR Functionality"

    name = fields.Char(string="Réf")
    date = fields.Date(string="Date d'importation", default=datetime.today())
    annee = fields.Char(string="Année de l'exercice")
    company = fields.Char(string="Désignation de l'entreprise")
    passif_lines = fields.One2many("import.ocr.passif.line", "passif_id", string="Lignes", domain=lambda self: self._get_domain())
    file_import = fields.Binary(string="Import de fichier")
    file_import_name = fields.Char(string="Fichier")
    hide_others = fields.Boolean(string="Filter que les lignes concernées")
    state = fields.Selection([("get_data", "Import données"),
                              ("validation", "Validation"),
                              ("valide", "Validé"),
                              ('modified', 'Modifié par le risque')], string="Etat", default="get_data")

    def _get_domain(self):
        if self.hide_others:
            return [('sequence', 'in', [2, 4, 8, 12, 14, 18, 20, 21, 22, 23, 24, 25])]
        else:
            []

    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('import.ocr.passif.seq')
        return super(ImportPassifOCR, self).create(vals)

    def process_import_passif_file(self, col):
        print("Processing import passif OCR")
        passif_lines = []
        try:
            file_content = base64.b64decode(self.file_import)
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            column_idx = openpyxl.utils.column_index_from_string(col.upper())

            if column_idx > sheet.max_column:
                raise UserError(f"Column {col} is out of range in the file.")

            sequences = [2, 4, 8, 12, 14, 18, 20, 21, 22, 23, 24, 25]

            for position, sequence in enumerate(sequences):
                row_idx = position + 23
                name = sheet.cell(row=row_idx, column=1).value
                montant_n = sheet.cell(row=row_idx, column=column_idx).value

                montant_n1 = None
                if col.upper() != 'E':
                    montant_n1 = sheet.cell(row=row_idx, column=column_idx + 1).value

                if montant_n is None:
                    montant_n = 0
                    print(f"Montant (N) was None, set to 0 for row {row_idx}.")

                if montant_n1 is None:
                    montant_n1 = 0
                    print(f"Montant (N-1) was None, set to 0 for row {row_idx}.")

                if isinstance(montant_n, float):
                    montant_n = int(montant_n)
                    print(f"Converted montant_n to integer: {montant_n}")

                if montant_n1 is not None and isinstance(montant_n1, float):
                    montant_n1 = int(montant_n1)
                    print(f"Converted montant_n1 to integer: {montant_n1}")

                rubrique_vals = {
                    'name': name,
                    'type': 'passif',
                    'sequence': sequence
                }
                rubrique = self.env['import.ocr.config'].create(rubrique_vals)

                passif_line_vals = {
                    'name': name,
                    'montant_n': montant_n,
                    'passif_id': self.id,
                    'rubrique': rubrique.id
                }

                if montant_n1 is not None:
                    passif_line_vals['montant_n1'] = montant_n1

                new_line = self.env['import.ocr.passif.line'].create(passif_line_vals)
                passif_lines.append(new_line)

            self.passif_lines = [(6, 0, [line.id for line in passif_lines])]
            print("here are passif_lines from self")
            print(self.passif_lines)
            print("here are passif_lines")
            print(passif_lines)

        except Exception as e:
            print(f"Error processing file: {e}")


    def open_file(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.extract_bilan_wizard_form').id
            context = dict(self.env.context or {})
            context['pdf_1'] = rec.file_import
            context['passif_id'] = rec.id
            wizard = self.env['extract.bilan.wizard'].create({'pdf_1': rec.file_import})
            return {
                'name': 'Passif',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'extract.bilan.wizard',
                'res_id': wizard.id,
                'view_id': view_id,
                'target': 'new',
                'context': context,
            }
    def extract_data(self):
        for rec in self:
            if rec.file_import:
                pattern_alpha = r'^[a-zA-Z\séèàôâê\'();,*+-1]+$'
                pattern_num = r'^[0-9\s()\-]+$'
                if rec.passif_lines:
                    rec.passif_lines.unlink()
                data = str(rec.file_import)
                data = data.replace("b'", '\n')
                data = data.replace("'", '')
                data = data.replace('\r\n', '\n')  # Replace Windows-style newline with Unix-style
                data = data.replace('\r', '\n')
                data = 'data:application/pdf;base64,' + data
                test_file = ocr_space_file(filename=data, api_key='K87496787788957', language='fre',
                                           isTable=True)
                json_dumps = test_file.content.decode()
                print(test_file.status_code)
                print(json_dumps)
                if test_file.status_code == 200:
                    json_loads = json.loads(json_dumps)
                    if json_loads['IsErroredOnProcessing'] == False:
                        lines = json_loads['ParsedResults'][0]['TextOverlay']['Lines']
                        lines = group_words_by_line(lines)
                        list_tcr = self.env['import.ocr.config'].search([('type', '=', 'passif')])
                        for line in lines:
                            rubrique = self.env['import.ocr.config'].search([('name', '=', line['Words'][0]['WordText'])])
                            value_text = line['Words'][0]['WordText']
                            if not rubrique:
                                print(value_text)
                                value_text = value_text.replace(';', '')
                                value_text = value_text.replace(',', '')
                                value_text = value_text.replace('-', '')
                                for i in list_tcr:
                                    val = i.name
                                    val = val.replace(';', '')
                                    val = val.replace(',', '')
                                    val = val.replace('-', '')
                                    if value_text == val:
                                        print(value_text == val)
                            if rubrique:
                                value = rec.env['import.ocr.passif.line'].create({'passif_id': rec.id,
                                                                               'name': line['Words'][0]['WordText'],
                                                                               'rubrique': rubrique.id,
                                                                               })
                                if len(line['Words']) == 3:
                                    try:
                                        value.write({'montant_n': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                    except:
                                        value.write({'montant_n': 0})
                                    try:
                                        value.write({'montant_n1': int(re.sub(r'[^0-9]', '', line['Words'][2]['WordText']))})
                                    except:
                                        value.write({'montant_n1': 0})

                                elif len(line['Words']) == 2:
                                    separator = line['Words'][1]['Left'] - line['Words'][0]['Left']
                                    if separator > 400:
                                        try:
                                            value.write(
                                                {'montant_n1': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                        except:
                                            value.write({'montant_n1': 0})
                                    else:
                                        try:
                                            value.write({'montant_n': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                        except:
                                            value.write({'montant_n': 0})
                        rec.state = "validation"
                        for line in rec.passif_lines:
                            line.montant_n = line.montant_n / 1000
                            line.montant_n1 = line.montant_n1 / 1000
                    else:
                        raise UserError('Vous devriez verifier la qualité, le nombre et la taille du fichier \n Le fichier ne doit pas dépasser 1024KB. \n Le fichier doit contenir une seule page.')
                else:
                    raise UserError('Un probleme est survenu, vous devriez réessayer ulterieurement.')

    def action_validation(self):
        for rec in self:
            list_validation = [12, 14, 20, 23, 24, 25]
            passifs = rec.passif_lines.filtered(lambda r: r.rubrique.sequence in list_validation)
            if len(passifs) != 6:
                raise ValidationError("Vous devriez confirmer les valeurs suivantes: \n"
                                      "- Total I \n"
                                      "- Emprunts et dettes financières \n"
                                      "- Fournisseurs et comptes rattachés \n"
                                      "- Trésorerie passifs \n"
                                      "- Total III \n"
                                      "- Total General Passif (I+II+III)")
            view_id = self.env.ref('financial_modeling.confirmation_wizard_form')
            context = dict(self.env.context or {})
            context['passif_id'] = rec.id
            context['state'] = 'valide'
            print(context)
            if not self._context.get('warning'):
                return {
                    'name': 'Validation',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'import.ocr.wizard',
                    'view_id': view_id.id,
                    'target': 'new',
                    'context': context,
                }

    def action_annulation(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.confirmation_wizard_form')
            context = dict(self.env.context or {})
            context['passif_id'] = rec.id
            context['state'] = 'validation'
            print(context)
            if not self._context.get('warning'):
                return {
                    'name': 'Annulation',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'import.ocr.wizard',
                    'view_id': view_id.id,
                    'target': 'new',
                    'context': context,
                }
def ocr_space_file(filename, overlay=True, api_key='helloworld', language='eng',isTable=False):
    payload = {'isOverlayRequired': overlay,
               'apikey': api_key,
               'language': language,
               'base64Image': filename,
               'isTable': isTable,
               'OCREngine': 2
               }
    print('called')
    r = requests.post('https://api.ocr.space/parse/image',
                          data=payload,
                          )
    return r

class ImportPassifOcrLine(models.Model):
    _name = "import.ocr.passif.line"
    _description = "Line de bilan passif importé"

    name = fields.Char(string="RUBRIQUES")
    mintop = fields.Integer(string='Rang')
    height = fields.Integer(string='Height')
    sequence = fields.Integer(related='rubrique.sequence')
    rubrique = fields.Many2one('import.ocr.config', string='Rubriques confirmés', domain="[('type','=','passif')]")
    montant_n = fields.Float(string="N")
    montant_n1 = fields.Float(string="N-1")
    passif_id = fields.Many2one('import.ocr.passif', string="Passif ID")

def assign_amounts(actif, amounts, intervals):
    for amount in amounts:
        if intervals[0][0] <= amount['left'] <= intervals[0][-1]:
            actif.montant_n = amount['amount']
        elif intervals[1][0] <= amount['left'] <= intervals[1][-1]:
            actif.montant_n1 = amount['amount']
        else:
            if intervals[0][0] <= amount['left'] - amount['left'] <= intervals[0][-1]:
                actif.montant_n = amount['amount']
            elif intervals[1][0] <= amount['left'] - amount['left'] <= intervals[1][-1]:
                actif.montant_n1 = amount['amount']


def group_words_by_line(json_data):
    # Trier les mots par leur position verticale (Top)
    sorted_words = sorted(json_data, key=lambda x: x['MinTop'])

    lines = []
    current_line = {'LineText': '', 'Left': None, 'Words': []}
    previous_top = None
    previous_max_height = None

    for item in sorted_words:
        top = item['MinTop']
        max_height = item['MaxHeight']
        word_text = item['LineText']
        left = item['Words'][0]['Left']  # Get the Left position of the first word in this line

        if previous_top is None:
            current_line['LineText'] = word_text
            current_line['Left'] = left
            current_line['Words'].append({'WordText': word_text, 'Left': left})
            previous_top = top
            previous_max_height = max_height
        elif top >= previous_top and top <= (previous_top + previous_max_height):
            current_line['LineText'] += " " + word_text
            current_line['Words'].append({'WordText': word_text, 'Left': left})
            previous_max_height = max(previous_max_height, max_height)
        else:
            # Trier les mots de la ligne par leur position horizontale (Left)
            current_line['Words'] = sorted(current_line['Words'], key=lambda x: x['Left'])
            lines.append(current_line)
            current_line = {'LineText': word_text, 'Left': left, 'Words': [{'WordText': word_text, 'Left': left}]}
            previous_top = top
            previous_max_height = max_height

    current_line['Words'] = sorted(current_line['Words'], key=lambda x: x['Left'])
    lines.append(current_line)  # Append the last line

    return lines
