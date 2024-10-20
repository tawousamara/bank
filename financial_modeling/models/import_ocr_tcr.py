from odoo.exceptions import UserError
from odoo import models, fields, api, _
from datetime import datetime
from odoo.exceptions import ValidationError
import re
import json
import requests
import io
import PyPDF2
import base64
from PIL import Image
import openpyxl
class ImportTcrOCR(models.Model):
    _name = 'import.ocr.tcr'
    _description = "Import Tcr Data by OCR Functionality"

    name = fields.Char(string="Réf")
    date = fields.Date(string="Date d'importation", default=datetime.today())
    annee = fields.Char(string="Année de l'exercice")
    company = fields.Char(string="Désignation de l'entreprise")
    tcr_lines = fields.One2many("import.ocr.tcr.line", "tcr_id", string="Lignes", domain=lambda self: self._get_domain())
    file_import = fields.Binary(string="Import de fichier")
    file_import2 = fields.Binary(string="Import de fichier")
    file_import_name = fields.Char(string="Fichier")
    hide_others = fields.Boolean(string="Filter que les lignes concernées")
    state = fields.Selection([("get_data", "Import données"),
                              ("validation", "Validation"),
                              ("valide", "Validé"),
                              ('modified', 'Modifié par le risque')], string="Etat", default="get_data")

    def _get_domain(self):
        if self.hide_others:
            return [('sequence', 'in', [7, 33, 50, 36, 12, 13, 14, 30])]
        else:
            []

    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('import.ocr.tcr.seq')
        return super(ImportTcrOCR, self).create(vals)
    
    # marwa##########################################################################
    def process_import_tcr_file(self, col):
        print("Processing import TCR OCR")
        tcr_lines = []

        try:

            file_content = base64.b64decode(self.file_import)
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            column_idx = openpyxl.utils.column_index_from_string(col.upper())

            if column_idx > sheet.max_column:
                raise UserError(f"Column {col} is out of range in the file.")

            sequences = [7, 33, 50, 36, 13, 14, 12, 30]

            for position, sequence in enumerate(sequences):
                row_idx = position + 3
                name = sheet.cell(row=row_idx, column=1).value
                montant_n = sheet.cell(row=row_idx, column=column_idx).value

                montant_n1 = None
                if col.upper() != 'E':
                    montant_n1 = sheet.cell(row=row_idx, column=column_idx + 1).value


                if montant_n is None:
                    montant_n = 0
                if montant_n1 is None:
                    montant_n1 = 0

                rubrique = self.env['import.ocr.config'].create({
                    'name': name,
                    'type': 'tcr',
                    'sequence': sequence
                })

                tcr_line_vals = {
                    'name': name,
                    'montant_n': montant_n,
                    'tcr_id': self.id,
                    'rubrique': rubrique.id
                }

                if montant_n1 is not None:
                    tcr_line_vals['montant_n1'] = montant_n1

                new_line = self.env['import.ocr.tcr.line'].create(tcr_line_vals)
                tcr_lines.append(new_line)

            self.tcr_lines = [(6, 0, [line.id for line in tcr_lines])]

        except Exception as e:
            print(f"Error processing file: {e}")

    def open_file(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.extract_bilan_wizard_form').id
            context = dict(self.env.context or {})
            context['pdf_1'] = rec.file_import
            context['pdf_2'] = rec.file_import2
            context['tcr_id'] = rec.id
            wizard = self.env['extract.bilan.wizard'].create({'pdf_1': rec.file_import,
                                                              'pdf_2': rec.file_import2,})
            return {
                'name': 'TCR',
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'extract.bilan.wizard',
                'res_id': wizard.id,
                'view_id': view_id,
                'target': 'new',
                'context': context,
            }
    def extract_data(self):
        for rec in self:
            if rec.file_import:
                pattern_alpha = r'^[a-zA-Z\séèàôâê\'-();,*+]+$'
                pattern_num = r'^[0-9\s]+$'
                credit = [1,2,3,4,5,6,7,8,9,10,11,30,33,34,40,41,44,45,49,50]
                debit = [12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,31,32,35,36,37,38,39,42,43,46,47,48]
                if rec.tcr_lines:
                    rec.tcr_lines.unlink()
                if rec.file_import:
                    data = str(rec.file_import)
                    data = data.replace("b'", '\n')
                    data = data.replace("'", '')
                    data = data.replace('\r\n', '\n')  # Replace Windows-style newline with Unix-style
                    data = data.replace('\r', '\n')
                    data = 'data:application/pdf;base64,' + data
                    test_file = ocr_space_file(filename=data, api_key='K87496787788957', language='fre',
                                               isTable=True)
                    json_dumps = test_file.content.decode()
                    json_loads = json.loads(json_dumps)
                    lines = json_loads['ParsedResults'][0]['TextOverlay']['Lines']

                    list_tcr = self.env['import.ocr.config'].search([('type', '=', 'tcr')])
                    lines = group_words_by_line(lines, list_tcr)
                    for line in lines:
                        rubrique = self.env['import.ocr.config'].search([('name', '=', line['Words'][0]['WordText'])])
                        value_text = line['Words'][0]['WordText']
                        if not rubrique:
                            print(value_text)
                            value_text = value_text.replace(';', '')
                            value_text = value_text.replace(',', '')
                            value_text = value_text.replace('-', '')
                            for i in list_tcr:
                                val = i.name
                                val = val.replace(';', '')
                                val = val.replace(',', '')
                                val = val.replace('-', '')
                                if value_text == val:
                                    print(value_text == val)
                        if rubrique:
                            value = rec.env['import.ocr.tcr.line'].create({'tcr_id': rec.id,
                                                                   'name': line['Words'][0]['WordText'],
                                                                   'rubrique': rubrique.id,
                                                                   })
                            if len(line['Words']) == 3:
                                try:
                                    value.write({'montant_n': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                except:
                                    value.write({'montant_n': 0})
                                try:
                                    value.write({'montant_n1': int(re.sub(r'[^0-9]', '', line['Words'][2]['WordText']))})
                                except:
                                    value.write({'montant_n1': 0})

                            elif len(line['Words']) == 2:
                                separator = line['Words'][1]['Left'] - line['Words'][0]['Left']
                                if separator > 400:
                                    try:
                                        value.write({'montant_n1': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                    except:
                                        value.write({'montant_n1': 0})
                                else:
                                    try:
                                        value.write({'montant_n': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                    except:
                                        value.write({'montant_n': 0})


                                '''dicty = {'min_top': line['MinTop'],
                                         'type': rubrique.sequence,
                                         'amounts': []}
                                same_line.append(dicty)'''

                if rec.file_import2:
                    data = str(rec.file_import2)
                    data = data.replace("b'", '\n')
                    data = data.replace("'", '')
                    data = data.replace('\r\n', '\n')  # Replace Windows-style newline with Unix-style
                    data = data.replace('\r', '\n')
                    data = 'data:application/pdf;base64,' + data
                    test_file = ocr_space_file(filename=data, api_key='K87496787788957', language='fre',
                                               isTable=True)
                    json_dumps = test_file.content.decode()
                    json_loads = json.loads(json_dumps)
                    lines = json_loads['ParsedResults'][0]['TextOverlay']['Lines']
                    list_tcr = self.env['import.ocr.config'].search([('type', '=', 'tcr')])
                    lines = group_words_by_line(lines, list_tcr)
                    for line in lines:
                        rubrique = self.env['import.ocr.config'].search([('name', '=', line['Words'][0]['WordText'])])
                        value_text = line['Words'][0]['WordText']
                        if not rubrique:
                            print(value_text)
                            value_text = value_text.replace(';', '')
                            value_text = value_text.replace(',', '')
                            value_text = value_text.replace('-', '')
                            for i in list_tcr:
                                val = i.name
                                val = val.replace(';', '')
                                val = val.replace(',', '')
                                val = val.replace('-', '')
                                if value_text == val:
                                    print(value_text == val)
                        if rubrique:
                            value = rec.env['import.ocr.tcr.line'].create({'tcr_id': rec.id,
                                                                   'name': line['Words'][0]['WordText'],
                                                                   'rubrique': rubrique.id,
                                                                   })
                            if len(line['Words']) == 3:
                                try:
                                    value.write({'montant_n': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                except:
                                    value.write({'montant_n': 0})
                                try:
                                    value.write(
                                        {'montant_n1': int(re.sub(r'[^0-9]', '', line['Words'][2]['WordText']))})
                                except:
                                    value.write({'montant_n1': 0})

                            elif len(line['Words']) == 2:
                                separator = line['Words'][1]['Left'] - line['Words'][0]['Left']
                                if separator > 400:
                                    try:
                                        value.write(
                                            {'montant_n1': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                    except:
                                        value.write({'montant_n1': 0})
                                else:
                                    try:
                                        value.write(
                                            {'montant_n': int(re.sub(r'[^0-9]', '', line['Words'][1]['WordText']))})
                                    except:
                                        value.write({'montant_n': 0})

                rec.state = "validation"
                for line in rec.tcr_lines:
                    line.montant_n = line.montant_n / 1000
                    line.montant_n1 = line.montant_n1 / 1000

    def action_validation(self):
        for rec in self:
            list_validation = [7, 12, 13, 14, 33, 50, 42]
            tcr = rec.tcr_lines.filtered(lambda r: r.rubrique.sequence in list_validation)
            if len(tcr) != 7:
                raise ValidationError("Vous devriez confirmer les valeurs suivantes: \n "
                                      "- Chiffre d'affaires net des rabais, Remises, Ristournes \n"
                                      "- Achats de marchandises vendues \n"
                                      "- Matières premieres \n"
                                      "- Autres approvisionnements \n"
                                      "- Excédent brut de l'exploitation \n"
                                      "- Charges financières \n"
                                      "- Résultat  net de l'exercice")
            view_id = self.env.ref('financial_modeling.confirmation_wizard_form')
            context = dict(self.env.context or {})
            context['tcr_id'] = rec.id
            context['state'] = 'valide'
            print(context)
            if not self._context.get('warning'):
                return {
                    'name': 'Validation',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'import.ocr.wizard',
                    'view_id': view_id.id,
                    'target': 'new',
                    'context': context,
                }

    def action_annulation(self):
        for rec in self:
            view_id = self.env.ref('financial_modeling.confirmation_wizard_form')
            context = dict(self.env.context or {})
            context['tcr_id'] = rec.id
            context['state'] = 'validation'
            print(context)
            if not self._context.get('warning'):
                return {
                    'name': 'Annulation',
                    'type': 'ir.actions.act_window',
                    'view_mode': 'form',
                    'res_model': 'import.ocr.wizard',
                    'view_id': view_id.id,
                    'target': 'new',
                    'context': context,
                }


class ImportTcrOcrLine(models.Model):
    _name = "import.ocr.tcr.line"
    _description = "Line de tcr importé"

    name = fields.Char(string="RUBRIQUES")
    sequence = fields.Integer(related='rubrique.sequence')
    mintop = fields.Integer(string='Rang')
    height = fields.Integer(string='Height')
    montant_n = fields.Float(string="N")
    montant_n1 = fields.Float(string="N-1")
    rubrique = fields.Many2one('import.ocr.config', string='Rubriques confirmés', domain="[('type','=','tcr')]")
    tcr_id = fields.Many2one('import.ocr.tcr', string="TCR ID")


class ConfigRubrique(models.Model):
    _name = 'import.ocr.config'
    _description = 'Liste des rubriques'

    name = fields.Char(string='rubrique')
    type = fields.Selection([('tcr', 'TCR'),
                             ('actif', 'Actif'),
                             ('passif', 'Passif')], string='Type')
    sequence = fields.Integer(string="Sequence")

def ocr_space_file(filename, overlay=True, api_key='helloworld', language='eng',isTable=False):
    payload = {'isOverlayRequired': overlay,
               'apikey': api_key,
               'language': language,
               'base64Image': filename,
               'isTable': isTable,
               'OCREngine': 2
               }
    print('called')
    r = requests.post('https://api.ocr.space/parse/image',
                          data=payload,
                          )
    return r

def assign_amounts(actif, amounts, intervals):
    for amount in amounts:
        if intervals[0][0] <= amount['left'] <= intervals[0][-1]:
                actif.montant_n = amount['amount']
        elif intervals[1][0] <= amount['left'] <= intervals[1][-1]:
            actif.montant_n1 = amount['amount']
        else:
            if intervals[0][0] <= amount['left'] - amount['left'] <= intervals[0][-1]:
                actif.montant_n = amount['amount']
            elif intervals[1][0] <= amount['left'] - amount['left'] <= intervals[1][-1]:
                actif.montant_n1 = amount['amount']


def image_to_base64(image_bytes):
    return base64.b64encode(image_bytes).decode('utf-8')


def pdf_page_to_base64(pdf_bytes, page_number):
    pdf_file = io.BytesIO(pdf_bytes)
    pdf_reader = PyPDF2.PdfFileReader(pdf_file)
    page = pdf_reader.getPage(page_number)
    page_content = page.extractText()  # Obtenez le contenu de la page si nécessaire
    # Convertir la page PDF en image (vous pouvez utiliser n'importe quelle bibliothèque pour cela)
    # Ici, nous utilisons PyMuPDF
    pdf_image = page.render()
    image = Image.open(io.BytesIO(pdf_image))
    # Convertir l'image en base64
    return image_to_base64(image.tobytes())


def group_words_by_line(json_data, list_tcr):
    # Trier les mots par leur position verticale (Top)
    sorted_words = sorted(json_data, key=lambda x: x['MinTop'])

    lines = []
    current_line = {'LineText': '', 'Left': None, 'Words': []}
    previous_top = None
    previous_max_height = None

    for item in sorted_words:
        top = item['MinTop']
        max_height = item['MaxHeight']
        word_text = item['LineText']
        left = item['Words'][0]['Left']  # Get the Left position of the first word in this line

        if previous_top is None:
            current_line['LineText'] = word_text
            current_line['Left'] = left
            current_line['Words'].append({'WordText': word_text, 'Left': left})
            previous_top = top
            previous_max_height = max_height
        elif top >= previous_top and top <= (previous_top + previous_max_height):
            current_line['LineText'] += " " + word_text
            current_line['Words'].append({'WordText': word_text, 'Left': left})
            previous_max_height = max(previous_max_height, max_height)
        else:
            # Trier les mots de la ligne par leur position horizontale (Left)
            current_line['Words'] = sorted(current_line['Words'], key=lambda x: x['Left'])
            lines.append(current_line)
            current_line = {'LineText': word_text, 'Left': left, 'Words': [{'WordText': word_text, 'Left': left}]}
            previous_top = top
            previous_max_height = max_height

    current_line['Words'] = sorted(current_line['Words'], key=lambda x: x['Left'])

    # Append the last line
    lines.append(current_line)
    return lines
