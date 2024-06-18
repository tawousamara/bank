
from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
import datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO
import xlsxwriter
import openpyxl

import numpy as np
import matplotlib.pyplot as plt

TCR_LIST = [
    ('1', "Chiffre d'affaire"),
    ('2', "Revente en l'état"),
    ('3', 'Production vendue'),
    ('4', 'Travaux'),
    ('5', 'Service'),
    ('6', 'Achats consommés'),
    ('7', "Autres charges externes"),
    ('8', "Valeur ajoutée d'exploitation"),
    ('9', 'Charges de personnel'),
    ('10', 'Impôts, taxes et versements assimilés'),
    ('11', "Excédent Brut d'Exploitation"),
    ('12', 'Autres produits opérationnels'),
    ('13', 'Autres charges opérationnelles'),
    ('14', 'Dotations aux amortissements'),
    ('15', 'Résultat Opérationnel'),
    ('16', 'Charges financières'),
    ('17', 'Résultat Ordinaire Avant Impôts'),
    ('18', "Impôts sur les bénéfices"),
    ('19', "Résultat Net"),
]
TCR_LIST_ar = [
    ('1', "Chiffre d'affaire", 'رقم الأعمال - المبيعات '),
    ('2', "Revente en l'état", 'إعادة البيع على الحالة'),
    ('3', 'Production vendue', 'الإنتاج المثبت '),
    ('4', 'Travaux', 'الاشغال'),
    ('5', 'Service', 'خدمات'),
    ('6', 'Achats consommés', 'مشتريات مستهلكة'),
    ('7', "Autres charges externes", 'خدمات خارجية ومشتريات أخرى'),
    ('8', "Valeur ajoutée d'exploitation", 'القيمة المضافة للاستغلال '),
    ('9', 'Charges de personnel', 'أعباء المستخدمين '),
    ('10', 'Impôts, taxes et versements assimilés', 'الضرائب والرسوم والمدفوعات المماثلة '),
    ('11', "Excédent Brut d'Exploitation", 'إجمالي فائض الاستغلال'),
    ('12', 'Autres produits opérationnels', 'المنتجات العملياتية الأخرى'),
    ('13', 'Autres charges opérationnelles', 'الأعباء العملياتية الأخرى'),
    ('14', 'Dotations aux amortissements', 'مخصصات الاستهلاك ،المؤونات وخسائر القيمة'),
    ('15', 'Résultat Opérationnel', 'النتيجة العملياتية'),
    ('16', 'Charges financières', 'الأعباء المالية'),
    ('17', 'Résultat Ordinaire Avant Impôts', 'النتيجة العادية قبل الضرائب'),
    ('18', "Impôts sur les bénéfices", 'الضرائب الواجب دفعها على النتائج العادية'),
    ('19', "Résultat Net", 'النتيجة الصافية للنشاطات العادية'),
]

Ratio_LIST = [
    ("1", "Marge brute"),
    ("2", "Marge brute %"),
    ("3", "EBE / CA %"),
    ("4", "RNC / CA %"),
    ("5", "CAF"),
    ("6", "CAF / CA %"),
    ("7", "FF / EBE %"),
]
Ratio_list_ar = [
    ('1', 'هامش الربح الإجمالي'),
    ('2', 'معدل هامش الربح %'),
    ('3', 'إجمالي فائض الاستغلال / رقم الأعمال %'),
    ('4', 'النتيجة الصافية / رقم الأعمال %'),
    ('5', 'القدرة على التمويل الذاتي'),
    ('6', 'قدرة التمويل الذاتي / رقم الأعمال %'),
    ('7', 'الأعباء المالية  / إجمالي فائض الاستغلال %'),
]

list_bilan = [
    ('1', 'تدفقات داخلة'),
    ('2', 'المبيعات'),
    ('3', 'تدفقات خارجة'),
    ('4', 'كلفة المبيعات'),
    ('5', 'المصاريف الإدارية والعمومية'),
    ('6', 'المصاريف التمويلية المرتبطة بالطلب'),
    ('7', 'Cash-flow  التدفقات النقدية'),
    ('8', 'صافي الربح'),
    ('9', 'الإهتلاكات و المؤونات'),
    ('10', 'CAF قدرة التمويل الذاتي'),
    ('11', 'الأقساط السنوية المرتبطة بالتمويلات الاستثمارية الحالية'),
    ('12', 'الأقساط السنوية المرتبطة بالطلب'),
    ('13', 'نسبة تغطية قدرة التمويل الذاتي للأقساط الإجمالية')
]

list_evaluation = [
    (1, 'المخزون'),
    (2, 'الزبائن'),
    (3, 'المـوردون'),
    (4, 'صافي راس المال العامل')
]

list_eval = [
    (1, 'إجمالي فائض الاستغلال'),
    (2, 'الضرائب الواجب دفعها على النتائج العادية'),
    (3, 'التغير في متطلبات رأس المال العامل'),
    (4, 'الأعباء المالية'),
    (5, 'التدفق النقدي الحر'),
]


class TCRAnalysis(models.Model):
    _name = 'tcr.analysis.import'

    name = fields.Char(string="Reference")
    date = fields.Date(string="Date")

    line_ids = fields.One2many('tcr.analysis.import.line', string='Lignes', inverse_name='tcr_analysis_id')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)
    ratio_ids = fields.One2many('tcr.analysis.ratio.line', string='Lignes', inverse_name='tcr_analysis_id', )
    tcr_prev_ids = fields.One2many('tcr.analysis.prev', string='Lignes', inverse_name='tcr_analysis_id', )
    recap_tcr_prev_ids = fields.One2many('tcr.analysis.prev', string='Lignes', inverse_name='tcr_analysis_id', )
    tcr_prev_ratio_ids = fields.One2many('tcr.analysis.prev.ratio', string='Lignes', inverse_name='tcr_analysis_id', )

    file_template = fields.Binary(string='Modèle Excel', compute='compute_template')
    file_template_name = fields.Char(string='fichier', default='Télécharger le modèle Excel')
    file_import_name = fields.Char(string='fichier', default='Importer le fichier Excel')
    file_import_data = fields.Binary(string='Importer le fichier Excel')
    file_tester = fields.Boolean(default=False)

    graph_historical_bar = fields.Binary(string='Graphique empilé')
    graph_historical_bar_emp = fields.Binary(string='Graphique empilé')
    graph_historical_pie_ca_by_exercise = fields.Binary(string='Graphique empilé')
    graph_historical_pie_by_exercise = fields.Binary(string='Graphique empilé')

    graph_prev_bar = fields.Binary(string='Graphique empilé')
    graph_prev_bar_emp = fields.Binary(string='Graphique empilé')
    graph_prev_pie_ca_by_exercise = fields.Binary(string='Graphique empilé', compute='compute_graph_suiv')
    graph_prev_pie_by_exercise = fields.Binary(string='Graphique empilé')

    year_prec = fields.Selection([('0', 'N+1'), ('1', 'N+2'), ('2', 'N+3'), ('3', 'N+4'), ('4', 'N+5')], string='Année')
    year_suiv = fields.Selection([('0', 'N+1'), ('1', 'N+2'), ('2', 'N+3'), ('3', 'N+4'), ('4', 'N+5')], string='Année')

    capital = fields.Float(string='CAPITAL BRUT' , default=0)
    capital_differe = fields.Float(string='DIFFERE CAPITALISE', readonly=True)
    taux = fields.Float(string='TAUX', default=0)
    tva = fields.Float(string='TVA', default=0)
    periodicite = fields.Selection([('m', 'MENSUEL'),
                                    ('a', 'ANNUEL'),
                                    ('s', 'SEMESTRIEL'),
                                    ('t', 'TRIMESTRIEL'),
                                    ], string='PERODICITE', default='m')
    nbr_echeance = fields.Integer(string='NOMBRE ECHEANCE', default=0)
    amort = fields.Float(string='ANNUITE', readonly=True, default=0)
    differe = fields.Selection([('oui', 'Oui'),
                                ('non', 'Non')], string='DIFFERE')
    duree_differe = fields.Integer(string='DUREE DIFFERE (en mois)', default=0)
    date_debut = fields.Date(string='DATE DEBUT')
    date_differe = fields.Date(string='DATE FIN DIFFERE')
    date_fin = fields.Date(string='DATE FIN', readonly=True)
    echeance_ids = fields.One2many('tcr.analysis.echeance.line', 'tcr_analysis_id', string='Echeances')

    cashflow_ids = fields.One2many('tcr.analysis.cashflow.line', 'tcr_analysis_id', string='Cash-flow')

    evaluation_ids = fields.One2many('tcr.analysis.evaluation.projet', 'tcr_analysis_id', string='Evaluations')
    evaluation_line_ids = fields.One2many('tcr.analysis.evaluation.projet.line', 'tcr_analysis_id', string='Evaluations')
    taux_rend = fields.Float(string='معدل العائد الداخلي')

    van = fields.Float(string='القيمة الحالية للتدفقات النقدية')

    @api.model
    def create(self, vals):
        vals['name'] = self.env['ir.sequence'].next_by_code('tcr.analysis.seq')
        res = super(TCRAnalysis, self).create(vals)
        etape = self.env.context.get('etape_id')
        if etape:
            res.etape_id = etape
            res.etape_id.invest_id = res.id
        return res

    def calcul_echeance_action(self):
        for rec in self:
            annuite = 0
            rec.capital_differe = rec.capital + (rec.capital * rec.duree_differe * rec.taux * (1 + rec.tva)) / 12 if rec.differe == 'oui' else rec.capital
            nbr_mois = periode = taux = 0
            if rec.periodicite == 'm':
                nbr_mois = rec.nbr_echeance - 1
                periode = 1
                taux = 12
                annuite = rec.capital_differe * ((rec.taux * (1 + rec.tva)) / taux) / (1 - (1 + ((rec.taux * (1 + rec.tva)) / taux)) ** -rec.nbr_echeance)
                valeur = ((rec.taux * (1 + rec.tva)) / taux) / (1 - (1 + ((rec.taux * (1 + rec.tva))/ taux) - rec.nbr_echeance))
                print(annuite)
                print(valeur)
            elif rec.periodicite == 't':
                nbr_mois = rec.nbr_echeance * 3 - 1
                periode = 3
                taux = 4
                annuite = rec.capital_differe * ((rec.taux * (1 + rec.tva)) / taux) / (1 - (1 + ((rec.taux * (1 + rec.tva)) / taux)) ** -rec.nbr_echeance)
            elif rec.periodicite == 's':
                nbr_mois = rec.nbr_echeance * 6 - 1
                periode = 6
                taux = 2
                annuite = rec.capital_differe * ((rec.taux * (1 + rec.tva)) / taux) / (1 - (1 + ((rec.taux * (1 + rec.tva)) / taux)) ** -rec.nbr_echeance)
            elif rec.periodicite == 'a':
                nbr_mois = rec.nbr_echeance * 12 - 1
                periode = 12
                taux = 1
                annuite = rec.capital_differe * (rec.taux * (1 + rec.tva)) / (1 - (1 + (rec.taux * (1 + rec.tva))) ** -rec.nbr_echeance)
            start_date = rec.date_debut
            rec.amort = annuite
            end_date = start_date + relativedelta(months=nbr_mois)
            rec.date_differe = end_date.strftime('%Y-%m-%d')
            rec.date_fin = end_date.strftime('%Y-%m-%d')
            capital = rec.capital
            rec.echeance_ids.unlink()
            for i in range(rec.nbr_echeance):
                marge = capital * (rec.taux / taux)
                vals = {
                    'tcr_analysis_id': rec.id,
                    'name': i+1,
                    'date': rec.date_debut + relativedelta(months=periode * i),
                    'capital': capital,
                    'total': rec.amort,
                    'principal': rec.amort - marge - (marge * rec.tva),
                    'marge': marge,
                    'tva': marge * rec.tva,
                }
                self.env['tcr.analysis.echeance.line'].create(vals)
                capital -= rec.amort

    def calcul_cashflow(self):
        for rec in self:

            bilan_1 = rec.cashflow_ids.filtered(lambda l: l.bilan == 1)
            bilan_2 = rec.cashflow_ids.filtered(lambda l: l.bilan == 2)
            recap_1 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '1')
            bilan_1.amount_n = bilan_2.amount_n = recap_1.amount_n
            bilan_1.amount_n1 = bilan_2.amount_n1 = recap_1.amount_n1
            bilan_1.amount_n2 = bilan_2.amount_n2 = recap_1.amount_n2
            bilan_1.amount_n3 = bilan_2.amount_n3 = recap_1.amount_n3
            bilan_1.amount_n4 = bilan_2.amount_n4 = recap_1.amount_n4
            bilan_1.amount_n5 = bilan_2.amount_n5 = recap_1.amount_n5

            bilan_3 = rec.cashflow_ids.filtered(lambda l: l.bilan == 3)
            bilan_4 = rec.cashflow_ids.filtered(lambda l: l.bilan == 4)
            bilan_5 = rec.cashflow_ids.filtered(lambda l: l.bilan == 5)
            bilan_6 = rec.cashflow_ids.filtered(lambda l: l.bilan == 6)
            bilan_15 = rec.cashflow_ids.filtered(lambda l: l.bilan == 12)
            recap_2 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '6')
            bilan_4.amount_n = recap_2.amount_n
            bilan_4.amount_n1 = recap_2.amount_n1
            bilan_4.amount_n2 = recap_2.amount_n2
            bilan_4.amount_n3 = recap_2.amount_n3
            bilan_4.amount_n4 = recap_2.amount_n4
            bilan_4.amount_n5 = recap_2.amount_n5
            recap_3 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '7')
            bilan_5.amount_n = recap_3.amount_n
            bilan_5.amount_n1 = recap_3.amount_n1
            bilan_5.amount_n2 = recap_3.amount_n2
            bilan_5.amount_n3 = recap_3.amount_n3
            bilan_5.amount_n4 = recap_3.amount_n4
            bilan_5.amount_n5 = recap_3.amount_n5

            somme = somme_annee = count = 0
            year = rec.echeance_ids[0].date.year
            for item in rec.echeance_ids:
                if year == item.date.year:
                    somme += item.marge
                    somme_annee += item.total
                else:
                    if count == 0:
                        bilan_6.amount_n = somme
                        bilan_15.amount_n = somme_annee
                    elif count == 1:
                        bilan_6.amount_n1 = somme
                        bilan_15.amount_n1 = somme_annee
                    elif count == 2:
                        bilan_6.amount_n2 = somme
                        bilan_15.amount_n2 = somme_annee
                    elif count == 3:
                        bilan_6.amount_n3 = somme
                        bilan_15.amount_n3 = somme_annee
                    elif count == 4:
                        bilan_6.amount_n4 = somme
                        bilan_15.amount_n4 = somme_annee
                    elif count == 5:
                        bilan_6.amount_n5 = somme
                        bilan_15.amount_n5 = somme_annee
                    somme = item.marge
                    somme_annee = item.total
                    count += 1
                year = item.date.year
            if count == 0:
                bilan_6.amount_n = somme
                bilan_15.amount_n = somme_annee
            elif count == 1:
                bilan_6.amount_n1 = somme
                bilan_15.amount_n1 = somme_annee
            elif count == 2:
                bilan_6.amount_n2 = somme
                bilan_15.amount_n2 = somme_annee
            elif count == 3:
                bilan_6.amount_n3 = somme
                bilan_15.amount_n3 = somme_annee
            elif count == 4:
                bilan_6.amount_n4 = somme
                bilan_15.amount_n4 = somme_annee
            elif count == 5:
                bilan_6.amount_n5 = somme
                bilan_15.amount_n5 = somme_annee
            bilan_3.amount_n = bilan_4.amount_n + bilan_5.amount_n + bilan_6.amount_n
            bilan_3.amount_n1 = bilan_4.amount_n1 + bilan_5.amount_n1 + bilan_6.amount_n1
            bilan_3.amount_n2 = bilan_4.amount_n2 + bilan_5.amount_n2 + bilan_6.amount_n2
            bilan_3.amount_n3 = bilan_4.amount_n3 + bilan_5.amount_n3 + bilan_6.amount_n3
            bilan_3.amount_n4 = bilan_4.amount_n4 + bilan_5.amount_n4 + bilan_6.amount_n4
            bilan_3.amount_n5 = bilan_4.amount_n5 + bilan_5.amount_n5 + bilan_6.amount_n5
            bilan_7 = rec.cashflow_ids.filtered(lambda l: l.bilan == 7)
            bilan_8 = rec.cashflow_ids.filtered(lambda l: l.bilan == 8)
            bilan_9 = rec.cashflow_ids.filtered(lambda l: l.bilan == 9)
            bilan_10 = rec.cashflow_ids.filtered(lambda l: l.bilan == 10)
            bilan_11 = rec.cashflow_ids.filtered(lambda l: l.bilan == 11)
            bilan_13 = rec.cashflow_ids.filtered(lambda l: l.bilan == 13)

            bilan_7.amount_n = bilan_1.amount_n - bilan_3.amount_n
            bilan_7.amount_n1 = bilan_1.amount_n1 - bilan_3.amount_n1
            bilan_7.amount_n2 = bilan_1.amount_n2 - bilan_3.amount_n2
            bilan_7.amount_n3 = bilan_1.amount_n3 - bilan_3.amount_n3
            bilan_7.amount_n4 = bilan_1.amount_n4 - bilan_3.amount_n4
            bilan_7.amount_n5 = bilan_1.amount_n5 - bilan_3.amount_n5

            recap_4 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '19')
            bilan_8.amount_n = recap_4.amount_n
            bilan_8.amount_n1 = recap_4.amount_n1
            bilan_8.amount_n2 = recap_4.amount_n2
            bilan_8.amount_n3 = recap_4.amount_n3
            bilan_8.amount_n4 = recap_4.amount_n4
            bilan_8.amount_n5 = recap_4.amount_n5
            recap_5 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '14')
            bilan_9.amount_n = recap_5.amount_n
            bilan_9.amount_n1 = recap_5.amount_n1
            bilan_9.amount_n2 = recap_5.amount_n2
            bilan_9.amount_n3 = recap_5.amount_n3
            bilan_9.amount_n4 = recap_5.amount_n4
            bilan_9.amount_n5 = recap_5.amount_n5
            recap_6 = rec.tcr_prev_ratio_ids.filtered(lambda l: l.ratio == '5')
            bilan_10.amount_n = recap_6.amount_n
            bilan_10.amount_n1 = recap_6.amount_n1
            bilan_10.amount_n2 = recap_6.amount_n2
            bilan_10.amount_n3 = recap_6.amount_n3
            bilan_10.amount_n4 = recap_6.amount_n4
            bilan_10.amount_n5 = recap_6.amount_n5
            bilan_13.amount_n = (bilan_10.amount_n / (bilan_11.amount_n + bilan_15.amount_n))  if bilan_11.amount_n + bilan_15.amount_n != 0 else 0
            bilan_13.amount_n1 = (bilan_10.amount_n1 / (bilan_11.amount_n1 + bilan_15.amount_n1)) if bilan_11.amount_n1 + bilan_15.amount_n1 != 0 else 0
            bilan_13.amount_n2 = (bilan_10.amount_n2 / (bilan_11.amount_n2 + bilan_15.amount_n2)) if bilan_11.amount_n2 + bilan_15.amount_n2 != 0 else 0
            bilan_13.amount_n3 = (bilan_10.amount_n3 / (bilan_11.amount_n3 + bilan_15.amount_n3)) if bilan_11.amount_n3 + bilan_15.amount_n3 != 0 else 0
            bilan_13.amount_n4 = (bilan_10.amount_n4 / (bilan_11.amount_n4 + bilan_15.amount_n4)) if bilan_11.amount_n4 + bilan_15.amount_n4 != 0 else 0
            bilan_13.amount_n5 = (bilan_10.amount_n5 / (bilan_11.amount_n5 + bilan_15.amount_n5)) if bilan_11.amount_n5 + bilan_15.amount_n5 != 0 else 0

    def calcul_evaluation(self):
        for rec in self:
            bilan_1 = rec.evaluation_line_ids.filtered(lambda l: l.sequence == 1)
            recap_1 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '11')
            bilan_1.amount_n1 = recap_1.amount_n
            bilan_1.amount_n2 = recap_1.amount_n1
            bilan_1.amount_n3 = recap_1.amount_n2
            bilan_1.amount_n4 = recap_1.amount_n3
            bilan_1.amount_n5 = recap_1.amount_n4

            bilan_2 = rec.evaluation_line_ids.filtered(lambda l: l.sequence == 2)
            cash_2 = rec.cashflow_ids.filtered(lambda l: l.bilan == 6)
            recap_2 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '18')
            bilan_2.amount_n1 = recap_2.amount_n
            bilan_2.amount_n2 = recap_2.amount_n1
            bilan_2.amount_n3 = recap_2.amount_n2
            bilan_2.amount_n4 = recap_2.amount_n3
            bilan_2.amount_n5 = recap_2.amount_n4

            bilan_3 = rec.evaluation_line_ids.filtered(lambda l: l.sequence == 3)
            bilan_init_1 = rec.evaluation_ids.filtered(lambda l: l.sequence == 4)

            bilan_3.amount_n1 = bilan_init_1.amount_n - bilan_init_1.amount_n1
            bilan_3.amount_n2 = bilan_init_1.amount_n1 - bilan_init_1.amount_n2
            bilan_3.amount_n3 = bilan_init_1.amount_n2 - bilan_init_1.amount_n3
            bilan_3.amount_n4 = bilan_init_1.amount_n3 - bilan_init_1.amount_n4
            bilan_3.amount_n5 = bilan_init_1.amount_n4 - bilan_init_1.amount_n5

            bilan_4 = rec.evaluation_line_ids.filtered(lambda l: l.sequence == 4)
            recap_3 = rec.recap_tcr_prev_ids.filtered(lambda l: l.poste_comptable == '16')

            bilan_4.amount_n1 = recap_3.amount_n + cash_2.amount_n
            bilan_4.amount_n2 = recap_3.amount_n1 + cash_2.amount_n1
            bilan_4.amount_n3 = recap_3.amount_n2 + cash_2.amount_n2
            bilan_4.amount_n4 = recap_3.amount_n3 + cash_2.amount_n3
            bilan_4.amount_n5 = recap_3.amount_n4 + cash_2.amount_n4
            bilan_5 = rec.evaluation_line_ids.filtered(lambda l: l.sequence == 5)

            bilan_5.amount_n1 = bilan_1.amount_n1 - bilan_2.amount_n1 + bilan_3.amount_n1 - bilan_4.amount_n1
            bilan_5.amount_n2 = bilan_1.amount_n2 - bilan_2.amount_n2 + bilan_3.amount_n2 - bilan_4.amount_n2
            bilan_5.amount_n3 = bilan_1.amount_n3 - bilan_2.amount_n3 + bilan_3.amount_n3 - bilan_4.amount_n3
            bilan_5.amount_n4 = bilan_1.amount_n4 - bilan_2.amount_n4 + bilan_3.amount_n4 - bilan_4.amount_n4
            bilan_5.amount_n5 = bilan_1.amount_n5 - bilan_2.amount_n5 + bilan_3.amount_n5 - bilan_4.amount_n5

            cashflows = [bilan_5.amount_n1, bilan_5.amount_n2, bilan_5.amount_n3, bilan_5.amount_n4, bilan_5.amount_n5]
            van = calculate_npv(cashflows, rec.taux_rend)
            rec.van = van
    def action_set_data(self):
        print('set')
        self.ensure_one()
        self.line_ids.unlink()
        for item in TCR_LIST_ar:
            new_line = self.env['tcr.analysis.import.line'].create({
                'poste_comptable': item[0],
                'poste_arabe': item[2],
                'amount_n4': 0.00,
                'amount_n3': 0.00,
                'amount_n2': 0.00,
                'amount_n1': 0.00,
                'amount_n': 0.00,
                'tcr_analysis_id': self.id
            })

    def action_validate(self):
        self.ensure_one()
        self.file_tester = True
        self.tcr_prev_ids.unlink()
        for rec in self.line_ids:
            self.env['tcr.analysis.prev'].create({
                'tcr_analysis_id': self.id,
                'amount_n': rec.amount_n,
                'amount_n1': rec.amount_n1,
                'amount_n2': rec.amount_n2,
                'amount_n3': rec.amount_n3,
                'amount_n4': rec.amount_n4,
                'amount_n5': 0.00,
                'poste_comptable': rec.poste_comptable,
                'poste_arabe': rec.poste_arabe
            })

        print(get_data(self.line_ids, type_class=2))
        self.graph_historical_bar = create_stacked_chart(get_data(self.line_ids, type_class=2), type_class=2)
        self.graph_historical_bar_emp = create_bar(get_data(self.line_ids, type_class=2), type_class=2)

    def compute_template(self):
        data_tcr = TCR_LIST
        result_excel = BytesIO()

        # Create Excel workbook and worksheet
        workbook = xlsxwriter.Workbook(result_excel)
        worksheet = workbook.add_worksheet("modèle")
        bold = workbook.add_format({'bold': True})

        worksheet.write(0, 0, 'Poste Comptable', bold)
        worksheet.write(0, 1, 'N+1', bold)
        worksheet.write(0, 2, 'N+2', bold)
        worksheet.write(0, 3, 'N+3', bold)
        worksheet.write(0, 4, 'N+4', bold)
        worksheet.write(0, 5, 'N+5', bold)

        for index, entry in data_tcr:
            bold_list = [1, 7, 18, 19, 22, 27, 30, 32]
            if int(index) in bold_list:
                worksheet.write(int(index), 0, entry, bold)
            else:
                worksheet.write(int(index), 0, entry)

        workbook.close()

        buf = base64.b64encode(result_excel.getvalue())
        self.write({'file_template': buf})

        result_excel.close()

        for rec in self:
            if rec.etape_id:
                rec.etape_id.invest_id = rec.id

    def export_one2many_to_excel(self):
        # Assuming you have a model with a one2many field named 'line_ids'
        # Replace 'your.model' with the actual model name
        records = self.env['your.model'].search([])  # Retrieve all records

        # Create a new Excel workbook
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('Sheet1')

        # Write header row
        sheet.write(0, 0, 'Parent Field 1')
        sheet.write(0, 1, 'Parent Field 2')
        sheet.write(0, 2, 'Child Field 1')
        sheet.write(0, 3, 'Child Field 2')

        row = 1
        for record in records:
            # Write parent record data
            sheet.write(row, 0, record.parent_field1)
            sheet.write(row, 1, record.parent_field2)

            # Write child records data
            for line in record.line_ids:
                sheet.write(row, 2, line.child_field1)
                sheet.write(row, 3, line.child_field2)
                row += 1

        # Save the workbook
        workbook.save('/path/to/your/exported_file.xls')

    def action_import_data(self):
        self.ensure_one()
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file_import_data)), read_only=True)
        ws = wb.active
        search = self.env['tcr.analysis.import.line'].search([('tcr_analysis_id', '=', self.id)]).unlink()
        count = 0
        for record in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None, values_only=True):
            count += 1
            if count <= 19:
                print(count)
                print(TCR_LIST[count - 1])
                print(record)
                try:
                    self.env['tcr.analysis.import.line'].create({
                        'tcr_analysis_id': self.id,
                        'poste_comptable': TCR_LIST[count-1][0],
                        'poste_arabe': TCR_LIST_ar[count-1][2],
                        'amount_n4': record[5],
                        'amount_n3': record[4],
                        'amount_n2': record[3],
                        'amount_n1': record[2],
                        'amount_n': record[1],
                    })
                except:
                    self.env['tcr.analysis.import.line'].create({
                        'tcr_analysis_id': self.id,
                        'poste_comptable': TCR_LIST[count-1][0],
                        'poste_arabe': TCR_LIST_ar[count-1][2],
                        'amount_n4': 0,
                        'amount_n3': 0,
                        'amount_n2': 0,
                        'amount_n1': 0,
                        'amount_n': 0,
                    })
        calculFormule(self.line_ids)

    def action_count_ratio_hist(self):
        self.ensure_one()
        self.ratio_ids.unlink()
        if self.line_ids:
            for index, entry in Ratio_LIST:
                list_amounts = get_amount(self.line_ids, self.ratio_ids, index)
                ratio = self.env['tcr.analysis.ratio.line'].create({
                    'tcr_analysis_id': self.id,
                    'ratio': index,
                    'amount_n': list_amounts[0],
                    'amount_n1': list_amounts[1],
                    'amount_n2': list_amounts[2],
                    'amount_n3': list_amounts[3],
                    'amount_n4': list_amounts[4]})

    def action_count_prev(self):
        self.ensure_one()
        if self.tcr_prev_ids:
            for rec in self.tcr_prev_ids:
                if rec.poste_comptable not in ('1', '8', '11', '15', '17', '19'):
                    line = rec.line_ids.filtered(lambda l: l.poste_comptable == rec.poste_comptable)
                    rec.amount_n = line.amount_n
                    rec.amount_n1 = line.amount_n1
                    rec.amount_n2 = line.amount_n2
                    rec.amount_n3 = line.amount_n3
                    rec.amount_n4 = line.amount_n4
                    rec.amount_n5 = line.amount_n5
                    rec.amount_n = rec.amount_n * (rec.augment_hypothesis_n1 / 100 + 1)
                    rec.amount_n1 = rec.amount_n1 * (rec.augment_hypothesis_n2 / 100 + 1)
                    rec.amount_n2 = rec.amount_n2 * (rec.augment_hypothesis_n3 / 100 + 1)
                    rec.amount_n3 = rec.amount_n3 * (rec.augment_hypothesis_n4 / 100 + 1)
                    rec.amount_n4 = rec.amount_n4 * (rec.augment_hypothesis_n5 / 100 + 1)
                    rec.amount_n5 = rec.amount_n5 * (rec.augment_hypothesis_n5 / 100 + 1)
            calculFormule(self.tcr_prev_ids)
            self.tcr_prev_ratio_ids.unlink()
            for index, entry in Ratio_LIST:
                list_amounts = get_amount(self.tcr_prev_ids, self.tcr_prev_ratio_ids, index, is_prev=1)
                ratio = self.env['tcr.analysis.prev.ratio'].create({
                    'tcr_analysis_id': self.id,
                    'ratio': index,
                    'amount_n': list_amounts[0],
                    'amount_n1': list_amounts[1],
                    'amount_n2': list_amounts[2],
                    'amount_n3': list_amounts[3],
                    'amount_n4': list_amounts[4],
                    'amount_n5': list_amounts[5]})
            self.recap_tcr_prev_ids = self.tcr_prev_ids
            self.graph_prev_bar = create_stacked_chart(get_data(self.tcr_prev_ids))
            self.graph_prev_bar_emp = create_bar(get_data(self.tcr_prev_ids))
            self.cashflow_ids.unlink()
            for index, entry in list_bilan:
                print(type(index))
                print(entry)
                bilan = self.env['tcr.analysis.cashflow.line'].create({
                    'tcr_analysis_id': self.id,
                    'taux_change': self.taux_change,
                    'bilan': int(index),
                    'name': entry,
                    'amount_n': 0,
                    'amount_n1': 0,
                    'amount_n2': 0,
                    'amount_n3': 0,
                    'amount_n4': 0,
                    'amount_n5': 0})
            self.evaluation_ids.unlink()
            for index, entry in list_evaluation:
                bilan = self.env['tcr.analysis.evaluation.projet'].create({
                    'tcr_analysis_id': self.id,
                    'sequence': int(index),
                    'name': entry,
                    'amount_n': 0,
                    'amount_n1': 0,
                    'amount_n2': 0,
                    'amount_n3': 0,
                    'amount_n4': 0,
                    'amount_n5': 0})
            self.evaluation_line_ids.unlink()
            for index, entry in list_eval:
                bilan = self.env['tcr.analysis.evaluation.projet.line'].create({
                    'tcr_analysis_id': self.id,
                    'sequence': int(index),
                    'name': entry,
                    'amount_n1': 0,
                    'amount_n2': 0,
                    'amount_n3': 0,
                    'amount_n4': 0,
                    'amount_n5': 0})
        else:
            raise ValidationError("Vous devriez d'abord valider les données")

    @api.onchange('year_prec')
    def compute_graph_prec(self):
        if self.year_prec:
            records_ca = self.line_ids.filtered(lambda r: r.poste_comptable in ['2', '3', '4', '5'])
            labels = []
            sizes = []
            data = get_data(records_ca, type_class=2)
            for lab in data:
                if list(lab.values())[int(self.year_prec) + 1] != 0:
                    labels.append(list(lab.values())[0])
                    sizes.append(list(lab.values())[int(self.year_prec) + 1])

            img = create_pie(sizes, labels=labels)
            self.write({'graph_historical_pie_ca_by_exercise': img})

    @api.depends('year_suiv')
    def compute_graph_suiv(self):
        if self.year_suiv and self.recap_tcr_prev_ids:
            records_ca = self.recap_tcr_prev_ids.filtered(lambda r: r.poste_comptable in ['2', '3', '4', '5'])
            labels = []
            sizes = []
            data = get_data(records_ca, )
            print(len(data))
            for lab in data:
                print(int(self.year_suiv) + 1)
                print(len(list(lab.values())))
                if list(lab.values())[int(self.year_suiv) + 1] != 0:
                    labels.append(list(lab.values())[0])
                    sizes.append(list(lab.values())[int(self.year_suiv) + 1])
            print(labels)
            print(sizes)
            img = create_pie(sizes, labels=labels)
            self.write({'graph_prev_pie_ca_by_exercise': img})
        else:
            self.graph_prev_pie_by_exercise = None
            self.graph_prev_pie_ca_by_exercise = None


class TCRprev(models.Model):
    _name = 'tcr.analysis.prev'
    _description = "BFR Forecast"

    poste_comptable = fields.Selection(TCR_LIST, string='Poste Comptable')
    poste_arabe = fields.Char(string='Poste Comptable')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import')

    amount_n = fields.Float(string="N+1")
    amount_n1 = fields.Float(string="N+2")
    amount_n2 = fields.Float(string="N+3")
    amount_n3 = fields.Float(string="N+4")
    amount_n4 = fields.Float(string="N+5")
    amount_n5 = fields.Float(string="N+5")

    augment_hypothesis_n1 = fields.Float(string="Hypothèse croissance N+1", digits=(16, 2))
    augment_hypothesis_n2 = fields.Float(string="Hypothèse croissance N+2", digits=(16, 2))
    augment_hypothesis_n3 = fields.Float(string="Hypothèse croissance N+3", digits=(16, 2))
    augment_hypothesis_n4 = fields.Float(string="Hypothèse croissance N+4", digits=(16, 2))
    augment_hypothesis_n5 = fields.Float(string="Hypothèse croissance N+5", digits=(16, 2))

    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)

    active = fields.Boolean(string="Active", default=True)


class ImportTCRLine(models.Model):
    _name = 'tcr.analysis.import.line'

    poste_comptable = fields.Selection(TCR_LIST, string='Poste Comptable')
    poste_arabe = fields.Char(string='Poste Comptable')
    amount_n = fields.Float(string='N+1')
    amount_n1 = fields.Float(string='N+2')
    amount_n2 = fields.Float(string='N+3')
    amount_n3 = fields.Float(string='N+4')
    amount_n4 = fields.Float(string='N+5')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)


class RatioTCRLine(models.Model):
    _name = 'tcr.analysis.ratio.line'

    ratio = fields.Selection(Ratio_LIST, string='Ratio', readonly=True)
    amount_n = fields.Float(string='N+1')
    amount_n1 = fields.Float(string='N+2')
    amount_n2 = fields.Float(string='N+3')
    amount_n3 = fields.Float(string='N+4')
    amount_n4 = fields.Float(string='N+5')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)
    graph = fields.Binary(string='', compute='compute_graph')

    def compute_graph(self):
        for rec in self:
            data = [rec.amount_n3, rec.amount_n2, rec.amount_n1, rec.amount_n]

            fig, ax = plt.subplots(1, 1, figsize=(1.5, 0.25))
            ax.plot(data)
            for k, v in ax.spines.items():
                v.set_visible(False)
            ax.set_xticks([])
            ax.set_yticks([])

            plt.plot(len(data) - 1, data[len(data) - 1], )

            buf = BytesIO()
            plt.savefig(buf, format='jpeg', dpi=100)
            buf.seek(0)
            imageBase64 = base64.b64encode(buf.getvalue())
            buf.close()
            rec.graph = imageBase64


class RatioTCRLinePrev(models.Model):
    _name = 'tcr.analysis.prev.ratio'

    ratio = fields.Selection(Ratio_LIST, string='Ratio', readonly=True)
    ratio_ar = fields.Char(string='ratio', compute='compute_ratio')
    amount_n = fields.Float(string='N+1')
    amount_n1 = fields.Float(string='N+2')
    amount_n2 = fields.Float(string='N+3')
    amount_n3 = fields.Float(string='N+4')
    amount_n4 = fields.Float(string='N+5')
    amount_n5 = fields.Float(string='N+5')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')
    company_id = fields.Many2one('res.company', readonly=True, default=lambda self: self.env.company)
    graph = fields.Binary(string='', compute='compute_graph')

    def compute_ratio(self):
        for rec in self:
            rec.ratio_ar = Ratio_list_ar[int(rec.ratio)-1][1]

    def compute_graph(self):
        for rec in self:
            data = [rec.amount_n, rec.amount_n1, rec.amount_n2, rec.amount_n3, rec.amount_n4, rec.amount_n5]

            fig, ax = plt.subplots(1, 1, figsize=(1.5, 0.25))
            ax.plot(data)
            for k, v in ax.spines.items():
                v.set_visible(False)
            ax.set_xticks([])
            ax.set_yticks([])

            plt.plot(len(data) - 1, data[len(data) - 1], )

            buf = BytesIO()
            plt.savefig(buf, format='jpeg', dpi=100)
            buf.seek(0)
            imageBase64 = base64.b64encode(buf.getvalue())
            buf.close()
            rec.graph = imageBase64


class Evaluation(models.Model):
    _name = 'tcr.analysis.evaluation.projet.line'

    name = fields.Char(string=' ')
    sequence = fields.Integer(string='Sequence')
    amount_n1 = fields.Float(string='N+1')
    amount_n2 = fields.Float(string='N+2')
    amount_n3 = fields.Float(string='N+3')
    amount_n4 = fields.Float(string='N+4')
    amount_n5 = fields.Float(string='N+5')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')

class Evaluation(models.Model):
    _name = 'tcr.analysis.evaluation.projet'

    name = fields.Char(string=' ')
    sequence = fields.Integer(string='Sequence')
    amount_n = fields.Float(string='N')
    amount_n1 = fields.Float(string='N+1')
    amount_n2 = fields.Float(string='N+2')
    amount_n3 = fields.Float(string='N+3')
    amount_n4 = fields.Float(string='N+4')
    amount_n5 = fields.Float(string='N+5')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')
    computed = fields.Boolean(string='', compute='compute_montant')

    def compute_montant(self):
        for rec in self:
            if rec.sequence == 4:
                bilan_1 = self.env['tcr.analysis.evaluation.projet'].search([('tcr_analysis_id', '=', rec.tcr_analysis_id.id),
                                                                             ('sequence', '=', 1)])
                bilan_2 = self.env['tcr.analysis.evaluation.projet'].search([('tcr_analysis_id', '=', rec.tcr_analysis_id.id),
                                                                             ('sequence', '=', 2)])
                bilan_3 = self.env['tcr.analysis.evaluation.projet'].search([('tcr_analysis_id', '=', rec.tcr_analysis_id.id),
                                                                             ('sequence', '=', 3)])
                rec.amount_n = bilan_1.amount_n + bilan_2.amount_n - bilan_3.amount_n
                rec.amount_n1 = bilan_1.amount_n1 + bilan_2.amount_n1 - bilan_3.amount_n1
                rec.amount_n2 = bilan_1.amount_n2 + bilan_2.amount_n2 - bilan_3.amount_n2
                rec.amount_n3 = bilan_1.amount_n3 + bilan_2.amount_n3 - bilan_3.amount_n3
                rec.amount_n4 = bilan_1.amount_n4 + bilan_2.amount_n4 - bilan_3.amount_n4
                rec.amount_n5 = bilan_1.amount_n5 + bilan_2.amount_n5 - bilan_3.amount_n5
                rec.computed = True
            else:
                rec.computed = False

class Echeance(models.Model):
    _name = 'tcr.analysis.echeance.line'
    _description = 'Lignes des echeances'

    name = fields.Integer(string='NUM ECH')
    date = fields.Date(string='DATE')
    capital = fields.Float(string='CAPITAL')
    principal = fields.Float(string='PRINCIPAL')
    marge = fields.Float(string='MARGE')
    tva = fields.Float(string='TVA')
    total = fields.Float(string='TOTAL', compute='compute_total')
    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')
    taux_change = fields.Float(string='Taux de change')

    def compute_total(self):
        for rec in self:
            rec.total = rec.principal + rec.marge + rec.tva

class Cashflow(models.Model):
    _name = 'tcr.analysis.cashflow.line'
    _description = 'Lignes des cash-flow'

    name = fields.Char(string=' ')
    bilan = fields.Integer(string='Bilan')

    amount_n = fields.Float(string='N+1 م/دج')
    amount_n_dollar = fields.Float(string='م/$', compute='compute_dollar')

    amount_n1 = fields.Float(string='N+2 م/دج')
    amount_n1_dollar = fields.Float(string='م/$', compute='compute_dollar')

    amount_n2 = fields.Float(string='N+3 م/دج')
    amount_n2_dollar = fields.Float(string='م/$', compute='compute_dollar')

    amount_n3 = fields.Float(string='N+4 م/دج')
    amount_n3_dollar = fields.Float(string='م/$', compute='compute_dollar')

    amount_n4 = fields.Float(string='N+5 م/دج')
    amount_n4_dollar = fields.Float(string='م/$', compute='compute_dollar')

    amount_n5 = fields.Float(string='N+5 م/دج')
    amount_n5_dollar = fields.Float(string='م/$', compute='compute_dollar')

    tcr_analysis_id = fields.Many2one('tcr.analysis.import', string='tcr')
    taux_change = fields.Float(string='1$ = ?DA: سعر الصرف')

    def compute_dollar(self):
        for rec in self:
            rec.amount_n_dollar = rec.taux_change * rec.amount_n
            rec.amount_n1_dollar = rec.taux_change * rec.amount_n1
            rec.amount_n2_dollar = rec.taux_change * rec.amount_n2
            rec.amount_n3_dollar = rec.taux_change * rec.amount_n3
            rec.amount_n4_dollar = rec.taux_change * rec.amount_n4
            rec.amount_n5_dollar = rec.taux_change * rec.amount_n5


def get_value(value):
    data_get = ''
    for index, entry in TCR_LIST:
        if entry == value:
            data_get = index
    return data_get


def calculate_npv(cashflows, discount_rate):
    npv = 0.0
    for t, cf in enumerate(cashflows):
        print(t)
        print(cf)
        print(discount_rate)
        npv += cf / ((1 + discount_rate) ** (t + 1))
    return npv

def create_stacked_chart(data, type_class=1):
    data_tmp_1 = list(data[0].values())[1:]
    data_tmp_2 = list(data[10].values())[1:]
    data_tmp_3 = list(data[18].values())[1:]
    data1 = data_tmp_1
    data2 = data_tmp_2
    data3 = data_tmp_3
    year = ["N+1", "N+2", "N+3", "N+4", "N+5"]
    x = np.arange(len(year))  # the label locations
    width = 0.25  # the width of the bars

    fig, ax = plt.subplots()
    rects1 = ax.bar(x, data1, width, color="blue", label="Chiffre d'affaire")
    rects2 = ax.bar(x + width, data2, width, color="orange", label="EBE")
    rects3 = ax.bar(x + width * 2, data3, width, color="grey", label="Résultat Net")

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_ylabel('Montant')
    ax.set_title('Montant par année')
    ax.set_xticks(x + width, year)
    ax.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))

    ax.bar_label(rects1, padding=3)
    ax.bar_label(rects2, padding=3)
    ax.bar_label(rects3, padding=3)

    fig.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='jpeg', dpi=100)
    buf.seek(0)
    imageBase64 = base64.b64encode(buf.getvalue())
    buf.close()
    return imageBase64


def create_bar(data, type_class=1):
    data_tmp_1 = list(data[0].values())[1:]
    data_tmp_2 = list(data[5].values())[1:]
    data1 = data_tmp_1
    data2 = data_tmp_2
    year = ["N+1", "N+2", "N+3", "N+4", "N+5"]
    fig, ax = plt.subplots()
    width = 0.5
    rects1 = ax.bar(year, data1, width, color="green", label="Chiffre d'affaire")
    rects2 = ax.bar(year, data2, width, color="yellow", bottom=np.array(data1), label="Achats consommés")

    ax.legend(loc="lower left", bbox_to_anchor=(0.8, 1.0))
    fig.tight_layout()
    buf = BytesIO()
    plt.savefig(buf, format='jpeg', dpi=100)
    buf.seek(0)
    imageBase64 = base64.b64encode(buf.getvalue())
    buf.close()
    return imageBase64


def create_pie(data, labels):
    sizes = data
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    buf = BytesIO()
    plt.savefig(buf, format='jpeg', dpi=100)
    buf.seek(0)
    imageBase64 = base64.b64encode(buf.getvalue())
    buf.close()
    return imageBase64


def get_amount(line_ids, ratio_ids, index, is_prev=0):
    chiffre_daffaire = line_ids.filtered(lambda r: r.poste_comptable == '1')
    amount_n = amount_n1 = amount_n2 = amount_n3 = amount_n4 = amount_n5 = 0.00
    if index == "1":
        achat_consomme = line_ids.filtered(lambda r: r.poste_comptable == '6')
        amount_n = chiffre_daffaire.amount_n - achat_consomme.amount_n
        amount_n1 = chiffre_daffaire.amount_n1 - achat_consomme.amount_n1
        amount_n2 = chiffre_daffaire.amount_n2 - achat_consomme.amount_n2
        amount_n3 = chiffre_daffaire.amount_n3 - achat_consomme.amount_n3
        if is_prev != 0:
            amount_n4 = chiffre_daffaire.amount_n4 - achat_consomme.amount_n4
            amount_n5 = chiffre_daffaire.amount_n5 - achat_consomme.amount_n5
    elif index == "2":
        marge = ratio_ids.filtered(lambda r: r.ratio == '1')
        amount_n = (marge.amount_n / chiffre_daffaire.amount_n) * 100 if chiffre_daffaire.amount_n > 0 else 0
        amount_n1 = (marge.amount_n1 / chiffre_daffaire.amount_n1) * 100 if chiffre_daffaire.amount_n1 > 0 else 0
        amount_n2 = (marge.amount_n2 / chiffre_daffaire.amount_n2) * 100 if chiffre_daffaire.amount_n2 > 0 else 0
        amount_n3 = (marge.amount_n3 / chiffre_daffaire.amount_n3) * 100 if chiffre_daffaire.amount_n3 > 0 else 0
        if is_prev != 0:
            amount_n4 = (marge.amount_n4 / chiffre_daffaire.amount_n4) * 100 if chiffre_daffaire.amount_n4 > 0 else 0
            amount_n5 = (marge.amount_n5 / chiffre_daffaire.amount_n5) * 100 if chiffre_daffaire.amount_n5 > 0 else 0
    elif index == "3":
        EBE = line_ids.filtered(lambda r: r.poste_comptable == '11')
        amount_n = (EBE.amount_n / chiffre_daffaire.amount_n) * 100 if chiffre_daffaire.amount_n > 0 else 0
        amount_n1 = (EBE.amount_n1 / chiffre_daffaire.amount_n1) * 100 if chiffre_daffaire.amount_n1 > 0 else 0
        amount_n2 = (EBE.amount_n2 / chiffre_daffaire.amount_n2) * 100 if chiffre_daffaire.amount_n2 > 0 else 0
        amount_n3 = (EBE.amount_n3 / chiffre_daffaire.amount_n3) * 100 if chiffre_daffaire.amount_n3 > 0 else 0
        if is_prev != 0:
            amount_n4 = (EBE.amount_n4 / chiffre_daffaire.amount_n4) * 100 if chiffre_daffaire.amount_n4 > 0 else 0
            amount_n5 = (EBE.amount_n5 / chiffre_daffaire.amount_n5) * 100 if chiffre_daffaire.amount_n5 > 0 else 0
    elif index == "4":
        RNC = line_ids.filtered(lambda r: r.poste_comptable == '19')
        amount_n = (RNC.amount_n / chiffre_daffaire.amount_n) * 100 if chiffre_daffaire.amount_n > 0 else 0
        amount_n1 = (RNC.amount_n1 / chiffre_daffaire.amount_n1) * 100 if chiffre_daffaire.amount_n1 > 0 else 0
        amount_n2 = (RNC.amount_n2 / chiffre_daffaire.amount_n2) * 100 if chiffre_daffaire.amount_n2 > 0 else 0
        amount_n3 = (RNC.amount_n3 / chiffre_daffaire.amount_n3) * 100 if chiffre_daffaire.amount_n3 > 0 else 0
        if is_prev != 0:
            amount_n4 = (RNC.amount_n4 / chiffre_daffaire.amount_n4) * 100 if chiffre_daffaire.amount_n4 > 0 else 0
            amount_n5 = (RNC.amount_n5 / chiffre_daffaire.amount_n5) * 100 if chiffre_daffaire.amount_n5 > 0 else 0
    elif index == "5":
        RNC = line_ids.filtered(lambda r: r.poste_comptable == '19')
        DA = line_ids.filtered(lambda r: r.poste_comptable == '14')
        reprise = line_ids.filtered(lambda r: r.poste_comptable == '26')
        amount_n = RNC.amount_n + DA.amount_n - reprise.amount_n
        amount_n1 = RNC.amount_n1 + DA.amount_n1 - reprise.amount_n1
        amount_n2 = RNC.amount_n2 + DA.amount_n2 - reprise.amount_n2
        amount_n3 = RNC.amount_n3 + DA.amount_n3 - reprise.amount_n3
        if is_prev != 0:
            amount_n4 = RNC.amount_n4 + DA.amount_n4 - reprise.amount_n4
            amount_n5 = RNC.amount_n5 + DA.amount_n5 - reprise.amount_n5
    elif index == "6":
        caf = ratio_ids.filtered(lambda r: r.ratio == '5')
        amount_n = (caf.amount_n / chiffre_daffaire.amount_n) * 100 if chiffre_daffaire.amount_n > 0 else 0
        amount_n1 = (caf.amount_n1 / chiffre_daffaire.amount_n1) * 100 if chiffre_daffaire.amount_n1 > 0 else 0
        amount_n2 = (caf.amount_n2 / chiffre_daffaire.amount_n2) * 100 if chiffre_daffaire.amount_n2 > 0 else 0
        amount_n3 = (caf.amount_n3 / chiffre_daffaire.amount_n3) * 100 if chiffre_daffaire.amount_n3 > 0 else 0
        if is_prev != 0:
            amount_n4 = (caf.amount_n4 / chiffre_daffaire.amount_n4) * 100 if chiffre_daffaire.amount_n4 > 0 else 0
            amount_n5 = (caf.amount_n5 / chiffre_daffaire.amount_n5) * 100 if chiffre_daffaire.amount_n5 > 0 else 0
    elif index == "7":
        charge_fin = line_ids.filtered(lambda r: r.poste_comptable == '16')
        EBE = line_ids.filtered(lambda r: r.poste_comptable == '11')
        amount_n = (charge_fin.amount_n / EBE.amount_n) * 100 if EBE.amount_n > 0 else 0
        amount_n1 = (charge_fin.amount_n1 / EBE.amount_n1) * 100 if EBE.amount_n1 > 0 else 0
        amount_n2 = (charge_fin.amount_n2 / EBE.amount_n2) * 100 if EBE.amount_n2 > 0 else 0
        amount_n3 = (charge_fin.amount_n3 / EBE.amount_n3) * 100 if EBE.amount_n3 > 0 else 0
        if is_prev != 0:
            amount_n4 = (charge_fin.amount_n4 / EBE.amount_n4) * 100 if EBE.amount_n4 > 0 else 0
            amount_n5 = (charge_fin.amount_n5 / EBE.amount_n5) * 100 if EBE.amount_n5 > 0 else 0
    return [amount_n, amount_n1, amount_n2, amount_n3, amount_n4, amount_n5]


def get_data(data, type_class=1):
    recordset = []
    if type_class != 1:
        for i in data:
            element = {
                'poste_comptable': TCR_LIST[int(i.poste_comptable) - 1][1],
                'amount_n': i.amount_n,
                'amount_n1': i.amount_n1,
                'amount_n2': i.amount_n2,
                'amount_n3': i.amount_n3,
                'amount_n4': i.amount_n4,
            }
            recordset.append(element)
    else:
        for i in data:
            element = {
                'poste_comptable': TCR_LIST[int(i.poste_comptable) - 1][1],
                'amount_n1': i.amount_n1,
                'amount_n2': i.amount_n2,
                'amount_n3': i.amount_n3,
                'amount_n4': i.amount_n4,
                'amount_n5': i.amount_n5
            }
            recordset.append(element)
    return recordset


def calculFormule(line_ids):
    calcul = line_ids.filtered(lambda r: r.poste_comptable == '1')
    valeurs = line_ids.filtered(lambda r: r.poste_comptable in ['2', '3', '4', '5'])
    amount_n = amount_n1 = amount_n2 = amount_n3 = amount_n4 = amount_n5 = 0.00
    calcul.amount_n = sum(valeurs.mapped('amount_n'))
    calcul.amount_n1 = sum(valeurs.mapped('amount_n1'))
    calcul.amount_n2 = sum(valeurs.mapped('amount_n2'))
    calcul.amount_n3 = sum(valeurs.mapped('amount_n3'))
    calcul.amount_n4 = sum(valeurs.mapped('amount_n4'))
    calcul1 = line_ids.filtered(lambda r: r.poste_comptable == '8')
    valeurs = line_ids.filtered(lambda r: r.poste_comptable in ['6', '7'])
    calcul1.amount_n = calcul.amount_n - sum(valeurs.mapped('amount_n'))
    calcul1.amount_n1 = calcul.amount_n1 - sum(valeurs.mapped('amount_n1'))
    calcul1.amount_n2 = calcul.amount_n2 - sum(valeurs.mapped('amount_n2'))
    calcul1.amount_n3 = calcul.amount_n3 - sum(valeurs.mapped('amount_n3'))
    calcul1.amount_n4 = calcul.amount_n4 - sum(valeurs.mapped('amount_n4'))
    calcul2 = line_ids.filtered(lambda r: r.poste_comptable == '11')
    valeurs = line_ids.filtered(lambda r: r.poste_comptable in ['9', '10'])
    calcul2.amount_n = calcul1.amount_n - sum(valeurs.mapped('amount_n'))
    calcul2.amount_n1 = calcul1.amount_n1 - sum(valeurs.mapped('amount_n1'))
    calcul2.amount_n2 = calcul1.amount_n2 - sum(valeurs.mapped('amount_n2'))
    calcul2.amount_n3 = calcul1.amount_n3 - sum(valeurs.mapped('amount_n3'))
    calcul2.amount_n4 = calcul1.amount_n4 - sum(valeurs.mapped('amount_n4'))
    calcul3 = line_ids.filtered(lambda r: r.poste_comptable == '15')
    valeur1 = line_ids.filtered(lambda r: r.poste_comptable in ['12'])
    valeurs = line_ids.filtered(lambda r: r.poste_comptable in ['13', '14'])
    calcul3.amount_n = calcul2.amount_n + valeur1.amount_n - sum(valeurs.mapped('amount_n'))
    calcul3.amount_n1 = calcul2.amount_n1 + valeur1.amount_n1 - sum(valeurs.mapped('amount_n1'))
    calcul3.amount_n2 = calcul2.amount_n2 + valeur1.amount_n2 - sum(valeurs.mapped('amount_n2'))
    calcul3.amount_n3 = calcul2.amount_n3 + valeur1.amount_n3 - sum(valeurs.mapped('amount_n3'))
    calcul3.amount_n4 = calcul2.amount_n4 + valeur1.amount_n4 - sum(valeurs.mapped('amount_n4'))
    calcul = line_ids.filtered(lambda r: r.poste_comptable == '17')
    valeurs = line_ids.filtered(lambda r: r.poste_comptable in ['16'])
    calcul.amount_n = calcul3.amount_n - sum(valeurs.mapped('amount_n'))
    calcul.amount_n1 = calcul3.amount_n1 - sum(valeurs.mapped('amount_n1'))
    calcul.amount_n2 = calcul3.amount_n2 - sum(valeurs.mapped('amount_n2'))
    calcul.amount_n3 = calcul3.amount_n3 - sum(valeurs.mapped('amount_n3'))
    calcul.amount_n4 = calcul3.amount_n4 - sum(valeurs.mapped('amount_n4'))
    calcul1 = line_ids.filtered(lambda r: r.poste_comptable == '19')
    valeurs = line_ids.filtered(lambda r: r.poste_comptable in ['18'])
    calcul1.amount_n = calcul.amount_n - sum(valeurs.mapped('amount_n'))
    calcul1.amount_n1 = calcul.amount_n1 - sum(valeurs.mapped('amount_n1'))
    calcul1.amount_n2 = calcul.amount_n2 - sum(valeurs.mapped('amount_n2'))
    calcul1.amount_n3 = calcul.amount_n3 - sum(valeurs.mapped('amount_n3'))
    calcul1.amount_n4 = calcul.amount_n4 - sum(valeurs.mapped('amount_n4'))
